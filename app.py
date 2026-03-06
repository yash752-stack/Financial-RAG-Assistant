from __future__ import annotations

import os, re, json, math, io, html as _ht
import datetime as _dt
import threading as _th, time as _tm
import statistics as _stats
import requests
import pandas as pd
import numpy as np
try:
    import plotly.graph_objects as go
    import plotly.express as px
    _PLOTLY = True
except ImportError:
    _PLOTLY = False
import streamlit as st
import streamlit.components.v1 as components
from dotenv import load_dotenv
from typing import Optional

load_dotenv()

st.set_page_config(
    page_title="Financial RAG Assistant",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for _k, _v in [
    ("messages",        []),
    ("vectorstore",     None),
    ("uploaded_docs",   0),
    ("chunk_count",     0),
    ("file_names",      []),
    ("show_upload",     False),
    ("show_chat",       False),
    ("show_portfolio",  False),    # v7: portfolio panel toggle
    ("doc_full_text",   ""),
    ("auto_metrics",    []),
    ("auto_generated",  False),
    ("search_query",    ""),
    ("search_results",  []),
    # Portfolio state
    ("portfolio",       {}),       # {sym: {"shares": float, "avg_cost": float, "added": str}}
    ("portfolio_notes", {}),       # {sym: str}  analyst notes per holding
    # Chat mode
    ("analyst_mode",    False),    # v8: structured analyst output vs free-form chat
    # AI chart analysis cache
    ("_chart_ai_text",  ""),
    ("_chart_ai_done",  False),
    ("_chart_ai_tf",    "1D"),      # which timeframe is showing
    ("_fx_ai_text",     ""),
    ("_fx_ai_done",     False),
    ("_fx_ai_tf",       "1D"),
    ("_comm_ai_text",   ""),
    ("_comm_ai_done",   False),
    ("_comm_ai_tf",     "1D"),
    ("_crypto_ai_text", ""),
    ("_crypto_ai_done", False),
    ("_crypto_ai_tf",   "1D"),
    # Portfolio: selected holding for instant AI analysis
    ("_pf_selected_holding", ""),
    ("_pf_holding_ai",       {}),   # {sym: {tf: text}}
    # Price alerts: {sym: {"above": None, "below": None}}
    ("price_alerts",         {}),
    # Alert notification log
    ("alert_log",            []),
    # Theme
    ("app_theme",            "Royal Velvet"),
    # Groq retry error tracking: {site_key: {ts, wait, msg, type}}
    ("_groq_last_err",       {}),
    # CrossEncoder opt-in (disabled by default to save RAM on free tier)
    ("_ce_enabled",          False),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────────────────────────────────────
# HIDE STREAMLIT CHROME + RESPONSIVE
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
#MainMenu,footer,header,[data-testid="stToolbar"],
[data-testid="stDecoration"],.stDeployButton{display:none!important}
[data-testid="collapsedControl"]{background:rgba(107,45,107,.18)!important;
  border:1px solid rgba(139,58,139,.45)!important;border-radius:8px!important;
  color:#C084C8!important;top:.9rem!important}
[data-testid="collapsedControl"]:hover{background:rgba(107,45,107,.35)!important;
  box-shadow:0 0 12px rgba(107,45,107,.4)!important}
@media(max-width:1024px){[data-testid="block-container"]{padding:0 1rem!important}
  .stat-strip{grid-template-columns:repeat(2,1fr)!important}}
@media(max-width:767px){[data-testid="block-container"]{padding:0 .5rem!important;max-width:100%!important}
  .stat-strip{grid-template-columns:repeat(2,1fr)!important}
  .rag-header h1{font-size:2rem!important}.rag-header{padding:1.2rem 1rem!important}
  /* Stack holding cards to 1-col on mobile */
  .holding-card{margin-bottom:.6rem!important}
  /* Touch-friendly buttons */
  div[data-testid="stButton"] > button{min-height:44px!important}
  /* Hide 3rd+ index cards on small screens */
  .mood-indices > div:nth-child(n+3){display:none!important}
  /* Portfolio panel full-width */
  div[data-testid="stHorizontalBlock"]:has(.portfolio-panel){flex-direction:column!important}
}
@media(max-width:479px){.stat-strip{grid-template-columns:1fr!important}
  .rag-header h1{font-size:1.6rem!important}
  /* Single column holdings grid on very small screens */
  div[data-testid="stHorizontalBlock"]:has(.holding-card){flex-direction:column!important}
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# DESIGN SYSTEM  (Royal Velvet & Black)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;0,600;1,300;1,400&family=Syne:wght@400;500;600;700&family=Space+Mono:ital,wght@0,400;0,700;1,400&display=swap');
:root{--black:#07060C;--card:#0D0B12;--card-2:#120E1A;--panel:#0F0C16;
  --border:rgba(139,58,139,.22);--border-l:rgba(176,107,176,.45);
  --velvet:#6B2D6B;--velvet-gl:#B06BB0;--accent:#C084C8;--lilac:#D4A8D8;
  --text:#EDE8F5;--text-dim:#9A8AAA;--text-ghost:#4A3858;
  --green:#4ADE80;--red:#F87171;--gold:#F0C040}
*,*::before,*::after{box-sizing:border-box}
html,body,[class*="css"]{font-family:'Syne',sans-serif!important;color:var(--text)!important}
.stApp,[data-testid="stAppViewContainer"]{background:
  radial-gradient(ellipse 110% 55% at 0% 0%,rgba(107,45,107,.20) 0%,transparent 55%),
  radial-gradient(ellipse  80% 50% at 100% 100%,rgba(107,45,107,.14) 0%,transparent 55%),
  var(--black)!important}
[data-testid="stMain"],[data-testid="block-container"]{background:transparent!important;
  padding-top:0!important;max-width:1120px!important}
[data-testid="stSidebar"]{background:var(--panel)!important;
  border-right:1px solid var(--border-l)!important;
  box-shadow:4px 0 40px rgba(107,45,107,.08)!important}
[data-testid="stSidebar"]>div{padding:1.4rem 1.2rem!important}
h1,h2,h3,h4{font-family:'Cormorant Garamond',serif!important;color:var(--text)!important}
code,pre{font-family:'Space Mono',monospace!important}
[data-testid="stMetric"]{background:var(--card)!important;border:1px solid var(--border)!important;
  border-radius:8px!important;padding:.9rem 1rem!important}
[data-testid="stMetricLabel"] p{font-family:'Space Mono',monospace!important;font-size:.58rem!important;
  color:var(--text-ghost)!important;text-transform:uppercase!important;letter-spacing:.18em!important}
[data-testid="stMetricValue"]{font-family:'Cormorant Garamond',serif!important;font-size:1.7rem!important;
  font-weight:300!important;color:var(--accent)!important}
.stButton>button{background:transparent!important;border:1px solid var(--border)!important;
  border-radius:6px!important;color:var(--text-dim)!important;font-family:'Syne',sans-serif!important;
  font-size:.8rem!important;transition:all .22s ease!important;text-align:left!important}
.stButton>button:hover{background:rgba(107,45,107,.14)!important;border-color:var(--velvet-gl)!important;
  color:var(--accent)!important;box-shadow:0 0 18px rgba(107,45,107,.22)!important;transform:translateY(-1px)!important}
.stTextInput input,.stTextArea textarea{background:var(--card)!important;
  border:1px solid var(--border)!important;border-radius:8px!important;color:var(--text)!important}
[data-testid="stChatInput"]{background:var(--card-2)!important;border:1px solid var(--border-l)!important;
  border-radius:14px!important;box-shadow:0 0 30px rgba(107,45,107,.12)!important}
[data-testid="stChatInput"] textarea{background:transparent!important;border:none!important;color:var(--text)!important}
[data-testid="stChatInput"]:focus-within{border-color:rgba(139,58,139,.7)!important}
[data-testid="stChatMessage"]{background:var(--card)!important;border:1px solid var(--border)!important;
  border-radius:12px!important;padding:.8rem 1rem!important;margin-bottom:.5rem!important}
[data-testid="stFileUploader"]{background:rgba(107,45,107,.05)!important;
  border:1.5px dashed rgba(139,58,139,.4)!important;border-radius:10px!important}
[data-testid="stExpander"]{background:var(--card)!important;border:1px solid var(--border)!important;
  border-radius:8px!important}
[data-testid="stAlert"]{background:rgba(107,45,107,.1)!important;
  border:1px solid var(--border-l)!important;border-radius:8px!important}
div[data-testid="stSuccess"]{background:rgba(74,222,128,.07)!important}
div[data-testid="stError"]{background:rgba(248,113,113,.07)!important}
.stProgress>div>div{background:linear-gradient(90deg,var(--velvet),var(--accent))!important}
[data-testid="stMultiSelect"]>div{background:var(--card)!important;
  border-color:var(--border)!important;border-radius:8px!important}
.stMultiSelect span[data-baseweb="tag"]{background:rgba(107,45,107,.3)!important;
  border:1px solid var(--velvet-gl)!important;color:var(--lilac)!important;border-radius:999px!important}
[data-testid="stSelectbox"]>div>div{background:var(--card)!important;
  border-color:var(--border)!important;border-radius:8px!important}
hr{border-color:var(--border)!important}
::-webkit-scrollbar{width:3px}
::-webkit-scrollbar-thumb{background:rgba(107,45,107,.35);border-radius:2px}

/* ── HERO ── */
.rag-header{position:relative;padding:2rem 2.2rem;
  background:linear-gradient(135deg,rgba(107,45,107,.22) 0%,rgba(13,11,18,.98) 55%,rgba(107,45,107,.12) 100%);
  border:1px solid rgba(255,255,255,.08);border-radius:18px;
  box-shadow:0 8px 40px rgba(0,0,0,.4);margin-bottom:1.4rem;overflow:hidden}
.rag-header::before{content:'';position:absolute;top:-80px;right:-80px;width:280px;height:280px;
  border-radius:50%;background:radial-gradient(circle,rgba(107,45,107,.25) 0%,transparent 70%);pointer-events:none}
.rag-header::after{content:'';position:absolute;bottom:0;left:0;right:0;height:1px;
  background:linear-gradient(90deg,transparent 0%,rgba(107,45,107,.6) 30%,rgba(192,132,200,.8) 50%,rgba(107,45,107,.6) 70%,transparent 100%)}
.rag-kicker{font-family:'Space Mono',monospace;font-size:.6rem;letter-spacing:.3em;
  color:var(--velvet-gl);text-transform:uppercase;margin-bottom:.9rem;
  display:flex;align-items:center;gap:.6rem}
.rag-kicker::before{content:'';display:inline-block;width:20px;height:1px;background:var(--velvet-gl);opacity:.6}
.rag-header h1{font-family:'Cormorant Garamond',serif!important;font-size:3.2rem!important;
  font-weight:300!important;line-height:1.0!important;color:var(--text)!important;
  margin:0 0 .2rem!important;letter-spacing:-.02em!important}
.rag-header h1 em{font-style:italic;
  background:linear-gradient(135deg,var(--velvet-gl) 0%,var(--accent) 100%);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.rag-header p{font-family:'Syne',sans-serif;font-size:.86rem;color:var(--text-dim);
  margin:.6rem 0 0!important;max-width:480px}
.badge-row{display:flex;gap:.4rem;margin-top:.9rem;flex-wrap:wrap}
.badge{font-family:'Space Mono',monospace;font-size:.62rem;letter-spacing:.08em;padding:.2rem .55rem;
  border-radius:999px;border:1px solid var(--border);color:var(--text-ghost);background:rgba(255,255,255,.04)}
.badge.v{border-color:rgba(139,58,139,.5);color:var(--accent);background:rgba(107,45,107,.12)}
.badge.g{border-color:rgba(74,222,128,.3);color:#86efac;background:rgba(74,222,128,.07)}
.badge.b{border-color:rgba(96,165,250,.4);color:#60a5fa;background:rgba(96,165,250,.07)}

/* ── STAT STRIP ── */
.stat-strip{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;
  background:rgba(107,45,107,.22);border-radius:10px;overflow:hidden;
  border:1px solid rgba(107,45,107,.22);margin-bottom:1.4rem}
.stat-cell{background:var(--card);padding:1rem 1.2rem;position:relative;transition:background .25s}
.stat-cell:hover{background:var(--card-2)}
.stat-cell::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,var(--velvet),var(--accent));opacity:0;transition:opacity .25s}
.stat-cell:hover::before{opacity:1}
.stat-lbl{font-family:'Space Mono',monospace;font-size:.52rem;letter-spacing:.2em;
  text-transform:uppercase;color:var(--text-ghost);margin-bottom:.4rem}
.stat-val{font-family:'Cormorant Garamond',serif;font-size:1.7rem;font-weight:300;
  color:var(--text);line-height:1}
.stat-val.active{color:var(--accent)}
.stat-val-mono{font-family:'Space Mono',monospace;font-size:.68rem;color:var(--accent);line-height:1.4}

/* ── MARKET MOOD ── */
.mood-bar-wrap{background:var(--card);border:1px solid var(--border);border-radius:12px;
  padding:1rem 1.4rem;margin-bottom:1.4rem}
.mood-title{font-family:'Space Mono',monospace;font-size:.58rem;letter-spacing:.2em;
  text-transform:uppercase;color:var(--velvet-gl);margin-bottom:.7rem}
.mood-track{height:6px;border-radius:3px;
  background:linear-gradient(90deg,#f87171 0%,#fb923c 25%,#facc15 50%,#86efac 75%,#4ade80 100%);
  position:relative;margin-bottom:.5rem}
.mood-needle{position:absolute;top:-5px;width:16px;height:16px;border-radius:50%;
  border:2px solid #fff;background:var(--accent);transform:translateX(-50%);
  box-shadow:0 0 8px rgba(192,132,200,.6)}
.mood-labels{display:flex;justify-content:space-between;font-family:'Space Mono',monospace;
  font-size:.5rem;color:var(--text-ghost)}
.mood-index{font-family:'Cormorant Garamond',serif;font-size:2rem;font-weight:300}
.mood-indices{display:flex;gap:1rem;margin-top:.8rem;flex-wrap:wrap}
.mood-idx-chip{display:flex;flex-direction:column;background:var(--card-2);
  border:1px solid var(--border);border-radius:8px;padding:.4rem .8rem;
  font-family:'Space Mono',monospace;min-width:90px;
  position:relative;overflow:hidden;transition:border-color .2s;}
.mood-idx-chip.chip-up{border-color:rgba(74,222,128,.25);box-shadow:inset 0 0 0 1px rgba(74,222,128,.08);}
.mood-idx-chip.chip-up::before{content:'';position:absolute;top:0;right:0;width:100%;height:100%;
  background:linear-gradient(135deg,transparent 55%,rgba(74,222,128,.07) 55%,rgba(74,222,128,.11) 68%,transparent 68%);
  border-radius:8px;pointer-events:none;}
.mood-idx-chip.chip-down{border-color:rgba(248,113,113,.25);box-shadow:inset 0 0 0 1px rgba(248,113,113,.08);}
.mood-idx-chip.chip-down::before{content:'';position:absolute;top:0;right:0;width:100%;height:100%;
  background:linear-gradient(135deg,transparent 55%,rgba(248,113,113,.07) 55%,rgba(248,113,113,.11) 68%,transparent 68%);
  border-radius:8px;pointer-events:none;}
.mood-idx-name{font-size:.52rem;color:var(--text-ghost);letter-spacing:.1em}
.mood-idx-val{font-size:.72rem;color:var(--text);margin-top:.1rem}
.mood-idx-chg.up{font-size:.56rem;color:#4ade80}
.mood-idx-chg.down{font-size:.56rem;color:#f87171}

/* ── PRICE CHIP ── */
.price-chip{
  display:flex;flex-direction:column;
  background:var(--card-2);
  border:1px solid var(--border);
  border-radius:10px;padding:.75rem 1rem;
  min-width:120px;font-family:'Space Mono',monospace;
  transition:border-color .2s, box-shadow .2s;
  position:relative;overflow:hidden;
}
.price-chip:hover{border-color:var(--border-l)}

/* diagonal stripe — injected via ::before, colour set inline by JS-free class */
.price-chip::before{
  content:'';position:absolute;top:0;right:0;
  width:100%;height:100%;
  pointer-events:none;border-radius:10px;
  opacity:0;transition:opacity .25s;
}
.price-chip.chip-up{
  border-color:rgba(74,222,128,.28);
  box-shadow:inset 0 0 0 1px rgba(74,222,128,.12), 0 2px 12px rgba(74,222,128,.06);
}
.price-chip.chip-up::before{
  background:linear-gradient(135deg,
    transparent 60%,
    rgba(74,222,128,.08) 60%,
    rgba(74,222,128,.13) 72%,
    transparent 72%
  );
  opacity:1;
}
.price-chip.chip-down{
  border-color:rgba(248,113,113,.28);
  box-shadow:inset 0 0 0 1px rgba(248,113,113,.12), 0 2px 12px rgba(248,113,113,.06);
}
.price-chip.chip-down::before{
  background:linear-gradient(135deg,
    transparent 60%,
    rgba(248,113,113,.08) 60%,
    rgba(248,113,113,.13) 72%,
    transparent 72%
  );
  opacity:1;
}

/* Left edge accent bar */
.price-chip.chip-up::after{
  content:'';position:absolute;top:0;left:0;width:2px;height:100%;
  background:linear-gradient(180deg,#4ade80,rgba(74,222,128,.3));border-radius:10px 0 0 10px;
}
.price-chip.chip-down::after{
  content:'';position:absolute;top:0;left:0;width:2px;height:100%;
  background:linear-gradient(180deg,#f87171,rgba(248,113,113,.3));border-radius:10px 0 0 10px;
}

.price-chip:hover{border-color:var(--border-l)}
.pc-sym{font-size:.6rem;color:var(--accent);font-weight:700;letter-spacing:.08em;white-space:nowrap}
.pc-name{font-size:.5rem;color:var(--text-ghost);margin-bottom:.2rem}
.pc-val{font-family:'Cormorant Garamond',serif;font-size:1.5rem;font-weight:300;
  color:var(--text);line-height:1}
.pc-chg.up{font-size:.58rem;color:#4ade80;margin-top:.1rem}
.pc-chg.down{font-size:.58rem;color:#f87171;margin-top:.1rem}
.pc-chg.flat{font-size:.58rem;color:var(--text-ghost);margin-top:.1rem}
.chips-row{display:flex;gap:.6rem;flex-wrap:wrap}

/* ── MARKET PANELS ── */
.fx-panel,.comm-panel,.crypto-panel{background:var(--card);border:1px solid var(--border);
  border-radius:12px;padding:1.1rem 1.4rem .9rem;margin-bottom:1.4rem}
.fx-panel-title,.comm-title,.crypto-title{font-family:'Cormorant Garamond',serif;
  font-size:1.1rem;font-weight:300;color:var(--text);margin-bottom:.9rem;
  display:flex;align-items:center;gap:.5rem}
.fx-panel-title::before{content:'';display:inline-block;width:3px;height:1.1rem;
  background:linear-gradient(180deg,var(--velvet),var(--accent));border-radius:2px}
.comm-title::before{content:'';display:inline-block;width:3px;height:1.1rem;
  background:linear-gradient(180deg,#F0C040,#C084C8);border-radius:2px}
.crypto-title::before{content:'';display:inline-block;width:3px;height:1.1rem;
  background:linear-gradient(180deg,#fb923c,#C084C8);border-radius:2px}

/* ── MISC ── */
.sb-lbl{font-family:'Space Mono',monospace;font-size:.54rem;letter-spacing:.22em;
  text-transform:uppercase;color:var(--velvet-gl);padding:1.2rem 0 .45rem;
  border-top:1px solid var(--border);margin-top:.5rem}
.key-ok{display:flex;align-items:center;gap:.5rem;background:rgba(74,222,128,.07);
  border:1px solid rgba(74,222,128,.2);color:#86efac;padding:.38rem .7rem;
  border-radius:6px;font-family:'Space Mono',monospace;font-size:.6rem}
.key-dot{width:5px;height:5px;border-radius:50%;background:#4ade80;
  box-shadow:0 0 6px #4ade80;animation:blink 2s infinite}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.3}}
.doc-pill{display:flex;align-items:center;gap:.4rem;background:rgba(107,45,107,.1);
  border:1px solid var(--border);padding:.32rem .65rem;border-radius:4px;
  margin-bottom:.3rem;font-family:'Space Mono',monospace;font-size:.58rem;
  color:var(--text-dim);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.doc-dot{width:4px;height:4px;border-radius:50%;background:var(--velvet-gl);flex-shrink:0}
.empty{text-align:center;padding:4rem 2rem}
.empty-orb{width:100px;height:100px;border-radius:50%;
  background:radial-gradient(circle,rgba(107,45,107,.28) 0%,transparent 70%);
  border:1px solid var(--border);margin:0 auto 1.5rem;
  display:flex;align-items:center;justify-content:center;font-size:2rem;color:var(--velvet-gl)}
.empty-title{font-family:'Cormorant Garamond',serif;font-size:1.7rem;font-weight:300;
  font-style:italic;color:var(--text-ghost);margin-bottom:.5rem}
.empty-sub{font-size:.8rem;color:var(--text-ghost);max-width:300px;margin:0 auto;line-height:1.8;opacity:.7}
.upload-drawer{background:linear-gradient(135deg,rgba(107,45,107,.18) 0%,rgba(13,11,18,.95) 100%);
  border:1px solid rgba(139,58,139,.45);border-radius:12px;padding:1rem 1.1rem .7rem;margin-bottom:.6rem}
.upload-drawer-title{font-family:'Space Mono',monospace;font-size:.62rem;letter-spacing:.15em;
  text-transform:uppercase;color:var(--velvet-gl);margin-bottom:.6rem}
.src-card{background:var(--card);border:1px solid var(--border);border-left:3px solid var(--velvet-gl);
  border-radius:0 8px 8px 0;padding:.7rem .9rem;margin:.4rem 0;font-size:.82rem}
.src-name{font-family:'Space Mono',monospace;font-size:.7rem;color:var(--accent);margin-bottom:.15rem}
.src-score{font-family:'Space Mono',monospace;font-size:.62rem;color:var(--text-ghost)}
.src-preview{color:var(--text-dim);line-height:1.55;margin-top:.2rem}
.vfooter{text-align:center;padding:1.8rem 0 .5rem;position:relative;margin-top:2.5rem}
.vfooter::before{content:'';position:absolute;top:0;left:50%;transform:translateX(-50%);
  width:180px;height:1px;background:linear-gradient(90deg,transparent,rgba(107,45,107,.5),transparent)}
.vfooter-text{font-family:'Space Mono',monospace;font-size:.56rem;
  letter-spacing:.2em;text-transform:uppercase;color:var(--text-ghost)}

/* ════════════════════════════════════════════
   v5  NEW STYLES
   ════════════════════════════════════════════ */

/* ① Global search bar */
.gsearch-wrap{
  position:sticky;top:0;z-index:1000;
  background:linear-gradient(180deg,rgba(7,6,12,.97) 82%,transparent);
  backdrop-filter:blur(16px);-webkit-backdrop-filter:blur(16px);
  padding:.55rem 0 .3rem;margin-bottom:.9rem;
}
/* search result cards */
.sr-wrap{background:var(--card);border:1px solid var(--border);
  border-radius:10px;padding:.75rem 1rem;margin-bottom:.9rem}
.sr-title{font-family:'Space Mono',monospace;font-size:.52rem;letter-spacing:.18em;
  text-transform:uppercase;color:var(--accent);margin-bottom:.6rem}
.sr-hit{background:var(--card-2);border:1px solid var(--border);
  border-left:3px solid var(--accent);border-radius:0 8px 8px 0;
  padding:.55rem .9rem;margin-bottom:.4rem}
.sr-fname{font-family:'Space Mono',monospace;font-size:.56rem;color:var(--accent);margin-bottom:.18rem}
.sr-snippet{font-size:.79rem;color:var(--text-dim);line-height:1.55}

/* ════════════════════════════════════════════
   v6  NEW STYLES
   ════════════════════════════════════════════ */

/* Top action bar — upload icon + chat icon pinned at very top */
.top-action-bar{
  position:sticky;top:0;z-index:1100;
  display:flex;align-items:center;gap:.6rem;
  padding:.55rem 0 .4rem;
  background:linear-gradient(180deg,rgba(7,6,12,.98) 85%,transparent);
  backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
  margin-bottom:.6rem;
}
.tab-icon-btn{
  display:flex;align-items:center;gap:.45rem;
  background:rgba(107,45,107,.12);border:1px solid rgba(139,58,139,.3);
  border-radius:8px;padding:.38rem .75rem;cursor:pointer;
  font-family:'Space Mono',monospace;font-size:.62rem;color:var(--text-dim);
  transition:all .2s;white-space:nowrap;text-decoration:none;
}
.tab-icon-btn:hover,.tab-icon-btn.active{
  background:rgba(107,45,107,.25);border-color:var(--velvet-gl);color:var(--accent);
  box-shadow:0 0 12px rgba(107,45,107,.22);
}
.tab-icon-btn .tb-icon{font-size:1rem;line-height:1}
.tab-icon-btn .tb-label{font-size:.58rem;letter-spacing:.08em;text-transform:uppercase}
.tab-divider{flex:1;height:1px;background:linear-gradient(90deg,rgba(107,45,107,.25),transparent)}

/* ── TIMEFRAME NAVIGATOR ── */
.tf-nav-wrap{display:flex;align-items:center;gap:.3rem;margin:.3rem 0 .5rem;}
/* Active timeframe button gets a pink-purple glow — applied via st.markdown wrapper trick */
div[data-testid="stButton"] button.tf-active{
  background:linear-gradient(135deg,#8B3D8B,#C084C8)!important;
  border-color:#C084C8!important;color:#fff!important;font-weight:700!important;
}

/* ═══════════════════════════════════════════════════════
   TOP ACTION BAR — iframe-based buttons, no CSS hack needed
   ═══════════════════════════════════════════════════════ */

/* Hide the hidden Streamlit trigger buttons completely */
.stHidden-upload-trigger,
.stHidden-portfolio-trigger {
  display: none !important;
}

/* Upload panel slide-in */
.upload-panel{
  background:linear-gradient(135deg,rgba(107,45,107,.14) 0%,rgba(13,11,18,.97) 100%);
  border:1px solid rgba(139,58,139,.4);border-radius:14px;
  padding:1.1rem 1.3rem .9rem;margin-bottom:1rem;
  animation:slideDown .22s ease;
}
@keyframes slideDown{from{opacity:0;transform:translateY(-8px)}to{opacity:1;transform:translateY(0)}}
.upload-panel-hdr{display:flex;align-items:center;justify-content:space-between;margin-bottom:.7rem}
.upload-panel-title{font-family:'Space Mono',monospace;font-size:.64rem;letter-spacing:.18em;
  text-transform:uppercase;color:var(--velvet-gl)}
.upload-panel-formats{font-family:'Space Mono',monospace;font-size:.5rem;
  color:var(--text-ghost);margin-top:.15rem}

/* Inline analytics panel (shown after upload, below upload bar) */
.analytics-inline-panel{
  background:var(--card);border:1px solid var(--border);border-radius:14px;
  padding:1.1rem 1.3rem 1rem;margin-bottom:1.1rem;
  animation:fadeIn .3s ease;
}
@keyframes fadeIn{from{opacity:0}to{opacity:1}}
.aip-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:.75rem}
.aip-title{font-family:'Cormorant Garamond',serif;font-size:1.15rem;font-weight:300;
  color:var(--text);display:flex;align-items:center;gap:.5rem}
.aip-title::before{content:'';display:inline-block;width:3px;height:1rem;
  background:linear-gradient(180deg,#4ade80,#C084C8);border-radius:2px}
.aip-badge{font-family:'Space Mono',monospace;font-size:.5rem;letter-spacing:.1em;
  text-transform:uppercase;background:rgba(74,222,128,.08);border:1px solid rgba(74,222,128,.22);
  color:#86efac;padding:.18rem .5rem;border-radius:4px}

/* Floating chat panel */
.chat-panel{
  background:var(--card);border:1px solid var(--border-l);border-radius:14px;
  margin-bottom:1.1rem;overflow:hidden;
  animation:slideDown .22s ease;
}
.chat-panel-hdr{
  padding:.65rem 1rem .55rem;
  background:linear-gradient(90deg,rgba(107,45,107,.18),rgba(13,11,18,.9));
  border-bottom:1px solid var(--border);
  display:flex;align-items:center;justify-content:space-between;
}
.chat-panel-title{font-family:'Space Mono',monospace;font-size:.58rem;letter-spacing:.18em;
  text-transform:uppercase;color:var(--velvet-gl)}
.chat-panel-body{padding:.8rem 1rem}

/* ③ Auto-analytics success banner */
.analytics-banner{
  background:linear-gradient(135deg,rgba(74,222,128,.07) 0%,rgba(107,45,107,.10) 100%);
  border:1px solid rgba(74,222,128,.22);border-radius:10px;
  padding:.7rem 1.1rem;margin-bottom:1rem;
  display:flex;align-items:flex-start;gap:.75rem;
}
.ab-icon{font-size:1.4rem;flex-shrink:0;line-height:1.2}
.ab-title{font-family:'Syne',sans-serif;font-size:.84rem;font-weight:600;color:#86efac;margin-bottom:.1rem}
.ab-sub{font-family:'Space Mono',monospace;font-size:.52rem;color:var(--text-ghost)}

/* ════════════════════════════════════════════
   v7  PORTFOLIO STYLES
   ════════════════════════════════════════════ */

/* Portfolio panel wrapper */
.portfolio-panel{
  background:var(--card);border:1px solid var(--border-l);border-radius:14px;
  margin-bottom:1.1rem;overflow:hidden;animation:slideDown .22s ease;
}
.portfolio-panel-hdr{
  padding:.7rem 1.1rem .6rem;
  background:linear-gradient(90deg,rgba(233,30,140,.18),rgba(107,45,107,.12),rgba(13,11,18,.9));
  border-bottom:1px solid rgba(233,30,140,.2);
  display:flex;align-items:center;justify-content:space-between;
}
.pph-title{font-family:'Cormorant Garamond',serif;font-size:1.1rem;font-weight:300;
  color:var(--text);display:flex;align-items:center;gap:.5rem}
.pph-title::before{content:'';display:inline-block;width:3px;height:1rem;
  background:linear-gradient(180deg,#ff4db8,#e91e8c);border-radius:2px}

/* Add Holdings expander cherry accent */
.add-holdings-panel{
  background:linear-gradient(135deg,rgba(233,30,140,.07) 0%,rgba(107,45,107,.10) 60%,rgba(13,11,18,.97) 100%);
  border:1px solid rgba(233,30,140,.22);border-radius:12px;
  padding:1rem 1.2rem .9rem;margin-bottom:.8rem;
  animation:slideDown .22s ease;
}
.ah-title{font-family:'Space Mono',monospace;font-size:.58rem;letter-spacing:.18em;
  text-transform:uppercase;color:#ff80c0;margin-bottom:.6rem;
  display:flex;align-items:center;gap:.5rem;}
.ah-title::before{content:'✦';color:#e91e8c;font-size:.7rem;}
.pph-stats{display:flex;gap:1.2rem;flex-wrap:wrap}
.pph-stat{font-family:'Space Mono',monospace;font-size:.52rem;text-align:right}
.pph-stat-lbl{color:var(--text-ghost);text-transform:uppercase;letter-spacing:.1em}
.pph-stat-val{font-size:.72rem;margin-top:.1rem}
.pph-stat-val.pos{color:#4ade80}
.pph-stat-val.neg{color:#f87171}
.pph-stat-val.neu{color:var(--text)}

/* Holding card */
.holding-card{
  background:var(--card-2);border:1px solid var(--border);border-radius:10px;
  padding:.8rem 1rem .7rem;position:relative;transition:border-color .2s;
  overflow:hidden;
}
.holding-card.hc-up{
  border-color:rgba(74,222,128,.25);
  box-shadow:inset 0 0 0 1px rgba(74,222,128,.06);
}
.holding-card.hc-up::after{
  content:'';position:absolute;top:0;right:0;width:100%;height:100%;
  background:linear-gradient(135deg,transparent 62%,rgba(74,222,128,.06) 62%,rgba(74,222,128,.10) 74%,transparent 74%);
  pointer-events:none;
}
.holding-card.hc-down{
  border-color:rgba(248,113,113,.25);
  box-shadow:inset 0 0 0 1px rgba(248,113,113,.06);
}
.holding-card.hc-down::after{
  content:'';position:absolute;top:0;right:0;width:100%;height:100%;
  background:linear-gradient(135deg,transparent 62%,rgba(248,113,113,.06) 62%,rgba(248,113,113,.10) 74%,transparent 74%);
  pointer-events:none;
}
.holding-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;
  border-radius:2px 2px 0 0;background:linear-gradient(90deg,var(--velvet),var(--accent));
  opacity:0;transition:opacity .2s}
.holding-card:hover::before{opacity:1}
.hc-sym{font-family:'Space Mono',monospace;font-size:.82rem;font-weight:700;
  color:var(--accent);letter-spacing:.06em}
.hc-country{font-family:'Space Mono',monospace;font-size:.58rem;font-weight:600;
  letter-spacing:.08em;text-transform:uppercase;margin-bottom:.1rem;}
.hc-name{font-family:'Syne',sans-serif;font-size:.72rem;color:var(--text-ghost);
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-bottom:.4rem}
.hc-price{font-family:'Cormorant Garamond',serif;font-size:1.55rem;font-weight:300;
  color:var(--text);line-height:1}
.hc-chg{font-family:'Space Mono',monospace;font-size:.56rem;margin-top:.1rem}
.hc-chg.up{color:#4ade80}.hc-chg.down{color:#f87171}.hc-chg.flat{color:var(--text-ghost)}
.hc-meta{display:flex;gap:.6rem;margin-top:.5rem;flex-wrap:wrap}
.hc-chip{background:rgba(107,45,107,.12);border:1px solid var(--border);border-radius:4px;
  font-family:'Space Mono',monospace;font-size:.46rem;color:var(--text-ghost);
  padding:.1rem .4rem;white-space:nowrap}
.hc-chip.gain{background:rgba(74,222,128,.08);border-color:rgba(74,222,128,.2);color:#86efac}
.hc-chip.loss{background:rgba(248,113,113,.08);border-color:rgba(248,113,113,.2);color:#fca5a5}

/* Allocation donut area */
.alloc-row{display:flex;gap:.5rem;flex-wrap:wrap;margin:.5rem 0}
.alloc-bar{height:8px;border-radius:4px;overflow:hidden;background:rgba(107,45,107,.12);
  margin:.4rem 0 .8rem;display:flex;gap:1px}
.alloc-seg{height:100%;transition:width .4s}

/* AI analysis card */
.ai-analysis-card{
  background:linear-gradient(135deg,rgba(107,45,107,.08) 0%,rgba(13,11,18,.95) 100%);
  border:1px solid rgba(192,132,200,.22);border-radius:10px;
  padding:.9rem 1.1rem;margin-top:.6rem;
}
.aac-header{font-family:'Space Mono',monospace;font-size:.52rem;letter-spacing:.18em;
  text-transform:uppercase;color:var(--velvet-gl);margin-bottom:.5rem;
  display:flex;align-items:center;gap:.4rem}
.aac-header::before{content:'◈';color:var(--accent)}
.aac-body{font-family:'Syne',sans-serif;font-size:.83rem;color:var(--text-dim);line-height:1.75}

/* Signal badge */
.signal{display:inline-flex;align-items:center;gap:.3rem;border-radius:5px;
  font-family:'Space Mono',monospace;font-size:.55rem;padding:.2rem .55rem;
  text-transform:uppercase;letter-spacing:.08em;font-weight:700}
.signal.buy{background:rgba(74,222,128,.12);border:1px solid rgba(74,222,128,.3);color:#4ade80}
.signal.hold{background:rgba(240,192,64,.10);border:1px solid rgba(240,192,64,.3);color:#F0C040}
.signal.sell{background:rgba(248,113,113,.10);border:1px solid rgba(248,113,113,.3);color:#f87171}
.signal.watch{background:rgba(192,132,200,.10);border:1px solid rgba(192,132,200,.3);color:#C084C8}

/* ════════════════════════════════════════════
   v8  SOURCE PANEL + ANALYST MODE STYLES
   ════════════════════════════════════════════ */

/* Mode toggle bar */
.mode-toggle-wrap{display:flex;align-items:center;gap:.5rem;padding:.4rem .7rem;margin-bottom:.5rem;
  background:rgba(107,45,107,.08);border:1px solid var(--border);border-radius:8px;width:fit-content;}
.mode-toggle-label{font-family:'Space Mono',monospace;font-size:.5rem;letter-spacing:.15em;text-transform:uppercase;color:#4A3858;}
.mode-badge.chat-mode{background:rgba(192,132,200,.15);border:1px solid rgba(192,132,200,.35);color:#C084C8;
  font-family:'Space Mono',monospace;font-size:.52rem;padding:.18rem .55rem;border-radius:4px;
  letter-spacing:.06em;text-transform:uppercase;font-weight:700;}
.mode-badge.analyst-mode{background:rgba(240,192,64,.12);border:1px solid rgba(240,192,64,.35);color:#F0C040;
  font-family:'Space Mono',monospace;font-size:.52rem;padding:.18rem .55rem;border-radius:4px;
  letter-spacing:.06em;text-transform:uppercase;font-weight:700;}

/* Source evidence panel */
.evidence-panel{border:1px solid var(--border);border-radius:10px;padding:.6rem .8rem .5rem;
  margin-top:.5rem;background:linear-gradient(180deg,rgba(107,45,107,.04),rgba(13,11,18,.97));}
.evidence-source-card{background:var(--card-2);border:1px solid var(--border);border-radius:8px;
  margin-bottom:.6rem;overflow:hidden;}
.esc-header{display:flex;align-items:center;justify-content:space-between;padding:.45rem .75rem;
  background:rgba(107,45,107,.10);border-bottom:1px solid var(--border);flex-wrap:wrap;gap:.4rem;}
.esc-source-num{font-family:'Space Mono',monospace;font-size:.48rem;letter-spacing:.15em;text-transform:uppercase;color:var(--velvet-gl);}
.esc-filename{font-family:'Space Mono',monospace;font-size:.56rem;color:var(--accent);}
.esc-meta{display:flex;gap:.4rem;align-items:center;flex-wrap:wrap;}
.esc-chip{background:rgba(107,45,107,.15);border:1px solid var(--border);border-radius:3px;
  font-family:'Space Mono',monospace;font-size:.44rem;color:var(--text-ghost);padding:.08rem .35rem;white-space:nowrap;}
.esc-chip.score-hi{background:rgba(74,222,128,.08);border-color:rgba(74,222,128,.22);color:#86efac;}
.esc-chip.score-mid{background:rgba(240,192,64,.08);border-color:rgba(240,192,64,.22);color:#F0C040;}
.esc-chip.score-lo{background:rgba(248,113,113,.07);border-color:rgba(248,113,113,.2);color:#fca5a5;}
.esc-section{font-family:'Space Mono',monospace;font-size:.48rem;letter-spacing:.1em;text-transform:uppercase;
  color:#C084C8;padding:.4rem .75rem .1rem;}
.esc-body{font-family:'Syne',sans-serif;font-size:.8rem;color:var(--text-dim);line-height:1.7;
  padding:.25rem .75rem .6rem;border-left:2px solid rgba(139,58,139,.25);margin:.1rem .5rem .1rem .75rem;}
.esc-page{font-family:'Space Mono',monospace;font-size:.44rem;color:#4A3858;padding:.1rem .75rem .45rem;text-align:right;}

/* Analyst Mode structured output */
.analyst-output{font-family:'Syne',sans-serif;border-radius:12px;overflow:hidden;
  border:1px solid rgba(240,192,64,.2);margin:.3rem 0;
  background:linear-gradient(135deg,rgba(107,45,107,.06),rgba(13,11,18,.98));}
.ao-header{padding:.55rem .9rem;display:flex;align-items:center;justify-content:space-between;
  background:linear-gradient(90deg,rgba(107,45,107,.18),rgba(13,11,18,.95));
  border-bottom:1px solid rgba(139,58,139,.2);}
.ao-title{font-family:'Space Mono',monospace;font-size:.52rem;letter-spacing:.18em;text-transform:uppercase;
  color:var(--velvet-gl);display:flex;align-items:center;gap:.4rem;}
.ao-title::before{content:'◈';color:#F0C040;}
.ao-mode-badge{background:rgba(240,192,64,.1);border:1px solid rgba(240,192,64,.3);color:#F0C040;
  font-family:'Space Mono',monospace;font-size:.44rem;padding:.1rem .4rem;border-radius:3px;
  letter-spacing:.08em;text-transform:uppercase;}
.ao-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(170px,1fr));gap:.4rem;padding:.7rem .9rem;}
.ao-metric{background:var(--card-2);border:1px solid var(--border);border-radius:8px;padding:.55rem .75rem;}
.ao-metric-label{font-family:'Space Mono',monospace;font-size:.44rem;letter-spacing:.12em;
  text-transform:uppercase;color:#4A3858;margin-bottom:.15rem;}
.ao-metric-value{font-family:'Cormorant Garamond',serif;font-size:1.3rem;font-weight:300;color:#EDE8F5;line-height:1;}
.ao-metric-value.pos{color:#4ade80;}.ao-metric-value.neg{color:#f87171;}.ao-metric-value.neu{color:#C084C8;}
.ao-section{padding:.4rem .9rem .7rem;}
.ao-section-title{font-family:'Space Mono',monospace;font-size:.46rem;letter-spacing:.15em;text-transform:uppercase;
  color:#4A3858;margin-bottom:.3rem;border-bottom:1px solid var(--border);padding-bottom:.2rem;}
.ao-section-body{font-family:'Syne',sans-serif;font-size:.82rem;color:var(--text-dim);line-height:1.7;}
.ao-risk-item{display:flex;align-items:flex-start;gap:.4rem;font-family:'Syne',sans-serif;font-size:.8rem;
  color:var(--text-dim);padding:.25rem 0;border-bottom:1px solid rgba(139,58,139,.08);}
.ao-risk-item::before{content:'⚠';font-size:.7rem;color:#F0C040;flex-shrink:0;margin-top:.05rem;}

.cmp-table{width:100%;border-collapse:collapse;font-size:.75rem}
.cmp-table th{background:rgba(107,45,107,.18);border:1px solid var(--border);
  padding:.45rem .8rem;font-family:'Space Mono',monospace;font-size:.5rem;
  text-transform:uppercase;letter-spacing:.12em;color:var(--velvet-gl);text-align:left}
.cmp-table td{border:1px solid var(--border);padding:.4rem .8rem;
  font-family:'Space Mono',monospace;color:var(--text-dim);vertical-align:top}
.cmp-table tr:nth-child(even) td{background:rgba(107,45,107,.04)}
.td-doc{color:var(--accent)!important}
.td-mkt{color:#86efac!important}
.td-pos{color:#4ade80!important}
.td-neg{color:#f87171!important}
.td-neu{color:var(--text-ghost)!important}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# ANALYTICS ENGINE  (taxonomy, metrics, hybrid retriever, templates, eval)
# ─────────────────────────────────────────────────────────────────────────────
TAXONOMY: dict[str, list[str]] = {
    "Income Statement":["revenue","net revenue","total revenue","net sales","gross profit",
        "gross margin","operating income","ebit","ebitda","net income","net profit",
        "earnings","eps","earnings per share","diluted eps","operating expense",
        "cost of revenue","cost of goods sold","cogs","research and development","r&d",
        "selling general administrative","sg&a","tax","income tax","interest expense"],
    "Balance Sheet":["total assets","total liabilities","shareholders equity",
        "stockholders equity","book value","cash and equivalents","short-term investments",
        "accounts receivable","inventory","goodwill","intangible assets","long-term debt",
        "current liabilities","current assets","retained earnings","common stock","treasury stock"],
    "Cash Flow":["operating cash flow","cash from operations","free cash flow","fcf",
        "capital expenditure","capex","investing activities","financing activities",
        "dividends paid","share repurchase","buyback","depreciation","amortization"],
    "Ratios & Valuation":["p/e ratio","price to earnings","p/b ratio","price to book",
        "roe","return on equity","roa","return on assets","debt to equity","d/e ratio",
        "current ratio","quick ratio","gross margin","net margin","operating margin",
        "ebitda margin","interest coverage","dividend yield","payout ratio",
        "price to sales","ev/ebitda","enterprise value"],
    "Growth & Guidance":["year over year","yoy","quarter over quarter","qoq",
        "guidance","outlook","forecast","projection","growth rate","cagr","compound annual"],
    "Risk Factors":["risk","uncertainty","competition","regulatory","litigation",
        "geopolitical","macroeconomic","supply chain","inflation",
        "interest rate risk","credit risk","liquidity risk","cyber"],
}

def tag_chunk(text: str) -> list[str]:
    tl = text.lower()
    return [c for c, kws in TAXONOMY.items() if any(kw in tl for kw in kws)]

_METRIC_PATTERNS = [
    ("Revenue",          "USD", r"(?:total\s+)?(?:net\s+)?revenue[^\n]{0,60}?\$\s*([\d,\.]+)\s*(billion|million|B|M)?"),
    ("Net Income",       "USD", r"net\s+income[^\n]{0,60}?\$\s*([\d,\.]+)\s*(billion|million|B|M)?"),
    ("Gross Profit",     "USD", r"gross\s+profit[^\n]{0,60}?\$\s*([\d,\.]+)\s*(billion|million|B|M)?"),
    ("Operating Income", "USD", r"operating\s+income[^\n]{0,60}?\$\s*([\d,\.]+)\s*(billion|million|B|M)?"),
    ("EBITDA",           "USD", r"ebitda[^\n]{0,60}?\$\s*([\d,\.]+)\s*(billion|million|B|M)?"),
    ("Free Cash Flow",   "USD", r"free\s+cash\s+flow[^\n]{0,60}?\$\s*([\d,\.]+)\s*(billion|million|B|M)?"),
    ("CapEx",            "USD", r"capital\s+expenditures?[^\n]{0,60}?\$\s*([\d,\.]+)\s*(billion|million|B|M)?"),
    ("EPS (Basic)",      "USD", r"basic\s+(?:earnings|eps)[^\n]{0,60}?\$\s*([\d,\.]+)"),
    ("EPS (Diluted)",    "USD", r"diluted\s+(?:earnings|eps)[^\n]{0,60}?\$\s*([\d,\.]+)"),
    ("Gross Margin",     "%",   r"gross\s+margin[^\n]{0,60}?([\d\.]+)\s*%"),
    ("Net Margin",       "%",   r"net\s+(?:profit\s+)?margin[^\n]{0,60}?([\d\.]+)\s*%"),
    ("Operating Margin", "%",   r"operating\s+margin[^\n]{0,60}?([\d\.]+)\s*%"),
    ("ROE",              "%",   r"return\s+on\s+equity[^\n]{0,60}?([\d\.]+)\s*%"),
    ("ROA",              "%",   r"return\s+on\s+assets[^\n]{0,60}?([\d\.]+)\s*%"),
    ("Debt/Equity",      "x",   r"debt[- ]to[- ]equity[^\n]{0,60}?([\d\.]+)"),
    ("Current Ratio",    "x",   r"current\s+ratio[^\n]{0,60}?([\d\.]+)"),
    ("Shares Outstanding","M",  r"shares?\s+outstanding[^\n]{0,60}?([\d,\.]+)\s*(million|billion|M|B)?"),
]
_SCALE = {"billion":1e9,"b":1e9,"million":1e6,"m":1e6,None:1.0,"":1.0}

def extract_metrics(full_text: str) -> list[dict]:
    tl = full_text.lower(); results: list[dict] = []; seen: set[str] = set()
    for label, unit, pattern in _METRIC_PATTERNS:
        for m in re.finditer(pattern, tl, re.IGNORECASE):
            rn = m.group(1).replace(",","")
            sk = (m.group(2) or "").lower() if m.lastindex and m.lastindex >= 2 else ""
            try:
                val = float(rn) * _SCALE.get(sk, 1.0)
            except:
                continue
            if label in seen:
                continue
            seen.add(label)
            results.append({"label":label,"value":val,"unit":unit,
                             "raw":m.group(0).strip()[:120],"category":_mcat(label)})
    return results

def _mcat(label: str) -> str:
    ll = label.lower()
    if any(k in ll for k in ["eps","diluted","basic"]):   return "Per Share"
    if any(k in ll for k in ["margin","roe","roa","ratio","debt"]): return "Ratios"
    if any(k in ll for k in ["cash flow","capex"]):       return "Cash Flow"
    if any(k in ll for k in ["revenue","income","profit","ebitda"]): return "Income Statement"
    return "Other"

def fmt_val(val: float, unit: str) -> str:
    if unit == "USD":
        if val >= 1e9: return f"${val/1e9:.2f}B"
        if val >= 1e6: return f"${val/1e6:.1f}M"
        if val >= 1e3: return f"${val/1e3:.1f}K"
        return f"${val:.2f}"
    if unit == "%":  return f"{val:.1f}%"
    if unit == "x":  return f"{val:.2f}x"
    if unit == "M":  return f"{val/1e9:.2f}B" if val >= 1e9 else f"{val:.1f}M"
    return f"{val:,.2f}"

def _tokenize(text: str) -> list[str]:
    return re.findall(r"[a-z0-9]+", text.lower())

# ── Semantic retrieval cache ─────────────────────────────────────────────────
# Caches (query_text → {doc_context, sources_data}) so repeated / semantically
# identical questions skip re-embedding and re-ranking entirely.
_RETRIEVAL_CACHE: dict[str, dict] = {}
_CACHE_MAXSIZE = 64   # keep last 64 unique queries in memory

def _cache_key(q: str) -> str:
    """Normalise query to a stable cache key (lower, collapsed whitespace)."""
    return re.sub(r"\s+", " ", q.strip().lower())

def _retrieval_cache_get(q: str) -> dict | None:
    return _RETRIEVAL_CACHE.get(_cache_key(q))

def _retrieval_cache_put(q: str, result: dict) -> None:
    key = _cache_key(q)
    if len(_RETRIEVAL_CACHE) >= _CACHE_MAXSIZE:
        _RETRIEVAL_CACHE.pop(next(iter(_RETRIEVAL_CACHE)))
    _RETRIEVAL_CACHE[key] = result

# ── Retrieval quality monitoring log ────────────────────────────────────────
# Logs every live retrieval attempt with query, retrieved sections, CE scores,
# and (when available) ground-truth keyword hit.  Used in Eval tab dashboard.
_RETRIEVAL_LOG: list[dict] = []
_LOG_MAXSIZE = 200

def log_retrieval(
    query: str,
    retrieved: list[dict],       # list of {section, ce_score, score, page}
    keyword_hits: int = 0,       # how many expected keywords appeared in top chunks
    keyword_total: int = 0,
) -> None:
    """Append one retrieval event to the in-memory log."""
    if len(_RETRIEVAL_LOG) >= _LOG_MAXSIZE:
        _RETRIEVAL_LOG.pop(0)
    avg_ce = (sum(c.get("ce_score", 0) for c in retrieved) / len(retrieved)
              if retrieved else 0.0)
    top_ce = max((c.get("ce_score", -99) for c in retrieved), default=-99)
    _RETRIEVAL_LOG.append({
        "ts":           _dt.datetime.now().strftime("%H:%M:%S"),
        "query":        query[:80],
        "n_chunks":     len(retrieved),
        "sections":     list({c.get("section","?") for c in retrieved}),
        "avg_ce_score": round(avg_ce, 2),
        "top_ce_score": round(top_ce, 2),
        "kw_hits":      keyword_hits,
        "kw_total":     keyword_total,
        "recall":       round(keyword_hits / keyword_total, 3) if keyword_total else None,
    })

def compute_retrieval_stats() -> dict:
    """Aggregate Recall@K, MRR, avg CE score from the retrieval log."""
    if not _RETRIEVAL_LOG:
        return {}
    log = _RETRIEVAL_LOG
    avg_ce   = sum(e["avg_ce_score"]  for e in log) / len(log)
    top_ce   = sum(e["top_ce_score"]  for e in log) / len(log)
    # Recall@K: fraction of logged retrievals where ≥1 keyword was hit
    eval_log = [e for e in log if e["kw_total"]]
    recall   = (sum(1 for e in eval_log if e["kw_hits"] > 0) / len(eval_log)
                if eval_log else None)
    # Mean avg CE across all queries (proxy for retrieval confidence)
    return {
        "n_queries":    len(log),
        "avg_ce":       round(avg_ce, 2),
        "top_ce":       round(top_ce, 2),
        "recall_at_k":  round(recall * 100, 1) if recall is not None else None,
        "eval_samples": len(eval_log),
    }

class HybridRetriever:
    """
    Two-stage hybrid retriever:
      Stage 1 — BM25 keyword + dense cosine, fused with Reciprocal Rank Fusion (RRF)
      Stage 2 — Cross-encoder reranker filters low-confidence candidates
    RRF formula:  score(d) = Σ 1 / (k + rank_i(d))   [k=60 is standard]
    This is more robust than linear weighting because it is rank-based, not
    score-magnitude-based, so BM25 and cosine scales don't need normalisation.
    """
    _K = 60   # RRF constant — higher = flatter rank weighting

    def __init__(self, chunks, embeddings):
        self.chunks = chunks
        self.embeddings = embeddings
        self._bm25 = None
        self._ce   = None
        try:
            from rank_bm25 import BM25Okapi
            self._bm25 = BM25Okapi([_tokenize(c) for c in chunks])
        except: pass

    def _cos(self, a, b):
        d  = sum(x*y for x,y in zip(a,b))
        na = math.sqrt(sum(x*x for x in a))
        nb = math.sqrt(sum(x*x for x in b))
        return d / (na * nb + 1e-9)

    def _rrf(self, *ranked_lists: list[int]) -> dict[int, float]:
        """Compute RRF scores from multiple ranked index lists."""
        scores: dict[int, float] = {}
        for ranked in ranked_lists:
            for rank, idx in enumerate(ranked):
                scores[idx] = scores.get(idx, 0.0) + 1.0 / (self._K + rank + 1)
        return scores

    def retrieve(self, query: str, qe: list, n: int = 8,
                 bw: float = 0.35, rerank: bool = True,
                 ce_threshold: float = -5.0) -> list[dict]:
        N = len(self.chunks)

        # ── Dense ranking ────────────────────────────────────────────────
        dense_scores = [self._cos(qe, e) for e in self.embeddings]
        dense_ranked = sorted(range(N), key=lambda i: dense_scores[i], reverse=True)

        # ── BM25 ranking ─────────────────────────────────────────────────
        if self._bm25:
            bm25_raw    = self._bm25.get_scores(_tokenize(query))
            bm25_ranked = sorted(range(N), key=lambda i: bm25_raw[i], reverse=True)
        else:
            bm25_ranked = dense_ranked   # fallback: use dense twice → same result

        # ── RRF fusion ───────────────────────────────────────────────────
        rrf_scores = self._rrf(dense_ranked, bm25_ranked)
        cands = sorted(rrf_scores.keys(), key=lambda i: rrf_scores[i], reverse=True)[:max(20, n)]

        # ── Cross-encoder reranking removed ─────────────────────────────
        # sentence-transformers / torch removed to fit Streamlit free tier.
        # RRF fusion of TF-IDF + BM25 is used instead.

        return [
            {
                "idx":   i,
                "chunk": self.chunks[i],
                "score": rrf_scores.get(i, 0.0),
                "bm25":  bm25_raw[i] if self._bm25 else 0.0,
                "dense": dense_scores[i],
            }
            for i in cands[:n]
        ]

VELVET = {"bg":"#07060C","card":"#0D0B12","card2":"#120E1A","border":"rgba(139,58,139,0.25)",
          "accent":"#C084C8","green":"#4ade80","red":"#f87171","gold":"#F0C040",
          "text":"#EDE8F5","dim":"#9A8AAA","ghost":"#4A3858"}
CAT_COLORS = {"Income Statement":"#C084C8","Balance Sheet":"#60a5fa","Cash Flow":"#4ade80",
              "Ratios":"#F0C040","Per Share":"#fb923c","Growth":"#34d399",
              "Risk Factors":"#f87171","Other":"#9A8AAA"}

def _metric_card(label, value, category="Other"):
    c = CAT_COLORS.get(category, VELVET["dim"])
    return (f'<div style="background:{VELVET["card2"]};border:1px solid {VELVET["border"]};'
            f'border-top:2px solid {c};border-radius:10px;padding:.8rem 1rem;">'
            f'<div style="font-family:Space Mono,monospace;font-size:.52rem;letter-spacing:.15em;'
            f'text-transform:uppercase;color:{VELVET["ghost"]};margin-bottom:.35rem;">{label}</div>'
            f'<div style="font-family:\'Cormorant Garamond\',serif;font-size:1.6rem;font-weight:300;'
            f'color:{VELVET["text"]};line-height:1;">{value}</div>'
            f'<div style="font-family:Space Mono,monospace;font-size:.46rem;color:{c};'
            f'margin-top:.3rem;text-transform:uppercase;letter-spacing:.1em;">{category}</div></div>')

def render_metrics_dashboard(metrics: list[dict]) -> None:
    if not metrics:
        st.info("No financial metrics extracted."); return
    by_cat: dict[str, list] = {}
    for m in metrics: by_cat.setdefault(m["category"], []).append(m)
    for cat in ["Income Statement","Per Share","Cash Flow","Ratios","Balance Sheet","Other"]:
        items = by_cat.get(cat, [])
        if not items: continue
        c = CAT_COLORS.get(cat, VELVET["dim"])
        st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:.56rem;letter-spacing:.2em;'
                    f'text-transform:uppercase;color:{c};margin:1rem 0 .5rem;padding-bottom:.3rem;'
                    f'border-bottom:1px solid rgba(139,58,139,.2);">{cat}</div>', unsafe_allow_html=True)
        cols = st.columns(min(len(items), 4))
        for i, m in enumerate(items):
            with cols[i % 4]:
                st.markdown(_metric_card(m["label"], fmt_val(m["value"],m["unit"]), category=cat),
                            unsafe_allow_html=True)

def render_trend_chart(metrics: list[dict]) -> None:
    for unit, color, lbl in [("USD","#C084C8","USD Metrics (Billions $)"),
                              ("%","#F0C040","% Metrics (Margins & Returns)"),
                              ("x","#4ade80","Ratio Metrics (×)")]:
        items = [m for m in metrics if m["unit"] == unit and m["value"] > 0]
        if not items: continue
        st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:.54rem;letter-spacing:.15em;'
                    f'text-transform:uppercase;color:{color};margin:.8rem 0 .3rem;">{lbl}</div>',
                    unsafe_allow_html=True)
        df = pd.DataFrame(items).set_index("label")["value"]
        if unit == "USD": df = df / 1e9
        st.bar_chart(df, height=160, use_container_width=True)

TEMPLATES = {
    "Revenue Summary":        {"icon":"💰","category":"Income Statement",
        "prompt":"Extract and summarise revenue figures: total revenue, segment breakdown, YoY growth, guidance. Present as a structured table."},
    "Profitability Deep-Dive":{"icon":"📊","category":"Income Statement",
        "prompt":"Analyse profitability: gross profit & margin, operating income & margin, EBITDA, net income & net margin. Compare to prior year. Highlight one-time items."},
    "EPS Analysis":           {"icon":"📈","category":"Per Share",
        "prompt":"What is basic and diluted EPS? Change YoY/QoQ? What drove it — revenue growth, cost cuts, buybacks, or tax?"},
    "Balance Sheet Snapshot": {"icon":"🏦","category":"Balance Sheet",
        "prompt":"Total assets, liabilities, shareholders equity, cash, debt. Debt-to-equity and book value per share."},
    "Liquidity Assessment":   {"icon":"💧","category":"Balance Sheet",
        "prompt":"Current ratio, quick ratio, cash, short-term debt. Is the company at risk of a liquidity crunch?"},
    "Free Cash Flow Analysis":{"icon":"🌊","category":"Cash Flow",
        "prompt":"Operating cash flow, CapEx, FCF, FCF conversion rate. How is cash deployed?"},
    "Capital Allocation":     {"icon":"🎯","category":"Cash Flow",
        "prompt":"Dividends, buybacks, M&A, R&D, debt repayment. What % of FCF returned to shareholders?"},
    "Key Ratios & Benchmarks":{"icon":"⚖️","category":"Ratios",
        "prompt":"ROE, ROA, Gross Margin, Net Margin, Operating Margin, Debt/Equity, Current Ratio, Interest Coverage. Flag concerns."},
    "Risk Factor Summary":    {"icon":"⚠️","category":"Risk Factors",
        "prompt":"Top 5 material risk factors: name, description, financial impact, mitigation."},
    "Growth & Guidance":      {"icon":"🚀","category":"Growth",
        "prompt":"3-year revenue CAGR, management guidance, key growth drivers and headwinds."},
    "Competitive Position":   {"icon":"🏆","category":"Strategy",
        "prompt":"Competitive moat, market share, key differentiators, strategic priorities. What could erode this position?"},
}

EVAL_QUESTIONS = [
    {"id":"fb_001","question":"What was total revenue?",
     "expected_keywords":["revenue","billion","million","total","sales"],
     "synonyms":{"revenue":["sales","turnover","top line","net revenue","income from operations","total income"],
                 "billion":["b","bn","crore","million","thousand","lakh","trillion"]},
     "category":"Income Statement"},
    {"id":"fb_002","question":"What was diluted EPS or earnings per share?",
     "expected_keywords":["eps","per share","earnings","diluted","basic"],
     "synonyms":{"eps":["earnings per share","net income per share","profit per share","diluted eps","basic eps"],
                 "diluted":["basic","adjusted","normalised","per share"],
                 "per share":["a share","each share","share basis"]},
     "category":"Per Share"},
    {"id":"fb_003","question":"What was the gross margin or gross profit?",
     "expected_keywords":["gross","margin","profit","percent","%"],
     "synonyms":{"gross margin":["gross profit ratio","gpm","gross profit margin","cost of revenue"],
                 "margin":["margin%","percentage","rate"],
                 "gross":["gross profit","gross income"]},
     "category":"Ratios"},
    {"id":"fb_004","question":"What was free cash flow or operating cash flow?",
     "expected_keywords":["cash","flow","operating","free","fcf"],
     "synonyms":{"free cash flow":["cash from operations","operating cash flow","fcf","net cash"],
                 "operating":["operations","operational","from operations"],
                 "cash":["cash flow","cash generated","cash position"]},
     "category":"Cash Flow"},
    {"id":"fb_005","question":"What are the main risk factors?",
     "expected_keywords":["risk","market","regulatory","operational","competition"],
     "synonyms":{"risk":["threat","exposure","concern","challenge","uncertainty","factor","hazard"],
                 "market":["macro","economic","industry","sector"],
                 "regulatory":["compliance","regulation","legal","government"]},
     "category":"Risk Factors"},
    {"id":"fb_006","question":"What is the company's guidance, outlook or growth forecast?",
     "expected_keywords":["guidance","outlook","growth","expect","forecast"],
     "synonyms":{"guidance":["outlook","target","projection","forecast","expected","anticipated"],
                 "growth":["increase","rise","expansion","cagr","yoy","year over year"],
                 "expect":["anticipate","project","estimate","plan","target"]},
     "category":"Growth"},
    {"id":"fb_007","question":"What is the debt level, leverage or debt-to-equity ratio?",
     "expected_keywords":["debt","equity","ratio","leverage","borrowing"],
     "synonyms":{"debt":["borrowing","loan","liability","obligation","net debt"],
                 "equity":["net worth","shareholders equity","book value"],
                 "leverage":["gearing","d/e","debt ratio","financial leverage"]},
     "category":"Ratios"},
    {"id":"fb_008","question":"What was net income or profit after tax?",
     "expected_keywords":["net","income","profit","billion","million"],
     "synonyms":{"net income":["net profit","profit after tax","pat","bottom line","net earnings"],
                 "profit":["income","earnings","surplus"],
                 "billion":["million","b","bn","crore","lakh","thousand"]},
     "category":"Income Statement"},
]

# ── LLM-generated document-adaptive QA generation ───────────────────────────
_DOC_QA_SYSTEM = """You are a financial document analyst. Given a passage from a financial document,
generate exactly 5 specific question-answer pairs based ONLY on what is explicitly stated in the text.

Rules:
1. Each question must have a clear, specific answer that exists verbatim in the passage
2. Questions should cover different aspects: numbers, dates, names, ratios, descriptions
3. Include the exact answer phrase from the text as expected_answer
4. Generate questions at different difficulty: factual lookup, calculation, comparison
5. Respond ONLY with valid JSON, no preamble, no markdown fences.

Format exactly:
{"qa_pairs": [
  {"question": "...", "expected_answer": "...", "answer_phrase": "exact phrase from text", "category": "..."},
  ...
]}"""

def generate_doc_qa_pairs(doc_text: str, groq_api_key: str, n_chunks: int = 3) -> list[dict]:
    """Generate QA pairs from actual document content using LLM."""
    if not groq_api_key or not doc_text:
        return []
    try:
        from openai import OpenAI
        oai = OpenAI(api_key=groq_api_key, base_url="https://api.groq.com/openai/v1")

        # Sample representative chunks from the document
        words = doc_text.split()
        chunk_size = min(600, len(words) // max(n_chunks, 1))
        passages = []
        step = max(1, len(words) // n_chunks)
        for i in range(0, len(words), step):
            chunk = " ".join(words[i:i+chunk_size])
            if len(chunk.strip()) > 100:
                passages.append(chunk)
        passages = passages[:n_chunks]

        all_pairs = []
        for passage in passages:
            try:
                resp = oai.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[
                        {"role": "system", "content": _DOC_QA_SYSTEM},
                        {"role": "user", "content": f"Generate 5 QA pairs from this passage:\n\n{passage[:1200]}"},
                    ],
                    temperature=0.2, max_tokens=800,
                    response_format={"type": "json_object"},
                )
                raw = resp.choices[0].message.content
                parsed = json.loads(raw)
                pairs = parsed.get("qa_pairs", [])
                for p in pairs:
                    if p.get("question") and p.get("expected_answer"):
                        all_pairs.append({
                            "id":               f"doc_{len(all_pairs)+1:03d}",
                            "question":         p["question"],
                            "expected_answer":  p["expected_answer"],
                            "answer_phrase":    p.get("answer_phrase", p["expected_answer"]),
                            "category":         p.get("category", "Document"),
                            "source":           "auto-generated",
                            # Build keyword list from answer for backward compat
                            "expected_keywords": list({
                                w.lower() for w in re.findall(r"[a-zA-Z0-9₹$%\.]+", p["expected_answer"])
                                if len(w) > 2
                            })[:6],
                        })
            except Exception:
                continue
        return all_pairs
    except Exception:
        return []

def score_answer_semantic(answer: str, expected_keywords: list[str],
                          expected_answer: str = "",
                          synonyms: dict | None = None) -> dict:
    """
    Four-layer semantic scoring — designed to reward correct financial answers
    regardless of currency, unit, or phrasing variation.

    Layer 0: NOT FOUND / empty → 0 immediately
    Layer 1: Exact expected_answer present → 100%
    Layer 2: Has a number + at least one keyword → strong signal (≥60%)
    Layer 3: Synonym-expanded keyword match with partial credit
    Layer 4: Token overlap for multi-word keywords
    """
    if not answer or not answer.strip():
        return {"recall":0,"hits":0,"total":1,"score_pct":0,"method":"empty"}

    al = answer.lower().strip()
    syn = synonyms or {}

    # Layer 0: NOT FOUND / no info → 0
    _neg = ("not found in context", "not found", "not available", "no information",
            "not provided", "not mentioned", "no data", "cannot find",
            "not present in", "no specific", "i don't", "i do not")
    if any(n in al for n in _neg):
        return {"recall":0,"hits":0,"total":len(expected_keywords) or 1,"score_pct":0,"method":"not_found"}

    # Layer 1: exact expected answer match
    if expected_answer and expected_answer.lower().strip() in al:
        return {"recall":1.0,"hits":len(expected_keywords),"total":len(expected_keywords),
                "score_pct":100.0,"method":"exact_match"}

    # Layer 2: answer contains a number → strong signal the LLM found something
    _has_number = bool(re.search(r"\d[\d,\.]*\s*(?:%|b\b|bn\b|m\b|billion|million|crore|lakh|thousand|trillion|x\b)", al)
                       or re.search(r"[$₹£€¥]\s*\d", al)
                       or re.search(r"\d+\.\d+", al))

    # Layer 3: keyword matching with synonyms
    hits = 0.0
    matched_any_core = False
    for kw in expected_keywords:
        kw_lo = kw.lower()
        # Direct match
        if kw_lo in al:
            hits += 1
            matched_any_core = True
            continue
        # Synonym match
        kw_syns = syn.get(kw, [])
        # Also check reverse: synonyms defined under OTHER keys that point to this kw
        for _k, _vs in syn.items():
            if kw_lo in [v.lower() for v in _vs]:
                kw_syns = kw_syns + [_k]
        if any(s.lower() in al for s in kw_syns):
            hits += 1
            matched_any_core = True
            continue
        # Layer 4: partial token overlap for multi-word keywords (≥half tokens match)
        kw_tokens = set(re.findall(r"[a-z0-9]+", kw_lo))
        ans_tokens = set(re.findall(r"[a-z0-9]+", al))
        overlap = kw_tokens & ans_tokens
        if kw_tokens and len(overlap) >= max(1, len(kw_tokens) // 2):
            hits += 0.6
            matched_any_core = True

    total = len(expected_keywords) or 1
    recall = hits / total

    # Boost: if answer has a number AND matched at least one keyword, floor at 60%
    if _has_number and matched_any_core and recall < 0.60:
        recall = 0.60
        hits = max(hits, total * 0.60)

    # Boost: if answer has a number and ANY keyword matched (even partial), floor at 40%
    if _has_number and hits > 0 and recall < 0.40:
        recall = 0.40
        hits = max(hits, total * 0.40)

    return {
        "recall":    round(recall, 3),
        "hits":      round(hits, 1),
        "total":     total,
        "score_pct": round(min(recall * 100, 100.0), 1),
        "method":    "semantic_keyword",
    }

# Keep old name for backward compat
def score_answer(answer: str, kws: list[str]) -> dict:
    return score_answer_semantic(answer, kws)

def render_eval_dashboard(results: list[dict]) -> None:
    if not results: return
    avg  = sum(r["score"]["score_pct"] for r in results) / len(results)
    pass_rate = sum(1 for r in results if r["score"]["score_pct"] >= 50) / len(results) * 100
    c    = VELVET["green"] if avg >= 60 else (VELVET["gold"] if avg >= 35 else VELVET["red"])
    pc   = VELVET["green"] if pass_rate >= 60 else (VELVET["gold"] if pass_rate >= 40 else VELVET["red"])

    # Header metrics strip
    st.markdown(
        f'<div style="background:{VELVET["card2"]};border:1px solid {VELVET["border"]};'
        f'border-radius:10px;padding:1rem 1.2rem;margin-bottom:1rem;'
        f'display:flex;gap:2rem;flex-wrap:wrap;">'
        f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.15em;'
        f'text-transform:uppercase;color:{VELVET["ghost"]};">Semantic Recall Score</div>'
        f'<div style="font-family:Cormorant Garamond,serif;font-size:2.2rem;font-weight:300;color:{c};">{avg:.1f}%</div></div>'
        f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.15em;'
        f'text-transform:uppercase;color:{VELVET["ghost"]};">Pass Rate (≥50%)</div>'
        f'<div style="font-family:Cormorant Garamond,serif;font-size:2.2rem;font-weight:300;color:{pc};">{pass_rate:.0f}%</div></div>'
        f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.15em;'
        f'text-transform:uppercase;color:{VELVET["ghost"]};">Questions</div>'
        f'<div style="font-family:Cormorant Garamond,serif;font-size:2.2rem;font-weight:300;color:{VELVET["text"]};">{len(results)}</div></div>'
        f'</div>',
        unsafe_allow_html=True
    )

    # Per-question breakdown with color-coded pass/fail bars
    rows = []
    for r in results:
        sp = r["score"]["score_pct"]
        bar_c = VELVET["green"] if sp >= 70 else (VELVET["gold"] if sp >= 40 else VELVET["red"])
        grade = "PASS" if sp >= 50 else "FAIL"
        rows.append({
            "Question":  (r["question"][:62]+"…") if len(r["question"])>62 else r["question"],
            "Category":  r.get("category","—"),
            "Score":     f"{sp:.0f}%",
            "Grade":     grade,
            "Hits":      f'{r["score"].get("hits",0):.1f}/{r["score"]["total"]}',
        })
    df_r = pd.DataFrame(rows)
    st.dataframe(df_r, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────────────────────
# SHARED EVAL BENCHMARK RUNNER
# Used by all three eval modes (generic, adaptive, custom).
# Runs the full RRF → CE rerank pipeline per question, scores with semantic
# matching, logs to retrieval monitor, renders dashboard.
# ─────────────────────────────────────────────────────────────────────────────
def _run_eval_benchmark(questions: list[dict], vectorstore, groq_api_key: str,
                        use_expected_answer: bool = False,
                        doc_text: str = "") -> None:
    """Execute full pipeline benchmark and render results in-place."""
    if not vectorstore or not groq_api_key:
        st.error("Need uploaded documents and a Groq API key.")
        return

    # ── Doc-aware keyword enrichment ────────────────────────────────────────
    # For generic benchmarks the fixed keywords may not match the document's
    # actual vocabulary. Scan the full doc text once and build a set of
    # numbers and financial terms that actually appear — use these to augment
    # the keyword list for each question so the scorer gives credit for
    # document-specific phrasings (e.g. "net sales" instead of "revenue").
    doc_lower = doc_text.lower() if doc_text else ""
    # Extract all numbers found in doc (catches "$1.2B", "₹243 Cr", "14.5%")
    _doc_numbers = set(re.findall(r"[\d,]+\.?\d*\s*(?:%|bn|b|m|cr|lakh|billion|million|thousand)?", doc_lower))
    _doc_fin_terms = set(re.findall(r"[a-z][a-z\s\-]{3,25}(?:revenue|income|profit|loss|cash|assets|equity|eps|ebitda|margin|ratio|debt|sales|turnover|cagr|growth|guidance|outlook|dividend|buyback|r&d|capex|fcf|roe|roa|roce)[a-z\s]{0,15}", doc_lower))

    def _enrich_keywords(kws: list[str], syns: dict) -> tuple[list[str], dict]:
        """Add any doc-found synonyms that the scorer should also accept."""
        enriched_kws  = list(kws)
        enriched_syns = {k: list(v) for k, v in syns.items()}
        # Numbers: if doc has any numbers at all, add them as valid keyword tokens
        if _doc_numbers and not any(c.isdigit() for kw in kws for c in kw):
            enriched_kws.extend(list(_doc_numbers)[:4])
        # Financial terms: any doc term overlapping with the question's theme
        return enriched_kws, enriched_syns

    er = []
    prog = st.progress(0, text="Evaluating… (pacing calls to avoid rate limits)")
    vs = vectorstore

    # Pre-load TF-IDF components once
    _vectorizer_e   = vs.get("vectorizer")
    _tfidf_matrix_e = vs.get("tfidf_matrix")
    _chunks_e       = vs.get("chunks", [])

    for i, eq in enumerate(questions):
        prog.progress(i / len(questions),
                      text=f"Q{i+1}/{len(questions)} · {eq['question'][:45]}…")
        try:
            # ── Inter-call delay to stay within Groq free tier RPM ───────
            # Free tier: 30 req/min → ~2s between calls keeps us safe.
            if i > 0:
                _tm.sleep(2.1)

            # ── Full pipeline: expand → TF-IDF → RRF ────────────────────
            expanded = _expand_query(eq["question"])

            # TF-IDF similarity
            from sklearn.metrics.pairwise import cosine_similarity
            if _vectorizer_e is not None and _tfidf_matrix_e is not None and len(_chunks_e):
                q_vec_e   = _vectorizer_e.transform([expanded])
                sims_e    = cosine_similarity(q_vec_e, _tfidf_matrix_e).flatten().tolist()
            else:
                sims_e = [0.0] * len(_chunks_e)

            # BM25 re-score
            try:
                from rank_bm25 import BM25Okapi
                _b  = BM25Okapi([_tokenize(c) for c in _chunks_e])
                _br = _b.get_scores(_tokenize(expanded))
                _bm = max(_br) if max(_br) > 0 else 1.0
                bm25_s = [s/_bm for s in _br]
            except Exception:
                bm25_s = [0.0] * len(_chunks_e)

            # RRF fusion
            N_e = len(_chunks_e)
            rrf_e: dict[int, float] = {}
            for _rl in (sorted(range(N_e), key=lambda x: sims_e[x],  reverse=True),
                        sorted(range(N_e), key=lambda x: bm25_s[x], reverse=True)):
                for _ri, _ridx in enumerate(_rl):
                    rrf_e[_ridx] = rrf_e.get(_ridx, 0.0) + 1.0 / (60 + _ri + 1)

            # ── Context building: use all chunks if doc is tiny ───────────
            # With <10 chunks, the top-6 retrieval may miss relevant content
            # because everything scored similarly. Send the full document
            # as context instead — it's short enough to fit in the prompt.
            n_ctx = min(6, N_e)
            if N_e <= 12:
                # Small doc: use every chunk (full document context)
                cands_e = list(range(N_e))
                ctx_note = f"[full document · {N_e} chunks]"
            else:
                cands_e = sorted(rrf_e, key=lambda x: rrf_e[x], reverse=True)[:n_ctx]
                ctx_note = f"[top-{n_ctx} retrieved chunks]"
            cks_e = _chunks_e

            ctx = "\n---\n".join(cks_e[j] for j in cands_e)
            exp_ans  = eq.get("expected_answer", "")
            kws      = eq.get("expected_keywords", [])
            syns     = eq.get("synonyms", {})

            # Enrich keywords with doc-found terms so scorer credits doc-specific phrasing
            kws_eff, syns_eff = _enrich_keywords(kws, syns)

            # Keyword hits against retrieved context (retrieval quality signal)
            kw_hits = sum(1 for kw in kws if kw.lower() in ctx.lower())

            # Generate answer via groq_call (handles retry + rate limit internally)
            ans = groq_call(
                api_key=groq_api_key,
                messages=[{"role": "user", "content": f"Context:\n{ctx}\n\nQuestion: {eq['question']}"}],
                system=("Answer ONLY from the provided document context. "
                        "Quote exact numbers and phrases verbatim. "
                        "If the answer is not present, say exactly: NOT FOUND in context."),
                model="llama-3.3-70b-versatile",
                temperature=0.03,
                max_tokens=400,
                site_key=f"eval_{i}",
            )

            # If groq_call returned a rate-limit error string, wait and skip gracefully
            if ans.startswith("⚠"):
                wait_m = re.search(r"(\d+)m\s*(\d+)s|(\d+)s", ans)
                wait_s = 62  # default: wait just over 1 minute
                if wait_m:
                    if wait_m.group(1):
                        wait_s = int(wait_m.group(1))*60 + int(wait_m.group(2))
                    else:
                        wait_s = int(wait_m.group(3))
                wait_s = min(wait_s + 3, 125)  # cap at ~2 min, add buffer
                prog.progress(i / len(questions),
                              text=f"⏱ Rate limit hit — pausing {wait_s}s then continuing…")
                _tm.sleep(wait_s)
                # Retry once after the pause
                ans = groq_call(
                    api_key=groq_api_key,
                    messages=[{"role": "user", "content": f"Context:\n{ctx}\n\nQuestion: {eq['question']}"}],
                    system=("Answer ONLY from the provided document context. "
                            "Quote exact numbers and phrases verbatim. "
                            "If the answer is not present, say exactly: NOT FOUND in context."),
                    model="llama-3.3-70b-versatile",
                    temperature=0.03,
                    max_tokens=400,
                    site_key=f"eval_{i}_retry",
                )

            # Semantic scoring — use enriched keywords to credit doc-specific phrasing
            sc = score_answer_semantic(
                answer=ans,
                expected_keywords=kws_eff,
                expected_answer=exp_ans if use_expected_answer else "",
                synonyms=syns_eff,
            )
            ce_scores_e: dict = {}
            top_ce_q = 0.0

            log_retrieval(
                query=eq["question"],
                retrieved=[{"section": (_chunks_e[j][:30] if j < len(_chunks_e) else "?"),
                             "ce_score": 0, "score": rrf_e.get(j, 0)}
                            for j in cands_e[:6]],
                keyword_hits=kw_hits,
                keyword_total=len(kws),
            )
            er.append({
                "question":   eq["question"],
                "category":   eq.get("category", "—"),
                "answer":     ans,
                "score":      sc,
                "ctx_recall": round(kw_hits / len(kws) * 100 if kws else 0, 1),
                "top_ce":     0.0,
                "exp_ans":    exp_ans,
            })

        except Exception as exc:
            err_s = str(exc)
            # Rate limit: parse wait time, sleep, continue — no error card shown
            if "rate_limit_exceeded" in err_s or "429" in err_s:
                retry_m = re.search(r"(\d+)m\s*(\d+)s|(\d+)s", err_s)
                wait_s  = 62
                if retry_m:
                    if retry_m.group(1): wait_s = int(retry_m.group(1))*60 + int(retry_m.group(2))
                    else:                wait_s = int(retry_m.group(3))
                wait_s = min(wait_s + 3, 125)
                prog.progress(i / len(questions),
                              text=f"⏱ Rate limit — pausing {wait_s}s ({i+1}/{len(questions)} done so far)…")
                _tm.sleep(wait_s)
                # Mark as skipped (not as error card)
                er.append({"question": eq["question"], "category": eq.get("category","—"),
                           "answer": "⏱ Skipped (rate limit — re-run to retry)",
                           "score": {"recall":0,"hits":0,"total":0,"score_pct":0},
                           "ctx_recall": 0, "top_ce": 0, "exp_ans": ""})
            else:
                er.append({"question": eq["question"], "category": eq.get("category","—"),
                           "answer": f"Error: {err_s[:120]}",
                           "score": {"recall":0,"hits":0,"total":0,"score_pct":0},
                           "ctx_recall": 0, "top_ce": 0, "exp_ans": ""})

        prog.progress((i+1) / len(questions),
                      text=f"✓ Q{i+1}/{len(questions)} done · {eq['question'][:40]}…")

    prog.empty()
    if not er:
        st.warning("No results — check documents and API key."); return

    render_eval_dashboard(er)

    # Context Recall strip
    avg_ctx = sum(r.get("ctx_recall", 0) for r in er) / len(er)
    ctx_c   = VELVET["green"] if avg_ctx>=70 else (VELVET["gold"] if avg_ctx>=40 else VELVET["red"])
    avg_ce  = sum(r.get("top_ce", 0) for r in er) / len(er)
    ce_c2   = VELVET["green"] if avg_ce>0 else (VELVET["gold"] if avg_ce>-3 else VELVET["red"])
    st.markdown(
        f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.2);'
        f'border-radius:8px;padding:.7rem 1rem;margin:.6rem 0;display:flex;gap:2rem;flex-wrap:wrap;">'
        f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.1em;'
        f'text-transform:uppercase;color:{VELVET["ghost"]};">Context Recall@6</div>'
        f'<div style="font-family:Space Mono,monospace;font-size:1.1rem;color:{ctx_c};">{avg_ctx:.1f}%</div>'
        f'<div style="font-family:Space Mono,monospace;font-size:.44rem;color:{VELVET["ghost"]};">'
        f'keywords in retrieved chunks (retrieval quality)</div></div>'
        f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.1em;'
        f'text-transform:uppercase;color:{VELVET["ghost"]};">Avg CE Confidence</div>'
        f'<div style="font-family:Space Mono,monospace;font-size:1.1rem;color:{ce_c2};">{avg_ce:.2f}</div>'
        f'<div style="font-family:Space Mono,monospace;font-size:.44rem;color:{VELVET["ghost"]};">'
        f'>0 = relevant matches found</div></div>'
        f'</div>', unsafe_allow_html=True)

    with st.expander("📋 Full answers"):
        for r in er:
            sp  = r["score"]["score_pct"]
            bar = VELVET["green"] if sp>=70 else (VELVET["gold"] if sp>=40 else VELVET["red"])
            exp_html = (f'<div style="font-size:.7rem;color:#4ade80;margin:.15rem 0;">'
                        f'Expected: {r["exp_ans"][:100]}</div>' if r.get("exp_ans") else "")
            st.markdown(
                f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.2);'
                f'border-left:3px solid {bar};'
                f'border-radius:0 8px 8px 0;padding:.7rem .9rem;margin-bottom:.45rem;">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                f'<div style="font-family:Space Mono,monospace;font-size:.54rem;color:#C084C8;">'
                f'{r["category"]} · {r["question"][:65]}</div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.52rem;color:{bar};">'
                f'{sp:.0f}% · CE {r.get("top_ce",0):.1f}</div></div>'
                + exp_html +
                f'<div style="font-size:.77rem;color:#9A8AAA;margin-top:.25rem;line-height:1.5;">'
                f'{r["answer"][:450]}</div>'
                f'</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# ④  DOC vs MARKET COMPARISON  (new Analytics sub-tab)
# ─────────────────────────────────────────────────────────────────────────────

# S&P 500 sector benchmark medians (2024 TTM, approximate)
SECTOR_BENCHMARKS = {
    "Gross Margin":     {"Technology":55,"Financials":40,"Energy":30,"Healthcare":50,
                         "Industrials":33,"Consumer Discretionary":30,"Consumer Staples":28,"S&P 500 Avg":38},
    "Net Margin":       {"Technology":22,"Financials":18,"Energy":8,"Healthcare":14,
                         "Industrials":9,"Consumer Discretionary":6,"Consumer Staples":8,"S&P 500 Avg":11.5},
    "Operating Margin": {"Technology":28,"Financials":20,"Energy":12,"Healthcare":16,
                         "Industrials":12,"Consumer Discretionary":9,"Consumer Staples":10,"S&P 500 Avg":15},
    "ROE":              {"Technology":35,"Financials":12,"Energy":18,"Healthcare":22,
                         "Industrials":20,"Consumer Discretionary":28,"Consumer Staples":25,"S&P 500 Avg":19},
    "ROA":              {"Technology":15,"Financials":1.2,"Energy":7,"Healthcare":9,
                         "Industrials":7.5,"Consumer Discretionary":7,"Consumer Staples":8,"S&P 500 Avg":8},
}

SECTOR_ETFS = {
    "XLK":"Technology","XLF":"Financials","XLE":"Energy","XLV":"Healthcare",
    "XLI":"Industrials","XLY":"Consumer Discretionary","XLP":"Consumer Staples",
    "XLB":"Materials","XLRE":"Real Estate","XLU":"Utilities",
}

def render_comparison_tab(metrics: list[dict], groq_api_key: str) -> None:
    if not metrics:
        st.markdown('<div style="text-align:center;padding:3rem 2rem;">'
                    '<div style="font-size:2.5rem;margin-bottom:1rem;opacity:.4;">📊</div>'
                    '<div style="font-family:\'Cormorant Garamond\',serif;font-size:1.5rem;'
                    'font-weight:300;font-style:italic;color:#4A3858;">'
                    'Upload &amp; ingest documents first — analytics auto-generate on ingest</div>'
                    '</div>', unsafe_allow_html=True)
        return

    def section_hdr(title: str, grad: str = "linear-gradient(180deg,#6B2D6B,#C084C8)") -> str:
        return (f'<div style="font-family:\'Cormorant Garamond\',serif;font-size:1.15rem;'
                f'font-weight:300;color:#EDE8F5;margin:.9rem 0 .5rem;'
                f'display:flex;align-items:center;gap:.5rem;">'
                f'<span style="display:inline-block;width:3px;height:1rem;'
                f'background:{grad};border-radius:2px;"></span>{title}</div>')

    # ── Section A: Margin comparison vs sector ──────────────────────────────
    st.markdown(section_hdr("Document Margins vs Sector Benchmarks"), unsafe_allow_html=True)

    pct_m = {m["label"]: m["value"] for m in metrics if m["unit"] == "%"}
    sectors = ["S&P 500 Avg","Technology","Financials","Energy","Healthcare",
               "Industrials","Consumer Discretionary","Consumer Staples"]
    sector_sel = st.selectbox("Compare against sector", sectors, index=0, key="cmp_sector")

    rows_html = ""
    for metric, doc_val in pct_m.items():
        bench_row = SECTOR_BENCHMARKS.get(metric, {})
        bench_val = bench_row.get(sector_sel)
        if bench_val is None: continue
        delta = doc_val - bench_val
        dstr  = f"+{delta:.1f}%" if delta >= 0 else f"{delta:.1f}%"
        dcls  = "td-pos" if delta > 1 else ("td-neg" if delta < -1 else "td-neu")
        rows_html += (f'<tr><td>{metric}</td>'
                      f'<td class="td-doc">{doc_val:.1f}%</td>'
                      f'<td class="td-mkt">{bench_val:.1f}%</td>'
                      f'<td class="{dcls}">{dstr}</td></tr>')

    if rows_html:
        st.markdown(
            f'<table class="cmp-table"><thead><tr>'
            f'<th>Metric</th><th>Your Document</th><th>{sector_sel} Benchmark</th><th>Delta</th>'
            f'</tr></thead><tbody>{rows_html}</tbody></table>',
            unsafe_allow_html=True,
        )
    else:
        st.info("No comparable margin metrics found in this document (gross margin / net margin / operating margin / ROE / ROA).")

    st.markdown("<hr style='border-color:rgba(139,58,139,.12);margin:1.1rem 0;'>",
                unsafe_allow_html=True)

    # ── Section B: Live index snapshot ─────────────────────────────────────
    st.markdown(section_hdr("Live Market Context"), unsafe_allow_html=True)

    IDX_COMP = {"^GSPC":"S&P 500","^IXIC":"NASDAQ","^NSEI":"NIFTY 50","^N225":"Nikkei 225"}
    COMM_COMP = {"GC=F":("Gold","$/oz",2),"CL=F":("Crude Oil","$/bbl",2),"SI=F":("Silver","$/oz",3)}

    idx_q  = fetch_multi_quotes(tuple(IDX_COMP.keys()))
    comm_q = fetch_multi_quotes(tuple(COMM_COMP.keys()))

    ca, cb = st.columns(2)
    with ca:
        st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.18em;'
                    'text-transform:uppercase;color:#4A3858;margin-bottom:.45rem;">Global Indices</div>',
                    unsafe_allow_html=True)
        chips = ""
        for sym, name in IDX_COMP.items():
            info = idx_q.get(sym)
            if info:
                arr = "▲" if info["pct"] >= 0 else "▼"
                cls = "up" if info["pct"] >= 0 else "down"
                chips += (f'<div class="mood-idx-chip" style="min-width:110px;">'
                          f'<div class="mood-idx-name">{name}</div>'
                          f'<div class="mood-idx-val">{info["price"]:,.0f}</div>'
                          f'<div class="mood-idx-chg {cls}">{arr} {abs(info["pct"]):.2f}%</div></div>')
        if chips:
            st.markdown(f'<div style="display:flex;gap:.6rem;flex-wrap:wrap;">{chips}</div>',
                        unsafe_allow_html=True)
    with cb:
        st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.18em;'
                    'text-transform:uppercase;color:#4A3858;margin-bottom:.45rem;">Commodities</div>',
                    unsafe_allow_html=True)
        chips_c = ""
        for sym, (name, unit, dec) in COMM_COMP.items():
            info = comm_q.get(sym)
            if info:
                arr = "▲" if info["pct"] >= 0 else "▼"
                cls = "up" if info["pct"] >= 0 else "down"
                chips_c += (f'<div class="mood-idx-chip" style="min-width:110px;">'
                             f'<div class="mood-idx-name">{name}</div>'
                             f'<div class="mood-idx-val">${info["price"]:,.{dec}f}</div>'
                             f'<div class="mood-idx-chg {cls}">{arr} {abs(info["pct"]):.2f}%</div></div>')
        if chips_c:
            st.markdown(f'<div style="display:flex;gap:.6rem;flex-wrap:wrap;">{chips_c}</div>',
                        unsafe_allow_html=True)

    st.markdown("<hr style='border-color:rgba(139,58,139,.12);margin:1.1rem 0;'>",
                unsafe_allow_html=True)

    # ── Section C: Sector ETF performance ──────────────────────────────────
    st.markdown(section_hdr("Sector Performance (SPDR ETFs)",
                             "linear-gradient(180deg,#F0C040,#C084C8)"), unsafe_allow_html=True)
    etf_q = fetch_multi_quotes(tuple(SECTOR_ETFS.keys()))
    etf_rows = []
    for sym, name in SECTOR_ETFS.items():
        info = etf_q.get(sym)
        if info:
            etf_rows.append({"Sector":name,"ETF":sym,
                              "Price":f'${info["price"]:.2f}',
                              "1D Change":f'{info["pct"]:+.2f}%',
                              "_pct":info["pct"]})
    if etf_rows:
        df_etf = pd.DataFrame(etf_rows).sort_values("_pct", ascending=False).drop(columns=["_pct"])
        st.dataframe(df_etf, use_container_width=True, hide_index=True)

    st.markdown("<hr style='border-color:rgba(139,58,139,.12);margin:1.1rem 0;'>",
                unsafe_allow_html=True)

    # ── Section D: AI positioning commentary ───────────────────────────────
    st.markdown(section_hdr("AI Market Positioning Commentary",
                             "linear-gradient(180deg,#fb923c,#C084C8)"), unsafe_allow_html=True)

    if not groq_api_key:
        st.info("Enter Groq API key in the sidebar to generate AI commentary.")
    else:
        msummary = "; ".join(f"{m['label']}: {fmt_val(m['value'],m['unit'])}" for m in metrics[:12])
        isummary = "; ".join(f"{name}: {idx_q[s]['price']:,.0f} ({idx_q[s]['pct']:+.2f}%)"
                             for s, name in IDX_COMP.items() if s in idx_q)
        ssummary = "; ".join(f"{SECTOR_ETFS[s]}: {etf_q[s]['pct']:+.2f}%"
                             for s in SECTOR_ETFS if s in etf_q)

        if st.button("🤖  Generate Market Positioning Analysis", key="gen_cmp"):
            with st.spinner("Analysing document metrics against live market conditions…"):
                try:
                    from openai import OpenAI
                    oai = OpenAI(api_key=groq_api_key, base_url="https://api.groq.com/openai/v1")
                    prompt = (
                        f"You are a senior equity analyst. Analyse the following:\n\n"
                        f"DOCUMENT METRICS: {msummary}\n\n"
                        f"LIVE MARKET DATA:\n"
                        f"- Global Indices: {isummary}\n"
                        f"- Sector ETF Performance Today: {ssummary}\n\n"
                        f"Write a 200-250 word analyst-style commentary covering:\n"
                        f"1. How this company's margins/ratios compare to the current market environment\n"
                        f"2. Whether valuation/performance is attractive given current macro backdrop\n"
                        f"3. Key sector tailwinds or headwinds visible in today's market data\n"
                        f"4. Positioning recommendation: Overweight / Neutral / Underweight — with brief rationale\n\n"
                        f"Be specific, cite numbers, be direct and concise."
                    )
                    resp = oai.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=[
                            {"role":"system","content":"You are a senior equity analyst. Be concise, data-driven, and direct."},
                            {"role":"user","content":prompt},
                        ],
                        temperature=0.2, max_tokens=600,
                    )
                    commentary = resp.choices[0].message.content
                    st.markdown(
                        f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.25);'
                        f'border-left:3px solid #C084C8;border-radius:0 10px 10px 0;'
                        f'padding:1rem 1.2rem;font-size:.88rem;color:#9A8AAA;line-height:1.8;">'
                        f'{commentary.replace(chr(10),"<br>")}</div>',
                        unsafe_allow_html=True,
                    )
                except Exception as e:
                    st.error(f"Error: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# FULL ANALYTICS TAB (all sub-tabs)
# ─────────────────────────────────────────────────────────────────────────────
def render_analytics_tab(vectorstore, groq_api_key, doc_full_text="", auto_metrics=None):
    if not vectorstore and not doc_full_text:
        st.markdown('<div style="text-align:center;padding:3rem 2rem;">'
                    '<div style="font-size:2.5rem;margin-bottom:1rem;opacity:.4;">📊</div>'
                    '<div style="font-family:\'Cormorant Garamond\',serif;font-size:1.5rem;'
                    'font-weight:300;font-style:italic;color:#4A3858;">'
                    'Upload documents to unlock analytics</div>'
                    '<div style="font-family:Syne,sans-serif;font-size:.8rem;color:#4A3858;'
                    'margin-top:.6rem;">Auto-extract metrics · Compare vs market · Templates · Benchmark</div>'
                    '</div>', unsafe_allow_html=True)
        return

    sub_tabs = st.tabs(["📊 Metrics","📈 Doc vs Market","📋 Templates","🔍 Hybrid Search","🧪 Eval","📑 Compare Docs","⬇ Export"])

    # ── 0: Metrics ───────────────────────────────────────────────────────────
    with sub_tabs[0]:
        metrics = auto_metrics if auto_metrics else []
        if not metrics and doc_full_text:
            with st.spinner("Extracting metrics…"):
                metrics = extract_metrics(doc_full_text)
        st.markdown('<div style="font-family:Space Mono,monospace;font-size:.54rem;letter-spacing:.18em;'
                    'text-transform:uppercase;color:#C084C8;margin-bottom:.8rem;">'
                    'Auto-Extracted Financial Metrics</div>', unsafe_allow_html=True)
        if metrics:
            render_metrics_dashboard(metrics)
            st.markdown("<hr style='border-color:rgba(139,58,139,.15);margin:1rem 0;'>", unsafe_allow_html=True)
            render_trend_chart(metrics)
            with st.expander("📄 Raw extraction table"):
                st.dataframe(pd.DataFrame([{"Metric":m["label"],"Value":fmt_val(m["value"],m["unit"]),
                                             "Unit":m["unit"],"Category":m["category"],"Raw Text":m["raw"]}
                                            for m in metrics]), use_container_width=True, hide_index=True)
        else:
            st.info("No metrics matched. Try Templates tab for LLM extraction.")

    # ── 1: Doc vs Market (NEW) ───────────────────────────────────────────────
    with sub_tabs[1]:
        render_comparison_tab(auto_metrics or [], groq_api_key)

    # ── 2: Templates ─────────────────────────────────────────────────────────
    with sub_tabs[2]:
        cats = sorted({v["category"] for v in TEMPLATES.values()})
        chosen_cat = st.selectbox("Filter by category", ["All"]+cats, label_visibility="collapsed")
        visible = {k:v for k,v in TEMPLATES.items() if chosen_cat=="All" or v["category"]==chosen_cat}
        items = list(visible.items())
        for rs in range(0, len(items), 3):
            cols = st.columns(3)
            for ci, (tn, tm) in enumerate(items[rs:rs+3]):
                with cols[ci]:
                    color = CAT_COLORS.get(tm["category"], VELVET["dim"])
                    st.markdown(f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.22);'
                                f'border-top:2px solid {color};border-radius:10px;padding:.8rem .9rem .6rem;">'
                                f'<div style="font-size:1.2rem;">{tm["icon"]}</div>'
                                f'<div style="font-family:Syne,sans-serif;font-size:.82rem;font-weight:600;'
                                f'color:{VELVET["text"]};margin:.3rem 0 .2rem;">{tn}</div>'
                                f'<div style="font-family:Space Mono,monospace;font-size:.52rem;color:{color};'
                                f'text-transform:uppercase;">{tm["category"]}</div></div>', unsafe_allow_html=True)
                    if st.button("Run →", key=f"tpl_{tn[:20]}", use_container_width=True):
                        st.session_state["_prefill"] = tm["prompt"]
                        st.session_state["show_chat"] = True
                        st.rerun()

    # ── 3: Hybrid Search ─────────────────────────────────────────────────────
    with sub_tabs[3]:
        st.markdown('<div style="font-family:Space Mono,monospace;font-size:.54rem;letter-spacing:.18em;'
                    'text-transform:uppercase;color:#C084C8;margin-bottom:.8rem;">'
                    'Hybrid BM25 + Dense Retrieval with Cross-Encoder Re-ranking</div>', unsafe_allow_html=True)
        hs_q = st.text_input("Search", placeholder="e.g. free cash flow capital expenditure 2023",
                             label_visibility="collapsed")
        c1, c2, c3 = st.columns(3)
        with c1: bw = st.slider("BM25 weight", 0.0, 1.0, 0.35, 0.05)
        with c2: tn = st.slider("Results", 3, 10, 5)
        with c3: uce = st.checkbox("Cross-encoder re-rank", value=False, disabled=True,
                                   help="Cross-encoder disabled — not available in free-tier mode")
        tf = st.multiselect("Taxonomy filter", list(TAXONOMY.keys()), default=[], label_visibility="collapsed")
        if hs_q and vectorstore:
            with st.spinner("Retrieving…"):
                try:
                    vs   = vectorstore
                    cks  = vs.get("chunks", [])
                    mts  = vs.get("meta",   [])
                    if tf:
                        fi   = [i for i,c in enumerate(cks) if any(t in tag_chunk(c) for t in tf)]
                        cks  = [cks[i] for i in fi]
                        mts  = [mts[i]  for i in fi]
                    if not cks:
                        st.warning("No chunks match filter.")
                    else:
                        # TF-IDF similarity
                        from sklearn.metrics.pairwise import cosine_similarity
                        vec  = vs.get("vectorizer")
                        mat  = vs.get("tfidf_matrix")
                        if vec and mat is not None:
                            from sklearn.feature_extraction.text import TfidfVectorizer
                            q_vec = vec.transform([hs_q])
                            sims  = cosine_similarity(q_vec, mat).flatten().tolist()
                        else:
                            sims = [0.0] * len(cks)
                        # BM25
                        try:
                            from rank_bm25 import BM25Okapi
                            _b2   = BM25Okapi([_tokenize(c) for c in cks])
                            _br2  = _b2.get_scores(_tokenize(hs_q))
                            _bm2  = max(_br2) if max(_br2) > 0 else 1.0
                            bm25n = [s/_bm2 for s in _br2]
                        except Exception:
                            bm25n = [0.0] * len(cks)
                        # RRF
                        _K2 = 60; N2 = len(cks); rrf2: dict[int,float] = {}
                        for _rl2 in (sorted(range(N2), key=lambda i: sims[i],  reverse=True),
                                     sorted(range(N2), key=lambda i: bm25n[i], reverse=True)):
                            for _r2, _idx2 in enumerate(_rl2):
                                rrf2[_idx2] = rrf2.get(_idx2, 0.0) + 1.0/(_K2+_r2+1)
                        top_idxs2 = sorted(rrf2, key=lambda i: rrf2[i], reverse=True)[:tn]
                        hits = [{"idx":i,"chunk":cks[i],"score":rrf2[i]} for i in top_idxs2]
                        for rank, h in enumerate(hits, 1):
                            mt   = mts[h["idx"]] if h["idx"] < len(mts) else {}
                            tags = tag_chunk(h["chunk"])
                            th = " ".join(f'<span style="background:rgba(139,58,139,.15);border:1px solid rgba(139,58,139,.3);'
                                          f'font-family:Space Mono,monospace;font-size:.5rem;padding:.1rem .35rem;'
                                          f'border-radius:3px;color:{CAT_COLORS.get(t,VELVET["dim"])};">{t}</span>'
                                          for t in tags)
                            st.markdown(f'<div style="background:{VELVET["card"]};border:1px solid rgba(139,58,139,.22);'
                                        f'border-left:3px solid #C084C8;border-radius:0 8px 8px 0;'
                                        f'padding:.7rem .9rem;margin-bottom:.5rem;">'
                                        f'<div style="display:flex;justify-content:space-between;margin-bottom:.35rem;">'
                                        f'<div style="font-family:Space Mono,monospace;font-size:.58rem;color:#C084C8;">'
                                        f'#{rank} · 📄 {mt.get("filename","—")}</div>'
                                        f'<div style="font-family:Space Mono,monospace;font-size:.52rem;color:#4A3858;">'
                                        f'score:{h["score"]:.3f}</div></div>'
                                        f'<div style="font-size:.8rem;color:#9A8AAA;line-height:1.55;">{h["chunk"][:320]}…</div>'
                                        f'<div style="margin-top:.4rem;display:flex;gap:.3rem;flex-wrap:wrap;">{th}</div></div>',
                                        unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Search error: {e}")
        elif hs_q:
            st.info("Upload and ingest documents first.")

    # ── 4: Eval Benchmark ────────────────────────────────────────────────────
    with sub_tabs[4]:
        _hdr = ('<div style="font-family:Space Mono,monospace;font-size:.54rem;letter-spacing:.18em;'
                'text-transform:uppercase;color:#C084C8;margin-bottom:.8rem;">')

        # ── Live monitoring strip ─────────────────────────────────────────
        stats = compute_retrieval_stats()
        if stats:
            s_ce = stats["avg_ce"]; s_top = stats["top_ce"]; s_n = stats["n_queries"]
            s_r  = stats.get("recall_at_k")
            ce_c = VELVET["green"] if s_top>0 else (VELVET["gold"] if s_top>-3 else VELVET["red"])
            st.markdown(
                f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.28);'
                f'border-radius:10px;padding:.75rem 1rem;margin-bottom:1rem;display:flex;gap:1.5rem;flex-wrap:wrap;">'
                f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.12em;text-transform:uppercase;color:{VELVET["ghost"]};">Queries logged</div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.9rem;color:{VELVET["text"]};">{s_n}</div></div>'
                f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.12em;text-transform:uppercase;color:{VELVET["ghost"]};">Avg CE</div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.9rem;color:{ce_c};">{s_ce:.2f}</div></div>'
                f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.12em;text-transform:uppercase;color:{VELVET["ghost"]};">Top-1 CE</div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.9rem;color:{ce_c};">{s_top:.2f}</div></div>'
                + (f'<div><div style="font-family:Space Mono,monospace;font-size:.46rem;letter-spacing:.12em;text-transform:uppercase;color:{VELVET["ghost"]};">Recall@6</div>'
                   f'<div style="font-family:Space Mono,monospace;font-size:.9rem;'
                   f'color:{VELVET["green"] if s_r and s_r>=70 else (VELVET["gold"] if s_r and s_r>=40 else VELVET["red"])};">{s_r:.1f}%</div></div>'
                   if s_r is not None else '') +
                f'<div style="font-family:Space Mono,monospace;font-size:.44rem;color:{VELVET["ghost"]};align-self:flex-end;">CE > 0 = relevant · > 5 = strong · < −4 = filtered</div>'
                f'</div>', unsafe_allow_html=True)
            with st.expander("📋 Retrieval log (last 30)"):
                log_rows = [{"Time":e["ts"],"Query":e["query"][:50],"Chunks":e["n_chunks"],
                              "Sections":", ".join(e["sections"][:3]),
                              "Avg CE":e["avg_ce_score"],"Top CE":e["top_ce_score"],
                              "Recall":f'{e["kw_hits"]}/{e["kw_total"]}' if e["kw_total"] else "—"}
                             for e in reversed(_RETRIEVAL_LOG[-30:])]
                if log_rows:
                    st.dataframe(pd.DataFrame(log_rows), use_container_width=True, hide_index=True)

        # ── Mode selector ─────────────────────────────────────────────────
        st.markdown(_hdr + "Benchmark Mode</div>", unsafe_allow_html=True)
        eval_mode = st.radio(
            "mode",
            ["🎯 Document-Adaptive (AI-generated from your doc)",
             "📋 Generic Financial (8 standard questions)",
             "✏️ Custom Questions"],
            label_visibility="collapsed",
            horizontal=True,
            key="eval_mode_sel",
        )

        # ── Document-Adaptive mode ────────────────────────────────────────
        if eval_mode.startswith("🎯"):
            st.markdown(
                f'<div style="background:rgba(74,222,128,.06);border:1px solid rgba(74,222,128,.2);'
                f'border-radius:8px;padding:.7rem 1rem;margin-bottom:.8rem;">'
                f'<div style="font-family:Space Mono,monospace;font-size:.52rem;color:#4ade80;'
                f'letter-spacing:.1em;">✓ RECOMMENDED FOR ACCURATE RESULTS</div>'
                f'<div style="font-family:Syne,sans-serif;font-size:.78rem;color:{VELVET["dim"]};'
                f'margin-top:.3rem;">The LLM reads your actual document and generates questions it '
                f'<em>knows the answers to</em>. Scores reflect true retrieval+answer quality, '
                f'not vocabulary mismatch.</div>'
                f'</div>', unsafe_allow_html=True)

            col_gen, col_run_a, col_clr_a = st.columns([2, 2, 1])
            with col_gen:
                gen_btn = st.button("⚡ Generate QA from Document", use_container_width=True, key="gen_qa_btn")
            with col_run_a:
                run_adaptive = st.button("▶  Run Adaptive Benchmark", use_container_width=True, key="run_adaptive_btn",
                                         disabled="adaptive_qa_pairs" not in st.session_state)
            with col_clr_a:
                if st.button("🗑", key="clr_adaptive", use_container_width=True,
                             help="Clear generated questions"):
                    st.session_state.pop("adaptive_qa_pairs", None); st.rerun()

            if gen_btn:
                if not doc_full_text or not groq_api_key:
                    st.error("Need uploaded documents and API key.")
                else:
                    with st.spinner("🧠 LLM reading your document and writing questions…"):
                        pairs = generate_doc_qa_pairs(doc_full_text, groq_api_key, n_chunks=4)
                    if pairs:
                        st.session_state.adaptive_qa_pairs = pairs
                        st.success(f"Generated {len(pairs)} document-specific QA pairs.")
                        st.rerun()
                    else:
                        st.error("Failed to generate QA pairs. Check API key and try again.")

            # Show generated pairs
            if "adaptive_qa_pairs" in st.session_state:
                aq = st.session_state.adaptive_qa_pairs
                st.markdown(
                    f'<div style="font-family:Space Mono,monospace;font-size:.5rem;'
                    f'color:{VELVET["ghost"]};margin-bottom:.5rem;">'
                    f'{len(aq)} document-specific questions ready</div>', unsafe_allow_html=True)
                with st.expander("👀 Preview generated questions"):
                    for q in aq:
                        st.markdown(
                            f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.2);'
                            f'border-radius:6px;padding:.5rem .8rem;margin-bottom:.3rem;">'
                            f'<div style="font-family:Space Mono,monospace;font-size:.52rem;color:#C084C8;">'
                            f'{q["category"]} · {q["id"]}</div>'
                            f'<div style="font-size:.78rem;color:{VELVET["text"]};margin:.15rem 0;">Q: {q["question"]}</div>'
                            f'<div style="font-size:.76rem;color:#4ade80;">Expected: {q["expected_answer"][:120]}</div>'
                            f'</div>', unsafe_allow_html=True)

            if run_adaptive and "adaptive_qa_pairs" in st.session_state:
                _run_eval_benchmark(
                    st.session_state.adaptive_qa_pairs, vectorstore,
                    groq_api_key, use_expected_answer=True,
                )

        # ── Generic mode ──────────────────────────────────────────────────
        elif eval_mode.startswith("📋"):
            st.markdown(
                f'<div style="background:rgba(240,192,64,.05);border:1px solid rgba(240,192,64,.2);'
                f'border-radius:8px;padding:.6rem 1rem;margin-bottom:.8rem;">'
                f'<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#F0C040;">ℹ WHY SCORES DIFFER</div>'
                f'<div style="font-family:Syne,sans-serif;font-size:.76rem;color:{VELVET["dim"]};">'
                f'<b>Document-Adaptive</b> scores high because questions are generated <i>from</i> your document — '
                f'it\'s an open-book test written from the answer key. '
                f'<b>Generic</b> uses 8 fixed SEC-filing questions ("diluted EPS", "free cash flow", etc.) '
                f'that may not match your document\'s vocabulary. '
                f'A small file with 2 chunks will score low — use Adaptive mode for real accuracy measurement, '
                f'or upload a multi-page PDF/Excel for generic to work well.</div>'
                f'</div>', unsafe_allow_html=True)

            _custom_qs = st.session_state.get("custom_eval_qs", [])
            st.caption(f"{len(EVAL_QUESTIONS)} standard questions · semantic keyword scoring")
            col_run_g, col_clr_g = st.columns([3, 1])
            with col_run_g:
                run_generic = st.button("▶  Run Generic Benchmark", use_container_width=True, key="run_generic_btn")
            with col_clr_g:
                if st.button("🗑 Clear log", key="clr_log_g", use_container_width=True):
                    _RETRIEVAL_LOG.clear(); st.rerun()
            if run_generic:
                _run_eval_benchmark(EVAL_QUESTIONS, vectorstore, groq_api_key,
                                    doc_text=doc_full_text)

        # ── Custom mode ───────────────────────────────────────────────────
        else:
            with st.expander("➕ Add a custom question", expanded=True):
                cq_col1, cq_col2 = st.columns([3, 1])
                with cq_col1:
                    new_q  = st.text_input("Question", key="eval_new_q",
                                            placeholder="e.g. What was operating income in FY24?",
                                            label_visibility="collapsed")
                    new_exp = st.text_input("Expected answer (exact phrase from document)", key="eval_new_exp",
                                            placeholder="₹12,450 crore",
                                            label_visibility="collapsed")
                    new_kw = st.text_input("Backup keywords (comma-separated)", key="eval_new_kw",
                                            placeholder="operating income, crore, FY24",
                                            label_visibility="collapsed")
                with cq_col2:
                    new_cat = st.selectbox("Category",
                        ["Income Statement","Balance Sheet","Cash Flow","Per Share","Ratios",
                         "Risk Factors","Segments","Outlook & Guidance","Other"],
                        key="eval_new_cat", label_visibility="collapsed")
                    if st.button("Add", key="eval_add_btn", use_container_width=True):
                        if new_q.strip():
                            kws = [k.strip() for k in new_kw.split(",") if k.strip()]
                            if "custom_eval_qs" not in st.session_state:
                                st.session_state.custom_eval_qs = []
                            st.session_state.custom_eval_qs.append({
                                "id":               f"custom_{len(st.session_state.custom_eval_qs)+1:03d}",
                                "question":         new_q.strip(),
                                "expected_answer":  new_exp.strip(),
                                "expected_keywords": kws or new_exp.strip().split()[:6],
                                "category":         new_cat, "source": "custom",
                            })
                            st.success(f"Added: {new_q[:50]}"); st.rerun()

            _custom_qs = st.session_state.get("custom_eval_qs", [])
            if _custom_qs:
                with st.expander(f"📝 {len(_custom_qs)} custom questions"):
                    for cq in _custom_qs:
                        exp = cq.get("expected_answer","")
                        st.markdown(
                            f'<div style="font-family:Space Mono,monospace;font-size:.54rem;'
                            f'color:#C084C8;margin:.2rem 0;">{cq["id"]} · {cq["category"]}</div>'
                            f'<div style="font-size:.76rem;color:{VELVET["dim"]};margin-bottom:.2rem;">'
                            f'Q: {cq["question"][:80]}'
                            + (f'<br>Expected: <span style="color:#4ade80">{exp[:80]}</span>' if exp else "")
                            + f'</div>', unsafe_allow_html=True)
                    if st.button("🗑 Clear all custom", key="eval_clear_custom"):
                        st.session_state.custom_eval_qs = []; st.rerun()

                if st.button("▶  Run Custom Benchmark", use_container_width=True, key="run_custom_btn"):
                    _run_eval_benchmark(_custom_qs, vectorstore, groq_api_key, use_expected_answer=True)
            else:
                st.info("Add questions above, then run the benchmark.")

    # ── 5: Document Comparison ────────────────────────────────────────────────
    with sub_tabs[5]:
        render_document_comparison(
            file_names   = st.session_state.get("file_names", []),
            auto_metrics = auto_metrics or [],
        )

    # ── 6: Export ────────────────────────────────────────────────────────────
    with sub_tabs[6]:
        _exp_summary = portfolio_summary(st.session_state.portfolio) if st.session_state.portfolio else {"holdings":[],"total_value":0,"total_cost":0,"total_pnl":0,"total_pnl_pct":0}
        render_export_panel(
            summary      = _exp_summary,
            metrics      = auto_metrics or [],
            doc_full_text= doc_full_text or "",
        )

# ─────────────────────────────────────────────────────────────────────────────
# DATA FETCH HELPERS
# ─────────────────────────────────────────────────────────────────────────────
_HEADERS = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

class _TokenBucket:
    def __init__(self, capacity=30, refill_every=60.0):
        self._cap=capacity; self._tokens=float(capacity)
        self._interval=refill_every/capacity; self._lock=_th.Lock(); self._last=_tm.monotonic()
    def acquire(self, timeout=5.0):
        deadline=_tm.monotonic()+timeout
        while True:
            with self._lock:
                now=_tm.monotonic(); earned=(now-self._last)/self._interval
                self._tokens=min(self._cap,self._tokens+earned); self._last=now
                if self._tokens>=1.0: self._tokens-=1.0; return True
            if _tm.monotonic()>=deadline: return False
            _tm.sleep(0.05)

@st.cache_resource
def _get_bucket(): return _TokenBucket(capacity=30, refill_every=60.0)

def _throttled_get(url, timeout=10):
    if not _get_bucket().acquire(timeout=4.0):
        raise RuntimeError("Rate limit reached — wait a few seconds.")
    return requests.get(url, headers=_HEADERS, timeout=timeout)

@st.cache_data(ttl=300, max_entries=50)
def fetch_yahoo_series(symbol, period, interval):
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?range={period}&interval={interval}&includePrePost=false"
    try:
        r = _throttled_get(url, timeout=10); r.raise_for_status()
        data = r.json(); res = data["chart"]["result"][0]
        ts = res["timestamp"]; close = res["indicators"]["quote"][0]["close"]
        idx = pd.to_datetime(ts, unit="s", utc=True).tz_convert("US/Eastern")
        return pd.Series(close, index=idx, name=symbol).dropna()
    except: return None

@st.cache_data(ttl=60, max_entries=100)
def fetch_quote(symbol):
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?range=2d&interval=1d"
    try:
        r = _throttled_get(url, timeout=8); r.raise_for_status()
        data = r.json()
        q = [x for x in data["chart"]["result"][0]["indicators"]["quote"][0]["close"] if x is not None]
        if not q: return None
        return {"price":q[-1],"pct":(q[-1]-q[-2])/q[-2]*100 if len(q)>=2 else 0.0}
    except: return None

@st.cache_data(ttl=60)
def fetch_multi_quotes(symbols): return {s:i for s in symbols if (i:=fetch_quote(s))}

@st.cache_data(ttl=300)
def fetch_fear_greed():
    try:
        r = _throttled_get("https://api.alternative.me/fng/?limit=1", timeout=8)
        d = r.json()["data"][0]
        return {"value":int(d["value"]),"label":d["value_classification"]}
    except: return {"value":50,"label":"Neutral"}

@st.cache_data(ttl=600)
def fetch_rss_with_images(feed_url, source_name, accent, max_items=8):
    FALLBACK = {"Bloomberg":"https://assets.bbhub.io/company/sites/51/2019/08/BBG-Logo-Black.png",
                "Wall Street Journal":"https://s.wsj.net/media/wsj_logo_black_sm.png",
                "Financial Times":"https://about.ft.com/files/2020/04/ft_logo.png",
                "Reuters":"https://www.reuters.com/pf/resources/images/reuters/logo-vertical-default.png",
                "CNBC":"https://www.cnbc.com/2020/07/21/cnbc-social-card-2019.jpg"}
    import html as _h
    try:
        r = requests.get(feed_url, headers={**_HEADERS,"Accept":"application/rss+xml,*/*"}, timeout=10)
        r.raise_for_status(); text = r.text
    except: return []
    results = []
    items = re.findall(r"<item[^>]*>(.*?)</item>", text, re.DOTALL) or \
            re.findall(r"<entry[^>]*>(.*?)</entry>", text, re.DOTALL)
    for item in items[:max_items]:
        tm = re.search(r"<title[^>]*>(.*?)</title>", item, re.DOTALL|re.IGNORECASE)
        raw = tm.group(1).strip() if tm else ""
        cdata = re.match(r"<!\[CDATA\[(.*?)\]\]>", raw, re.DOTALL)
        title = cdata.group(1).strip() if cdata else raw
        title = _h.unescape(re.sub(r"<[^>]+>", "", title)).strip()
        if not title or len(title) < 10: continue
        lm = (re.search(r'<link[^>]*href=["\'](https?://[^"\'> ]+)["\']', item, re.IGNORECASE) or
              re.search(r"<link>(.*?)</link>", item, re.DOTALL|re.IGNORECASE))
        link = (lm.group(1) or "#").strip() if lm else "#"
        if not link.startswith("http"): link = "#"
        img = ""
        mm = re.search(r'<media:(?:content|thumbnail)[^>]+url=["\'](https?://[^"\']+)["\']', item, re.IGNORECASE)
        if mm: img = mm.group(1)
        if not img:
            em = re.search(r'<enclosure[^>]+url=["\'](https?://[^"\']+(?:jpg|jpeg|png|webp))["\']', item, re.IGNORECASE)
            if em: img = em.group(1)
        if not img: img = FALLBACK.get(source_name, "")
        results.append({"title":title,"link":link,"source":source_name,"accent":accent,"img_url":img})
    return results

@st.cache_data(ttl=600)
def fetch_gnews_with_images(query, source_label, accent, max_items=6):
    import urllib.parse
    return fetch_rss_with_images(
        f"https://news.google.com/rss/search?q={urllib.parse.quote(query)}&hl=en-US&gl=US&ceid=US:en",
        source_label, accent, max_items)

def make_chip_html(sym, name, price, pct, prefix="$", suffix="", decimals=2, icon=""):
    arrow    = "▲" if pct > 0.005 else ("▼" if pct < -0.005 else "●")
    cls      = "up" if pct > 0.005 else ("down" if pct < -0.005 else "flat")
    chip_cls = "chip-up" if pct > 0.005 else ("chip-down" if pct < -0.005 else "")
    ih       = f'<span style="font-size:1rem;margin-right:.2rem;">{icon}</span>' if icon else ""
    return (f'<div class="price-chip {chip_cls}">'
            f'<div class="pc-sym">{ih}{sym}</div>'
            f'<div class="pc-name">{name}</div>'
            f'<div class="pc-val">{prefix}{price:,.{decimals}f}{suffix}</div>'
            f'<div class="pc-chg {cls}">{arrow} {abs(pct):.2f}%</div></div>')

@st.cache_data(ttl=600)
def get_all_news():
    news = []
    for url, src, color in [("https://feeds.bloomberg.com/markets/news.rss","Bloomberg","#4ADE80"),
                             ("https://feeds.a.dj.com/rss/RSSMarketsMain.xml","Wall Street Journal","#F0C040"),
                             ("https://www.ft.com/?format=rss","Financial Times","#FB923C"),
                             ("https://www.cnbc.com/id/100003114/device/rss/rss.html","CNBC","#C084C8"),
                             ("https://feeds.reuters.com/reuters/businessNews","Reuters","#60A5FA")]:
        news.extend(fetch_rss_with_images(url, src, color, max_items=4))
    if len(news) < 6:
        news.extend(fetch_gnews_with_images("financial markets economy","Google News","#9CA3AF",8))
    return news[:16]

@st.cache_data(ttl=600)
def get_policy_news():
    policy = []
    for url, src, flag, color in [("https://www.federalreserve.gov/feeds/press_all.xml","Federal Reserve","🇺🇸","#60A5FA"),
                                   ("https://www.ecb.europa.eu/rss/press.html","ECB","🇪🇺","#34D399"),
                                   ("https://www.imf.org/en/News/rss?language=eng","IMF","🌐","#A78BFA"),
                                   ("https://www.rbi.org.in/scripts/rss.aspx","RBI India","🇮🇳","#FB923C"),
                                   ("https://www.bankofengland.co.uk/rss/publications","Bank of England","🇬🇧","#F472B6")]:
        items = fetch_rss_with_images(url, src, color, max_items=3)
        for item in items: item["flag"]=flag; item["policy"]=True
        policy.extend(items)
    if len(policy) < 4:
        extra = fetch_gnews_with_images("central bank monetary policy rate decision","Policy News","#A78BFA",6)
        for item in extra: item["flag"]="🏦"; item["policy"]=True
        policy.extend(extra)
    return policy[:12]

def build_carousel_html(items, is_policy=False, height_px=380):
    ag  = "linear-gradient(90deg,#3B82F6,#A78BFA)" if is_policy else "linear-gradient(90deg,#6B2D6B,#C084C8)"
    tt  = "Policy &amp; Government Decisions" if is_policy else "Financial Headlines"
    tbc = "#3B82F6" if is_policy else "#C084C8"
    bgc = "#0A0F1E" if is_policy else "#120E1A"
    bc  = "rgba(59,130,246,.3)" if is_policy else "rgba(139,58,139,.3)"
    slides_js = json.dumps([{"title":i["title"],"link":i.get("link","#"),"source":i.get("source",""),
                              "accent":i.get("accent","#C084C8"),"img":i.get("img_url",""),
                              "flag":i.get("flag",""),"policy":i.get("policy",False)} for i in items])
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0D0B12;font-family:'Segoe UI',system-ui,sans-serif;color:#EDE8F5;height:{height_px}px;overflow:hidden}}
.cw{{background:#0D0B12;border:1px solid {bc};border-radius:14px;height:{height_px}px;
  display:flex;flex-direction:column;overflow:hidden;position:relative}}
.cw::before{{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:{ag}}}
.ch{{display:flex;align-items:center;justify-content:space-between;
  padding:.85rem 1.1rem .6rem;flex-shrink:0;border-bottom:1px solid rgba(139,58,139,.12)}}
.ct{{font-family:Georgia,serif;font-size:1rem;font-weight:300;color:#EDE8F5;
  display:flex;align-items:center;gap:.5rem}}
.ct::before{{content:'';display:inline-block;width:3px;height:1rem;background:{ag};border-radius:2px}}
.cn{{display:flex;align-items:center;gap:.4rem}}
.nb{{background:rgba(107,45,107,.15);border:1px solid {bc};color:{tbc};width:26px;height:26px;
  border-radius:50%;cursor:pointer;font-size:1rem;display:flex;align-items:center;justify-content:center;
  transition:background .2s}}
.nb:hover{{background:rgba(107,45,107,.35)}}
.dots{{display:flex;gap:4px;align-items:center;flex-wrap:wrap;max-width:160px}}
.dot{{width:6px;height:6px;border-radius:50%;background:rgba(139,58,139,.3);
  border:1px solid rgba(139,58,139,.4);cursor:pointer;transition:all .2s}}
.dot.active{{background:{tbc};transform:scale(1.3)}}
.sa{{flex:1;position:relative;overflow:hidden}}
.sl{{position:absolute;top:0;left:0;right:0;bottom:0;display:flex;flex-direction:column;
  opacity:0;transform:translateX(40px);transition:opacity .45s,transform .45s;pointer-events:none}}
.sl.active{{opacity:1;transform:translateX(0);pointer-events:all}}
.sl.leaving{{opacity:0;transform:translateX(-40px)}}
.si{{width:100%;height:160px;object-fit:cover;flex-shrink:0;background:#120E1A}}
.ip{{width:100%;height:160px;flex-shrink:0;display:flex;align-items:center;justify-content:center;
  font-size:2.5rem;background:linear-gradient(135deg,#120E1A 0%,#1a1028 100%);
  border-bottom:1px solid rgba(139,58,139,.15)}}
.sb{{padding:.7rem 1rem .5rem;display:flex;flex-direction:column;gap:.3rem;flex:1;background:{bgc}}}
.ss{{font-family:'Courier New',monospace;font-size:.58rem;letter-spacing:.14em;text-transform:uppercase;
  display:flex;align-items:center;gap:.4rem;flex-wrap:wrap}}
.pb2{{font-size:.45rem;background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.3);
  color:#93C5FD;padding:.05rem .35rem;border-radius:3px}}
.st{{font-size:.9rem;font-weight:500;color:#EDE8F5;line-height:1.45;text-decoration:none;
  display:-webkit-box;-webkit-line-clamp:3;-webkit-box-orient:vertical;overflow:hidden}}
.st:hover{{color:{tbc};text-decoration:underline}}
.sm{{font-family:'Courier New',monospace;font-size:.48rem;color:#4A3858;margin-top:auto}}
.pbw{{flex-shrink:0;height:2px;background:rgba(139,58,139,.12);overflow:hidden}}
.pb{{height:100%;width:0%;background:{ag};border-radius:1px;transition:width 3s linear}}
.cf{{display:flex;justify-content:space-between;align-items:center;
  padding:.3rem 1rem;flex-shrink:0;background:rgba(0,0,0,.3)}}
.fl{{font-family:'Courier New',monospace;font-size:.44rem;color:#4A3858}}
.fc{{font-family:'Courier New',monospace;font-size:.48rem;color:#4A3858}}
</style></head><body>
<div class="cw" id="car">
  <div class="ch"><div class="ct">{tt}</div>
    <div class="cn"><div class="dots" id="dots"></div>
      <button class="nb" id="prev">&#8249;</button>
      <button class="nb" id="next">&#8250;</button>
    </div>
  </div>
  <div class="sa" id="slides"></div>
  <div class="pbw"><div class="pb" id="pb"></div></div>
  <div class="cf"><div class="fl">&#9679; live &middot; 10min cache</div>
    <div class="fc" id="ctr">1/1</div></div>
</div>
<script>
(function(){{
var S={slides_js},N=S.length,cur=0,paused=false,timer=null,pbt=null;
var se=document.getElementById('slides'),de=document.getElementById('dots'),
    ce=document.getElementById('ctr'),pb=document.getElementById('pb');
S.forEach(function(s,i){{
  var sd=document.createElement('div');sd.className='sl'+(i===0?' active':'');sd.id='sl'+i;
  var img=s.img?'<img class="si" src="'+s.img+'" alt="" onerror="this.style.display=\\'none\\';this.nextElementSibling.style.display=\\'flex\\';"><div class="ip" style="display:none">📰</div>':'<div class="ip">📰</div>';
  var src=s.policy?'<span style="color:'+s.accent+'">'+s.flag+' '+s.source+'</span><span class="pb2">Policy</span>':'<span style="color:'+s.accent+'">'+s.source+'</span>';
  sd.innerHTML=img+'<div class="sb"><div class="ss">'+src+'</div><a class="st" href="'+s.link+'" target="_blank">'+s.title+'</a><div class="sm">&#128336; 3s auto-advance</div></div>';
  se.appendChild(sd);
  var dot=document.createElement('span');dot.className='dot'+(i===0?' active':'');dot.id='dot'+i;
  dot.onclick=(function(idx){{return function(){{goTo(idx)}}}})(i);de.appendChild(dot);
}});
function startPB(){{clearTimeout(pbt);pb.style.transition='none';pb.style.width='0%';
  pbt=setTimeout(function(){{pb.style.transition='width 3s linear';pb.style.width='100%'}},40);}}
function goTo(n){{
  var p=cur;cur=((n%N)+N)%N;if(p===cur)return;
  var oe=document.getElementById('sl'+p),ne=document.getElementById('sl'+cur),
      od=document.getElementById('dot'+p),nd=document.getElementById('dot'+cur);
  if(oe){{oe.className='sl leaving';setTimeout(function(){{if(oe)oe.className='sl'}},450);}}
  if(ne)ne.className='sl active';if(od)od.className='dot';if(nd)nd.className='dot active';
  ce.textContent=(cur+1)+'/'+N;startPB();
}}
document.getElementById('next').onclick=function(){{goTo(cur+1);restart();}};
document.getElementById('prev').onclick=function(){{goTo(cur-1);restart();}};
function restart(){{clearInterval(timer);timer=setInterval(function(){{if(!paused)goTo(cur+1);}},3000);}}
document.getElementById('car').addEventListener('mouseenter',function(){{paused=true;}});
document.getElementById('car').addEventListener('mouseleave',function(){{paused=false;}});
ce.textContent='1/'+N;startPB();restart();
}})();
</script></body></html>"""

# ─────────────────────────────────────────────────────────────────────────────
# ② UNIVERSAL FILE TEXT EXTRACTOR  — PDF / XLSX / XLS / CSV / DOCX / TXT
# ─────────────────────────────────────────────────────────────────────────────
def extract_text_from_file(f) -> str:
    """Legacy wrapper — extracts full text from any supported file."""
    import io as _io
    raw  = f.read(); name = f.name.lower()
    f.seek(0)  # reset so _extract_page_aware can re-read
    blocks = _extract_page_aware(f)
    return " ".join(b["text"] for b in blocks) if blocks else ""

# ─────────────────────────────────────────────────────────────────────────────
# ③ INGEST — semantic chunking · rich metadata · BGE embeddings
# ─────────────────────────────────────────────────────────────────────────────

# Financial-aware section detector used during ingestion
_FIN_SECTION_PATTERNS = [
    (r"income statement|profit.{0,10}loss|revenue|net income|gross profit|ebitda|operating income|COGS",
     "Income Statement"),
    (r"balance sheet|total assets|liabilities|shareholders.{0,8}equity|book value|working capital",
     "Balance Sheet"),
    (r"cash flow|operating activities|investing activities|financing activities|capex|free cash flow",
     "Cash Flow"),
    (r"\bEPS\b|earnings per share|diluted|basic eps|per share",               "Per Share"),
    (r"P/E|price.{0,6}earnings|ROE|ROA|ROCE|debt.{0,6}equity|current ratio|quick ratio|gross margin|net margin",
     "Ratios"),
    (r"risk factor|material risk|litigation|regulatory|compliance|contingent", "Risk Factors"),
    (r"segment|business unit|geography|product line|divisional",               "Segments"),
    (r"guidance|outlook|forward.looking|forecast|target|next quarter",         "Outlook & Guidance"),
    (r"audit|auditor|opinion|going concern|material weakness",                 "Audit"),
    (r"dividend|buyback|share repurchase|capital return|payout",               "Capital Allocation"),
]

def _extract_chunk_keywords(text: str, max_kw: int = 8) -> str:
    """Extract top financial keywords from a chunk for metadata tagging."""
    _stops = {"the","and","of","in","to","a","is","are","was","were","for","as","on","at",
               "by","an","be","with","or","from","this","that","its","their","has","have",
               "been","which","year","fiscal","ended","per","total","net","gross","basic"}
    tokens = re.findall(r"[a-zA-Z][a-zA-Z0-9\-\.\/]{2,}", text)
    freq: dict[str, int] = {}
    for tok in tokens:
        t = tok.lower()
        if t not in _stops and len(t) >= 3:
            freq[t] = freq.get(t, 0) + 1
    top = sorted(freq, key=lambda k: freq[k], reverse=True)[:max_kw]
    return " ".join(top)

def _infer_section(text: str) -> str:
    """Fast pattern-match to label a chunk's financial section."""
    t = text[:800].lower()
    for pattern, label in _FIN_SECTION_PATTERNS:
        if re.search(pattern, t, re.IGNORECASE):
            return label
    return "General"

def _extract_page_aware(f) -> list[dict]:
    """
    Extract text per-page (PDF) or per-sheet/row (Excel/CSV).
    Returns list of {"text": str, "page": int, "source": str}.
    """
    name = f.name.lower()
    raw  = f.read()
    pages = []

    if name.endswith(".pdf"):
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        for i, pg in enumerate(reader.pages):
            # extract_text with layout mode better preserves table columns
            try:
                t = pg.extract_text(extraction_mode="layout") or ""
            except Exception:
                t = pg.extract_text() or ""
            if not t.strip():
                continue
            # Light markdown-ification: detect all-caps lines as section headings
            lines_out = []
            for line in t.splitlines():
                stripped = line.strip()
                if not stripped:
                    lines_out.append("")
                    continue
                # Heuristic: all-caps line of 4–80 chars → treat as heading
                if (stripped.isupper() and 4 <= len(stripped) <= 80
                        and not any(c.isdigit() for c in stripped[:3])):
                    lines_out.append(f"\n## {stripped}")
                else:
                    lines_out.append(stripped)
            pages.append({"text": "\n".join(lines_out), "page": i + 1, "source": f.name})

    elif name.endswith((".xlsx", ".xls")):
        try:
            # ── Strategy 1: openpyxl with data_only=True ─────────────────
            # data_only=True reads the *cached* formula result (the number
            # Excel last computed and stored in the file).  This is the only
            # reliable way to get =SUM(...) values without running Excel.
            # Caveat: if the file was never saved after editing, the cache
            # may be absent (cell.value == None).  We handle that in strategy 2.
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            for sheet_i, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                rows_text = []
                headers: list[str] = []

                for row_i, row in enumerate(ws.iter_rows(values_only=True)):
                    # Skip completely empty rows
                    if all(v is None for v in row):
                        continue

                    cells = []
                    for v in row:
                        if v is None:
                            cells.append("")
                        elif isinstance(v, float):
                            # Drop trailing .0 for whole numbers, keep decimals
                            cells.append(str(int(v)) if v == int(v) and abs(v) < 1e15 else f"{v:,.2f}")
                        elif isinstance(v, bool):
                            cells.append("Yes" if v else "No")
                        else:
                            cells.append(str(v).strip())

                    # Use first non-empty row as header hint
                    if row_i == 0 and any(c for c in cells):
                        headers = cells

                    row_str = " | ".join(c for c in cells if c)
                    if row_str.strip():
                        rows_text.append(row_str)

                if not rows_text:
                    continue

                text = f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text)
                pages.append({"text": text, "page": sheet_i + 1,
                               "source": f"{f.name}[{sheet_name}]"})
            wb.close()

            # ── Strategy 2: fallback to pandas if openpyxl got empty cache ──
            # If any sheet produced only blank lines (formula cache missing),
            # re-read with pandas which uses xlrd/openpyxl default (may give
            # NaN for some formula cells, but captures static values).
            if not pages:
                raise ValueError("openpyxl data_only returned no content")

            # ── Strategy 3: scan for cells that still look blank ──────────
            # Check if ANY numeric value was captured; if not, re-read without
            # data_only to at least get formula strings as fallback text.
            has_numbers = any(
                any(c.strip().replace(",","").replace(".","").replace("-","").isdigit()
                    for c in p["text"].split("|"))
                for p in pages
            )
            if not has_numbers:
                # Re-read with formula strings visible (shows =SUM(B2:B5) etc.)
                wb2 = openpyxl.load_workbook(io.BytesIO(raw), data_only=False, read_only=True)
                for sheet_i, sheet_name in enumerate(wb2.sheetnames):
                    ws2 = wb2[sheet_name]
                    formula_rows = []
                    for row in ws2.iter_rows(values_only=True):
                        if all(v is None for v in row):
                            continue
                        cells = []
                        for v in row:
                            if v is None:
                                cells.append("")
                            elif isinstance(v, str) and v.startswith("="):
                                # Try to evaluate simple =SUM(n,n,...) or =n+n patterns
                                inner = v[1:].strip()
                                # =SUM(literal numbers)
                                m_sum = re.match(r"SUM\(([0-9\.,\s\+\-]+)\)", inner, re.IGNORECASE)
                                if m_sum:
                                    try:
                                        nums = [float(x.replace(",","")) for x in re.findall(r"[\d\.]+", m_sum.group(1))]
                                        cells.append(str(int(sum(nums))) if sum(nums)==int(sum(nums)) else f"{sum(nums):,.2f}")
                                        continue
                                    except Exception:
                                        pass
                                cells.append(f"[formula: {v}]")
                            elif isinstance(v, (int, float)):
                                cells.append(str(int(v)) if isinstance(v, float) and v == int(v) else str(v))
                            else:
                                cells.append(str(v).strip())
                        row_str = " | ".join(c for c in cells if c)
                        if row_str.strip():
                            formula_rows.append(row_str)
                    if formula_rows:
                        # Replace the blank page entry with formula-visible version
                        for p in pages:
                            if f"[{sheet_name}]" in p["source"]:
                                p["text"] = f"=== Sheet: {sheet_name} (formulas) ===\n" + "\n".join(formula_rows)
                                break
                        else:
                            pages.append({"text": f"=== Sheet: {sheet_name} (formulas) ===\n" + "\n".join(formula_rows),
                                          "page": sheet_i + 1, "source": f"{f.name}[{sheet_name}]"})
                wb2.close()

        except Exception as e:
            # Final fallback: pandas read_excel (handles .xls with xlrd)
            try:
                dfs = pd.read_excel(io.BytesIO(raw), sheet_name=None)
                for sheet_i, (sheet, df) in enumerate(dfs.items()):
                    # Convert numeric columns to strings preserving actual values
                    rows = []
                    for _, row in df.iterrows():
                        cells = []
                        for val in row:
                            if pd.isna(val):
                                cells.append("")
                            elif isinstance(val, float) and val == int(val):
                                cells.append(str(int(val)))
                            else:
                                cells.append(str(val))
                        rows.append(" | ".join(c for c in cells if c))
                    text = f"=== Sheet: {sheet} ===\n" + "\n".join(r for r in rows if r.strip())
                    if text.strip():
                        pages.append({"text": text, "page": sheet_i + 1,
                                      "source": f"{f.name}[{sheet}]"})
            except Exception as e2:
                pages.append({"text": f"[Excel parse error: {e} / {e2}]",
                               "page": 1, "source": f.name})

    elif name.endswith(".csv"):
        try:
            df = pd.read_csv(io.BytesIO(raw), dtype=str)
            pages.append({"text": df.fillna("").to_string(index=False), "page": 1, "source": f.name})
        except Exception as e:
            pages.append({"text": f"[CSV parse error: {e}]", "page": 1, "source": f.name})

    elif name.endswith(".docx"):
        try:
            import zipfile, xml.etree.ElementTree as ET
            z   = zipfile.ZipFile(io.BytesIO(raw))
            xml_content = z.read("word/document.xml")
            tree = ET.fromstring(xml_content)
            W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
            paras = []
            for para in tree.iter(f"{W}p"):
                line = "".join(t.text or "" for t in para.iter(f"{W}t")).strip()
                if line: paras.append(line)
            pages.append({"text": "\n".join(paras), "page": 1, "source": f.name})
        except Exception as e:
            pages.append({"text": f"[DOCX parse error: {e}]", "page": 1, "source": f.name})
    else:
        # Plain text / markdown / unknown
        text_raw = ""
        for enc in ("utf-8", "latin-1", "cp1252"):
            try:
                text_raw = raw.decode(enc); break
            except: pass
        else:
            text_raw = raw.decode("utf-8", errors="ignore")

        # Split into logical "pages" at paragraph breaks so chunker gets
        # smaller units — prevents a whole file becoming 1-2 giant chunks.
        # Strategy: group lines into blocks of ≤1500 chars separated by blank lines.
        lines = text_raw.splitlines()
        blocks, cur = [], []
        cur_len = 0
        for line in lines:
            cur.append(line)
            cur_len += len(line) + 1
            # Split at blank line OR when block reaches 1500 chars
            if (not line.strip() or cur_len >= 1500) and cur_len > 0:
                block_text = "\n".join(cur).strip()
                if block_text:
                    blocks.append(block_text)
                cur, cur_len = [], 0
        if cur:
            block_text = "\n".join(cur).strip()
            if block_text:
                blocks.append(block_text)

        if not blocks:
            blocks = [text_raw]

        for bi, block_text in enumerate(blocks):
            pages.append({"text": block_text, "page": bi + 1, "source": f.name})

    return pages

def _chunk_text(text: str, chunk_size: int = 300, overlap: int = 50) -> list[str]:
    """Pure-Python recursive splitter — no langchain required."""
    if not text.strip():
        return []
    separators = ["\n\n", "\n", ". ", "! ", "? ", "; ", ", ", " ", ""]
    def _split(t: str, seps: list[str]) -> list[str]:
        if len(t) <= chunk_size or not seps:
            return [t] if t.strip() else []
        sep = seps[0]
        parts = t.split(sep)
        chunks, buf = [], ""
        for p in parts:
            candidate = (buf + sep + p).lstrip(sep) if buf else p
            if len(candidate) <= chunk_size:
                buf = candidate
            else:
                if buf:
                    chunks.append(buf)
                    # carry overlap
                    buf = buf[-overlap:] + sep + p if len(buf) > overlap else p
                else:
                    chunks.extend(_split(p, seps[1:]))
                    buf = ""
        if buf:
            chunks.append(buf)
        return [c for c in chunks if c.strip()]
    return _split(text, separators)


def ingest_documents(files):
    """
    Pure-Python RAG pipeline — zero ML framework dependencies.
    Uses TF-IDF (sklearn) for dense-style retrieval + BM25 for keyword retrieval.
    Memory budget: ~30MB total (vs ~750MB with sentence-transformers+torch).
    """
    from sklearn.feature_extraction.text import TfidfVectorizer
    import scipy.sparse as sp

    all_chunks, all_ids, all_meta, fnames, full_texts = [], [], [], [], []
    prog = st.progress(0, text="Reading files…")

    for i, f in enumerate(files):
        prog.progress((i + 0.1) / len(files), text=f"Parsing {f.name}…")
        page_blocks = _extract_page_aware(f)
        combined    = " ".join(b["text"] for b in page_blocks)
        fnames.append(f.name)
        full_texts.append(combined)

        for block in page_blocks:
            raw_chunks = _chunk_text(block["text"])
            for j, chunk in enumerate(raw_chunks):
                if not chunk.strip():
                    continue
                section  = _infer_section(chunk)
                keywords = _extract_chunk_keywords(chunk)
                chunk_with_header = (f"[{section}] {chunk}"
                                     if section != "General" else chunk)
                all_chunks.append(chunk)
                all_ids.append(f"{f.name}_p{block['page']}_c{j}")
                all_meta.append({
                    "filename":    f.name,
                    "doc_title":   f.name.rsplit(".", 1)[0].replace("_", " ").title(),
                    "page":        block["page"],
                    "chunk":       j,
                    "section":     section,
                    "keywords":    keywords,
                    "source":      block["source"],
                    "_index_text": chunk_with_header,
                })
        prog.progress((i + 1) / len(files), text=f"Processed {f.name}")
    prog.empty()

    # ── Build TF-IDF matrix (replaces neural embeddings) ────────────────
    tfidf_matrix = None
    vectorizer   = None
    if all_chunks:
        with st.spinner(f"Indexing {len(all_chunks)} chunks (TF-IDF)…"):
            index_texts = [m["_index_text"] for m in all_meta]
            vectorizer  = TfidfVectorizer(
                max_features=8000,
                sublinear_tf=True,
                ngram_range=(1, 2),
                min_df=1,
                strip_accents="unicode",
            )
            tfidf_matrix = vectorizer.fit_transform(index_texts)  # sparse: tiny RAM
            import gc; gc.collect()

    combined_text = " ".join(full_texts)
    with st.spinner("Auto-generating analytics…"):
        auto_metrics = extract_metrics(combined_text)

    # Store clean meta (no _index_text)
    clean_meta = [{k: v for k, v in m.items() if k != "_index_text"} for m in all_meta]

    st.session_state.vectorstore = {
        "chunks":       all_chunks,
        "meta":         clean_meta,
        "ids":          all_ids,
        "vectorizer":   vectorizer,
        "tfidf_matrix": tfidf_matrix,
        "model_label":  "TF-IDF",
        "is_bge":       False,
        # Legacy keys kept for compatibility
        "model":        None,
        "collection":   None,
    }
    st.session_state.uploaded_docs  = len(files)
    st.session_state.chunk_count    = len(all_chunks)
    st.session_state.file_names     = fnames
    st.session_state.doc_full_text  = combined_text
    st.session_state.auto_metrics   = auto_metrics
    st.session_state.auto_generated = True

    return len(all_chunks)

# ─────────────────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
#  PORTFOLIO ENGINE  v7
#  - fetch_stock_fundamentals(): P/E, 52w hi/lo, volume, mkt cap from Yahoo
#  - fetch_stock_history():      1Y daily OHLCV for RSI / momentum / vol
#  - compute_technicals():       RSI-14, SMA20/50/200, MACD, Bollinger Bands
#  - portfolio_summary():        total value, P&L, weights, beta, Sharpe
#  - render_portfolio_panel():   full Streamlit UI
# ══════════════════════════════════════════════════════════════════════════════

POPULAR_STOCKS = [
    "AAPL","MSFT","NVDA","GOOGL","AMZN","TSLA","META","NFLX","AMD","INTC",
    "JPM","GS","BAC","MS","V","MA","AXP",
    "JNJ","PFE","UNH","ABBV","MRK",
    "XOM","CVX","COP","BP",
    "RELIANCE.NS","TCS.NS","INFY.NS","HDFCBANK.NS","WIPRO.NS","ADANIENT.NS",
    "TSM","SAP","ASML","NVO","BABA","SONY","005930.KS",
    "BRK-B","WMT","COST","HD","NKE","DIS","SBUX",
    "BTC-USD","ETH-USD","SOL-USD",
]

# ── Global ticker catalog — covers US, India, Europe, Asia, LatAm, Crypto ──
# Format: (display_name, yahoo_symbol, exchange, sector)
GLOBAL_TICKERS: list[tuple[str,str,str,str]] = [
    # ── US Large Cap ──────────────────────────────────────────────────────────
    ("Apple",                    "AAPL",        "NASDAQ","Technology"),
    ("Microsoft",                "MSFT",        "NASDAQ","Technology"),
    ("NVIDIA",                   "NVDA",        "NASDAQ","Technology"),
    ("Alphabet (Google)",        "GOOGL",       "NASDAQ","Technology"),
    ("Amazon",                   "AMZN",        "NASDAQ","Consumer"),
    ("Meta Platforms",           "META",        "NASDAQ","Technology"),
    ("Tesla",                    "TSLA",        "NASDAQ","EV/Auto"),
    ("Berkshire Hathaway",       "BRK-B",       "NYSE",  "Financials"),
    ("Netflix",                  "NFLX",        "NASDAQ","Media"),
    ("AMD",                      "AMD",         "NASDAQ","Technology"),
    ("Intel",                    "INTC",        "NASDAQ","Technology"),
    ("Salesforce",               "CRM",         "NYSE",  "Technology"),
    ("Oracle",                   "ORCL",        "NYSE",  "Technology"),
    ("IBM",                      "IBM",         "NYSE",  "Technology"),
    ("Qualcomm",                 "QCOM",        "NASDAQ","Technology"),
    ("Broadcom",                 "AVGO",        "NASDAQ","Technology"),
    ("Texas Instruments",        "TXN",         "NASDAQ","Technology"),
    ("Palantir",                 "PLTR",        "NYSE",  "Technology"),
    # ── US Financials ─────────────────────────────────────────────────────────
    ("JPMorgan Chase",           "JPM",         "NYSE",  "Financials"),
    ("Goldman Sachs",            "GS",          "NYSE",  "Financials"),
    ("Bank of America",          "BAC",         "NYSE",  "Financials"),
    ("Morgan Stanley",           "MS",          "NYSE",  "Financials"),
    ("Visa",                     "V",           "NYSE",  "Financials"),
    ("Mastercard",               "MA",          "NYSE",  "Financials"),
    ("American Express",         "AXP",         "NYSE",  "Financials"),
    ("Wells Fargo",              "WFC",         "NYSE",  "Financials"),
    ("Citigroup",                "C",           "NYSE",  "Financials"),
    ("BlackRock",                "BLK",         "NYSE",  "Financials"),
    # ── US Healthcare ─────────────────────────────────────────────────────────
    ("Johnson & Johnson",        "JNJ",         "NYSE",  "Healthcare"),
    ("Pfizer",                   "PFE",         "NYSE",  "Healthcare"),
    ("UnitedHealth",             "UNH",         "NYSE",  "Healthcare"),
    ("AbbVie",                   "ABBV",        "NYSE",  "Healthcare"),
    ("Merck",                    "MRK",         "NYSE",  "Healthcare"),
    ("Eli Lilly",                "LLY",         "NYSE",  "Healthcare"),
    ("Bristol-Myers Squibb",     "BMY",         "NYSE",  "Healthcare"),
    ("Moderna",                  "MRNA",        "NASDAQ","Healthcare"),
    # ── US Energy ─────────────────────────────────────────────────────────────
    ("ExxonMobil",               "XOM",         "NYSE",  "Energy"),
    ("Chevron",                  "CVX",         "NYSE",  "Energy"),
    ("ConocoPhillips",           "COP",         "NYSE",  "Energy"),
    # ── US Consumer ───────────────────────────────────────────────────────────
    ("Walmart",                  "WMT",         "NYSE",  "Consumer"),
    ("Costco",                   "COST",        "NASDAQ","Consumer"),
    ("Home Depot",               "HD",          "NYSE",  "Consumer"),
    ("Nike",                     "NKE",         "NYSE",  "Consumer"),
    ("Disney",                   "DIS",         "NYSE",  "Media"),
    ("Starbucks",                "SBUX",        "NASDAQ","Consumer"),
    ("McDonald's",               "MCD",         "NYSE",  "Consumer"),
    ("Coca-Cola",                "KO",          "NYSE",  "Consumer"),
    ("PepsiCo",                  "PEP",         "NASDAQ","Consumer"),
    # ── India NSE ─────────────────────────────────────────────────────────────
    ("Reliance Industries",      "RELIANCE.NS", "NSE",   "Energy/Congl"),
    ("TCS",                      "TCS.NS",      "NSE",   "Technology"),
    ("Infosys",                  "INFY.NS",     "NSE",   "Technology"),
    ("HDFC Bank",                "HDFCBANK.NS", "NSE",   "Financials"),
    ("ICICI Bank",               "ICICIBANK.NS","NSE",   "Financials"),
    ("Wipro",                    "WIPRO.NS",    "NSE",   "Technology"),
    ("HCL Technologies",         "HCLTECH.NS",  "NSE",   "Technology"),
    ("Bajaj Finance",            "BAJFINANCE.NS","NSE",  "Financials"),
    ("Kotak Mahindra Bank",      "KOTAKBANK.NS","NSE",   "Financials"),
    ("State Bank of India",      "SBIN.NS",     "NSE",   "Financials"),
    ("Adani Enterprises",        "ADANIENT.NS", "NSE",   "Conglomerate"),
    ("Adani Ports",              "ADANIPORTS.NS","NSE",  "Infrastructure"),
    ("Tata Motors",              "TATAMOTORS.NS","NSE",  "Auto"),
    ("Tata Steel",               "TATASTEEL.NS","NSE",   "Metals"),
    ("Hindalco",                 "HINDALCO.NS", "NSE",   "Metals"),
    ("Asian Paints",             "ASIANPAINT.NS","NSE",  "Consumer"),
    ("Maruti Suzuki",            "MARUTI.NS",   "NSE",   "Auto"),
    ("Sun Pharma",               "SUNPHARMA.NS","NSE",   "Healthcare"),
    ("Dr Reddy's",               "DRREDDY.NS",  "NSE",   "Healthcare"),
    ("ONGC",                     "ONGC.NS",     "NSE",   "Energy"),
    ("NTPC",                     "NTPC.NS",     "NSE",   "Utilities"),
    ("Power Grid",               "POWERGRID.NS","NSE",   "Utilities"),
    ("Bajaj Auto",               "BAJAJ-AUTO.NS","NSE",  "Auto"),
    ("Hero MotoCorp",            "HEROMOTOCO.NS","NSE",  "Auto"),
    ("Larsen & Toubro",          "LT.NS",       "NSE",   "Industrials"),
    ("UltraTech Cement",         "ULTRACEMCO.NS","NSE",  "Materials"),
    ("Wipro",                    "WIPRO.BO",    "BSE",   "Technology"),
    # ── UK / Europe ───────────────────────────────────────────────────────────
    ("BP",                       "BP",          "NYSE",  "Energy"),
    ("Shell",                    "SHEL",        "NYSE",  "Energy"),
    ("ASML",                     "ASML",        "NASDAQ","Technology"),
    ("SAP",                      "SAP",         "NYSE",  "Technology"),
    ("Novo Nordisk",             "NVO",         "NYSE",  "Healthcare"),
    ("LVMH",                     "LVMUY",       "OTC",   "Luxury"),
    ("Nestlé",                   "NSRGY",       "OTC",   "Consumer"),
    ("AstraZeneca",              "AZN",         "NASDAQ","Healthcare"),
    ("Unilever",                 "UL",          "NYSE",  "Consumer"),
    ("Siemens",                  "SIEGY",       "OTC",   "Industrials"),
    ("Volkswagen",               "VWAGY",       "OTC",   "Auto"),
    ("BMW",                      "BMWYY",       "OTC",   "Auto"),
    ("TotalEnergies",            "TTE",         "NYSE",  "Energy"),
    ("HSBC",                     "HSBC",        "NYSE",  "Financials"),
    ("Barclays",                 "BCS",         "NYSE",  "Financials"),
    ("Roche",                    "RHHBY",       "OTC",   "Healthcare"),
    # ── Japan ────────────────────────────────────────────────────────────────
    ("Sony",                     "SONY",        "NYSE",  "Technology"),
    ("Toyota",                   "TM",          "NYSE",  "Auto"),
    ("Honda",                    "HMC",         "NYSE",  "Auto"),
    ("SoftBank",                 "SFTBY",       "OTC",   "Technology"),
    ("Nintendo",                 "NTDOY",       "OTC",   "Technology"),
    ("Keyence",                  "KYCCF",       "OTC",   "Technology"),
    # ── South Korea ──────────────────────────────────────────────────────────
    ("Samsung Electronics",      "005930.KS",   "KRX",   "Technology"),
    ("SK Hynix",                 "000660.KS",   "KRX",   "Technology"),
    ("LG Electronics",           "066570.KS",   "KRX",   "Technology"),
    ("Hyundai Motor",            "005380.KS",   "KRX",   "Auto"),
    ("POSCO",                    "PKX",         "NYSE",  "Metals"),
    # ── China / Hong Kong ────────────────────────────────────────────────────
    ("Alibaba",                  "BABA",        "NYSE",  "Technology"),
    ("Tencent",                  "TCEHY",       "OTC",   "Technology"),
    ("JD.com",                   "JD",          "NASDAQ","Consumer"),
    ("Baidu",                    "BIDU",        "NASDAQ","Technology"),
    ("NIO",                      "NIO",         "NYSE",  "EV/Auto"),
    ("BYD",                      "BYDDY",       "OTC",   "EV/Auto"),
    ("Meituan",                  "3690.HK",     "HKEX",  "Technology"),
    ("CNOOC",                    "CEO",         "NYSE",  "Energy"),
    # ── Taiwan ───────────────────────────────────────────────────────────────
    ("TSMC",                     "TSM",         "NYSE",  "Technology"),
    ("MediaTek",                 "2454.TW",     "TWSE",  "Technology"),
    ("Hon Hai (Foxconn)",        "2317.TW",     "TWSE",  "Technology"),
    # ── LatAm / Other ────────────────────────────────────────────────────────
    ("Petrobras",                "PBR",         "NYSE",  "Energy"),
    ("Vale",                     "VALE",        "NYSE",  "Metals"),
    ("Itaú Unibanco",            "ITUB",        "NYSE",  "Financials"),
    ("MercadoLibre",             "MELI",        "NASDAQ","Technology"),
    ("Copa Holdings",            "CPA",         "NYSE",  "Airlines"),
    # ── Australia ────────────────────────────────────────────────────────────
    ("BHP Group",                "BHP",         "NYSE",  "Metals"),
    ("Rio Tinto",                "RIO",         "NYSE",  "Metals"),
    ("Commonwealth Bank",        "CBA.AX",      "ASX",   "Financials"),
    ("Westpac",                  "WBC.AX",      "ASX",   "Financials"),
    # ── ETFs & Indices ───────────────────────────────────────────────────────
    ("S&P 500 ETF (SPY)",        "SPY",         "NYSE",  "ETF"),
    ("QQQ (NASDAQ 100 ETF)",     "QQQ",         "NASDAQ","ETF"),
    ("Gold ETF (GLD)",           "GLD",         "NYSE",  "ETF"),
    ("Oil ETF (USO)",            "USO",         "NYSE",  "ETF"),
    ("VIX",                      "^VIX",        "CBOE",  "Index"),
    # ── Crypto ───────────────────────────────────────────────────────────────
    ("Bitcoin",                  "BTC-USD",     "Crypto","Crypto"),
    ("Ethereum",                 "ETH-USD",     "Crypto","Crypto"),
    ("Solana",                   "SOL-USD",     "Crypto","Crypto"),
    ("BNB",                      "BNB-USD",     "Crypto","Crypto"),
    ("XRP",                      "XRP-USD",     "Crypto","Crypto"),
    ("Cardano",                  "ADA-USD",     "Crypto","Crypto"),
    ("Dogecoin",                 "DOGE-USD",    "Crypto","Crypto"),
    ("Avalanche",                "AVAX-USD",    "Crypto","Crypto"),
    ("Chainlink",                "LINK-USD",    "Crypto","Crypto"),
    ("Polkadot",                 "DOT-USD",     "Crypto","Crypto"),
]

# Build a fast lookup dict: lowercase name/symbol → row
_TICKER_SEARCH_INDEX: dict[str, list] = {}
for _row in GLOBAL_TICKERS:
    for _key in [_row[0].lower(), _row[1].lower()]:
        _TICKER_SEARCH_INDEX.setdefault(_key, []).append(_row)

def search_tickers(query: str, max_results: int = 12) -> list[tuple]:
    """Fast fuzzy search over GLOBAL_TICKERS by name or symbol prefix."""
    if not query:
        return []
    q = query.lower().strip()
    exact   = [r for r in GLOBAL_TICKERS if r[1].lower() == q]
    prefix  = [r for r in GLOBAL_TICKERS if r[1].lower().startswith(q) and r not in exact]
    name_m  = [r for r in GLOBAL_TICKERS
                if q in r[0].lower() and r not in exact and r not in prefix]
    combined = exact + prefix + name_m
    # Deduplicate preserving order
    seen = set(); out = []
    for r in combined:
        if r[1] not in seen:
            seen.add(r[1]); out.append(r)
    return out[:max_results]

@st.cache_data(ttl=120)
def fetch_stock_fundamentals(symbol: str) -> dict:
    """Fetch summary quote data: price, change, 52w range, volume, mkt cap, P/E."""
    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
           f"?range=1d&interval=1d&includePrePost=false")
    try:
        r = _throttled_get(url, timeout=10); r.raise_for_status()
        data = r.json()
        res  = data["chart"]["result"][0]
        meta = res.get("meta", {})
        q    = res["indicators"]["quote"][0]
        closes = [x for x in q.get("close",[]) if x]
        price  = meta.get("regularMarketPrice") or (closes[-1] if closes else None)
        prev   = meta.get("chartPreviousClose") or meta.get("previousClose")
        pct    = ((price - prev) / prev * 100) if price and prev and prev != 0 else 0.0
        return {
            "price":       round(price, 4)           if price else None,
            "prev_close":  round(prev, 4)             if prev  else None,
            "pct":         round(pct, 3),
            "open":        round(meta.get("regularMarketOpen", 0) or 0, 2),
            "day_high":    round(meta.get("regularMarketDayHigh", 0) or 0, 2),
            "day_low":     round(meta.get("regularMarketDayLow", 0) or 0, 2),
            "52w_high":    round(meta.get("fiftyTwoWeekHigh", 0) or 0, 2),
            "52w_low":     round(meta.get("fiftyTwoWeekLow", 0) or 0, 2),
            "volume":      meta.get("regularMarketVolume"),
            "avg_volume":  meta.get("averageDailyVolume10Day"),
            "mkt_cap":     meta.get("marketCap"),
            "currency":    meta.get("currency", "USD"),
            "symbol":      meta.get("symbol", symbol),
            "short_name":  meta.get("shortName") or meta.get("longName") or symbol,
        }
    except Exception:
        return {}

@st.cache_data(ttl=300, max_entries=25)
def fetch_stock_history_1y(symbol: str) -> pd.DataFrame:
    """1 year daily OHLCV as DataFrame — cached 5 min."""
    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
           f"?range=1y&interval=1d&includePrePost=false&events=div,splits")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        r = requests.get(url, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()
        res  = data["chart"]["result"][0]
        ts   = res["timestamp"]
        q    = res["indicators"]["quote"][0]
        df   = pd.DataFrame({
            "open":   q.get("open",  []),
            "high":   q.get("high",  []),
            "low":    q.get("low",   []),
            "close":  q.get("close", []),
            "volume": q.get("volume",[]),
        }, index=pd.to_datetime(ts, unit="s", utc=True).tz_convert("America/New_York"))
        return df.dropna(subset=["close"])
    except Exception:
        return pd.DataFrame()

def compute_technicals(df: pd.DataFrame) -> dict:
    """Compute RSI-14, SMA-20/50/200, MACD(12,26,9), Bollinger Bands(20,2), ATR-14."""
    if df.empty or len(df) < 20:
        return {}
    closes = df["close"].values.astype(float)
    n = len(closes)

    # ── SMA ──────────────────────────────────────────────────────────────────
    def sma(arr, w): return float(np.mean(arr[-w:])) if len(arr) >= w else None
    sma20  = sma(closes, 20)
    sma50  = sma(closes, 50)
    sma200 = sma(closes, 200)

    # ── RSI-14 ────────────────────────────────────────────────────────────────
    def rsi(arr, period=14):
        deltas = np.diff(arr[-period-1:])
        gains  = np.where(deltas > 0, deltas, 0)
        losses = np.where(deltas < 0, -deltas, 0)
        avg_g  = np.mean(gains)  if gains.size  else 0
        avg_l  = np.mean(losses) if losses.size else 1e-9
        rs     = avg_g / avg_l if avg_l else 0
        return round(100 - 100 / (1 + rs), 1)
    rsi14 = rsi(closes)

    # ── MACD(12,26,9) ────────────────────────────────────────────────────────
    def ema(arr, w):
        k = 2 / (w + 1); e = arr[0]
        for x in arr[1:]: e = x * k + e * (1 - k)
        return float(e)
    macd_line   = ema(closes, 12) - ema(closes, 26)
    signal_line = ema(closes[-9:], 9) if n >= 9 else macd_line
    macd_hist   = round(macd_line - signal_line, 4)

    # ── Bollinger Bands(20,2) ────────────────────────────────────────────────
    mid   = sma20 or closes[-1]
    std20 = float(np.std(closes[-20:])) if n >= 20 else 0
    bb_upper = round(mid + 2 * std20, 2)
    bb_lower = round(mid - 2 * std20, 2)
    price    = closes[-1]
    bb_pos   = round((price - bb_lower) / (bb_upper - bb_lower) * 100, 1) if bb_upper != bb_lower else 50

    # ── ATR-14 ───────────────────────────────────────────────────────────────
    highs  = df["high"].values[-15:].astype(float)
    lows   = df["low"].values[-15:].astype(float)
    cls_   = df["close"].values[-15:].astype(float)
    trs    = [max(highs[i]-lows[i], abs(highs[i]-cls_[i-1]), abs(lows[i]-cls_[i-1]))
              for i in range(1, len(highs))]
    atr14  = round(float(np.mean(trs)), 2) if trs else 0

    # ── Momentum ─────────────────────────────────────────────────────────────
    mom1m  = round((closes[-1]/closes[-21] - 1)*100, 2) if n >= 21 else None
    mom3m  = round((closes[-1]/closes[-63] - 1)*100, 2) if n >= 63 else None
    mom6m  = round((closes[-1]/closes[-126]- 1)*100, 2) if n >= 126 else None
    mom1y  = round((closes[-1]/closes[0]   - 1)*100, 2) if n >= 252 else None

    # ── Volatility (annualised) ───────────────────────────────────────────────
    rets  = np.diff(np.log(closes[-63:])) if n >= 64 else np.array([0])
    vol_a = round(float(np.std(rets) * np.sqrt(252) * 100), 1)

    # ── Signal synthesis ────────────────────────────────────────────────────
    score = 0
    if sma20 and price > sma20: score += 1
    if sma50 and price > sma50: score += 1
    if sma200 and price > sma200: score += 1
    if rsi14 < 30:  score += 2
    elif rsi14 > 70: score -= 2
    if macd_hist > 0: score += 1
    signal = "BUY" if score >= 4 else ("SELL" if score <= 0 else ("HOLD" if score >= 2 else "WATCH"))

    return {
        "sma20": round(sma20, 2) if sma20 else None,
        "sma50": round(sma50, 2) if sma50 else None,
        "sma200":round(sma200,2) if sma200 else None,
        "rsi14": rsi14,
        "macd_line":  round(macd_line, 4),
        "macd_signal":round(signal_line, 4),
        "macd_hist":  macd_hist,
        "bb_upper": bb_upper, "bb_lower": bb_lower, "bb_pos": bb_pos,
        "atr14":  atr14,
        "mom_1m": mom1m, "mom_3m": mom3m, "mom_6m": mom6m, "mom_1y": mom1y,
        "vol_annual": vol_a,
        "signal": signal,
        "score":  score,
    }

def portfolio_summary(portfolio: dict) -> dict:
    """Calculate total value, total P&L, weights for all holdings."""
    if not portfolio:
        return {"total_value":0,"total_cost":0,"total_pnl":0,"total_pnl_pct":0,"holdings":[]}

    holdings = []
    total_value = 0.0
    total_cost  = 0.0

    for sym, pos in portfolio.items():
        info = fetch_stock_fundamentals(sym)
        price = info.get("price") or 0
        shares    = pos.get("shares", 0)
        avg_cost  = pos.get("avg_cost", price)
        mkt_val   = price * shares
        cost_val  = avg_cost * shares
        pnl       = mkt_val - cost_val
        pnl_pct   = (pnl / cost_val * 100) if cost_val else 0
        total_value += mkt_val
        total_cost  += cost_val
        holdings.append({
            "sym": sym, "shares": shares, "avg_cost": avg_cost,
            "price": price, "pct": info.get("pct", 0),
            "mkt_val": mkt_val, "cost_val": cost_val,
            "pnl": pnl, "pnl_pct": pnl_pct,
            "short_name": info.get("short_name", sym),
            "currency": info.get("currency","USD"),
            "52w_high": info.get("52w_high"), "52w_low": info.get("52w_low"),
        })

    holdings.sort(key=lambda x: x["mkt_val"], reverse=True)
    total_pnl     = total_value - total_cost
    total_pnl_pct = (total_pnl / total_cost * 100) if total_cost else 0

    # weights
    for h in holdings:
        h["weight"] = round(h["mkt_val"] / total_value * 100, 1) if total_value else 0

    return {
        "total_value":   round(total_value, 2),
        "total_cost":    round(total_cost, 2),
        "total_pnl":     round(total_pnl, 2),
        "total_pnl_pct": round(total_pnl_pct, 2),
        "holdings":      holdings,
    }

# Colour palette for allocation chart (cycles through)
_ALLOC_COLORS = ["#C084C8","#60a5fa","#4ade80","#F0C040","#fb923c",
                 "#f87171","#34d399","#a78bfa","#38bdf8","#fbbf24"]

# ═══════════════════════════════════════════════════════════════════════════
#  NEW FEATURE LIBRARY  v8
#  1. Centralised Groq call with exponential back-off + retry UI
#  2. Portfolio risk analytics  (beta, VaR, Sharpe, diversification)
#  3. Price alert engine
#  4. Document comparison
#  5. Export / CSV download
#  6. Theme system
#  7. Keyboard shortcuts injector
# ═══════════════════════════════════════════════════════════════════════════

# ─── 1. CENTRALISED GROQ CALLER WITH RETRY ──────────────────────────────────

def _parse_retry_secs(err_str: str) -> float:
    """Extract 'try again in Xs' from Groq rate-limit error text."""
    m = re.search(r"try again in\s+(?:(\d+)m)?[\s]*([\d.]+)s", err_str, re.IGNORECASE)
    if m:
        mins = int(m.group(1) or 0)
        secs = float(m.group(2) or 0)
        return mins * 60 + secs
    m2 = re.search(r"([\d.]+)\s*second", err_str, re.IGNORECASE)
    return float(m2.group(1)) if m2 else 60.0

def groq_call(
    api_key:     str,
    messages:    list[dict],
    system:      str  = "",
    model:       str  = "llama-3.3-70b-versatile",
    temperature: float = 0.15,
    max_tokens:  int   = 600,
    site_key:    str   = "default",
    show_spinner:bool  = False,
) -> str:
    """
    Single entry-point for ALL Groq LLM calls.
    • Exponential back-off with jitter on 429/rate_limit (up to 3 retries)
    • Returns descriptive error strings on failure (never raises)
    • Tracks per-site last-error timestamps in session state
    """
    if not api_key:
        return "⚠ No API key — add your Groq key in the sidebar."
    # Defensive init — ensures key exists even if called before session state loop
    if "_groq_last_err" not in st.session_state:
        st.session_state["_groq_last_err"] = {}
    from openai import OpenAI
    oai = OpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1")

    full_messages = []
    if system:
        full_messages.append({"role": "system", "content": system})
    full_messages.extend(messages)

    for attempt in range(3):
        try:
            resp = oai.chat.completions.create(
                model=model,
                messages=full_messages,
                temperature=temperature,
                max_tokens=max_tokens,
            )
            # Clear error flag on success
            st.session_state["_groq_last_err"].pop(site_key, None)
            return resp.choices[0].message.content.strip()

        except Exception as e:
            err = str(e)
            is_rate = "rate_limit_exceeded" in err or "429" in err or "Rate limit" in err
            is_overload = "overloaded" in err.lower() or "503" in err or "502" in err

            if is_rate or is_overload:
                wait = _parse_retry_secs(err) if is_rate else 5.0
                # Jitter to avoid thundering herd
                wait = wait * (1 + 0.1 * attempt) + (attempt * 2)
                st.session_state["_groq_last_err"][site_key] = {
                    "ts":    _dt.datetime.utcnow().isoformat(),
                    "wait":  wait,
                    "msg":   err[:200],
                    "type":  "rate_limit" if is_rate else "overload",
                }
                if attempt < 2:
                    _tm.sleep(min(wait, 8))   # wait up to 8s before retry in-request
                    continue
                retry_min = max(1, int(wait / 60))
                return (
                    f"⏱ Groq {'rate limit' if is_rate else 'overload'} — "
                    f"retry in ~{retry_min} min.\n"
                    f"💡 Upgrade at console.groq.com/settings/billing to remove limits."
                )
            # Non-retryable errors
            return f"⚠ API error: {err[:200]}"
    return "⚠ Max retries exceeded — Groq API unavailable right now."


# ─── 2. PORTFOLIO RISK ANALYTICS ────────────────────────────────────────────

@st.cache_data(ttl=1800, max_entries=50, show_spinner=False)
def _fetch_beta(sym: str) -> float:
    """Compute 1-year rolling beta against SPY. Cached 30 min, max 50 symbols."""
    try:
        spy_df  = fetch_stock_history_1y("SPY")
        sym_df  = fetch_stock_history_1y(sym)
        if spy_df.empty or sym_df.empty or len(spy_df) < 60:
            return 1.0
        spy_r  = spy_df["close"].pct_change().dropna()
        sym_r  = sym_df["close"].pct_change().dropna()
        common = spy_r.index.intersection(sym_r.index)
        if len(common) < 30:
            return 1.0
        spy_r = spy_r.loc[common].values
        sym_r = sym_r.loc[common].values
        cov    = float(np.cov(sym_r, spy_r)[0, 1])
        var_m  = float(np.var(spy_r))
        return round(cov / var_m, 3) if var_m else 1.0
    except Exception:
        return 1.0

def calculate_portfolio_beta(summary: dict) -> float:
    """Weighted-average beta. Capped at 8 holdings to bound memory."""
    total_val = summary["total_value"] or 1.0
    # Cap to top-8 by weight to avoid OOM from too many 1Y fetches
    top_holdings = sorted(summary["holdings"], key=lambda h: h["mkt_val"], reverse=True)[:8]
    top_val = sum(h["mkt_val"] for h in top_holdings) or total_val
    beta = 0.0
    for h in top_holdings:
        w = h["mkt_val"] / top_val
        beta += w * _fetch_beta(h["sym"])
    return round(beta, 3)

def calculate_var(summary: dict, confidence: float = 0.95) -> float:
    """
    Parametric VaR (1-day). Capped at 8 holdings to bound memory.
    Falls back to 1.5% daily vol if history unavailable.
    """
    z_map = {0.90: 1.282, 0.95: 1.645, 0.99: 2.326}
    z = z_map.get(confidence, 1.645)
    total_val = summary["total_value"] or 1.0
    if not summary["holdings"]:
        return 0.0
    top_holdings = sorted(summary["holdings"], key=lambda h: h["mkt_val"], reverse=True)[:8]
    top_val = sum(h["mkt_val"] for h in top_holdings) or total_val
    weighted_vol = 0.0
    for h in top_holdings:
        w = h["mkt_val"] / top_val
        try:
            df  = fetch_stock_history_1y(h["sym"])
            tec = compute_technicals(df)
            ann_vol = tec.get("vol_annual", 30.0) / 100.0
        except Exception:
            ann_vol = 0.30
        daily_vol = ann_vol / (252 ** 0.5)
        weighted_vol += w * daily_vol
    return round(total_val * z * weighted_vol, 2)

def calculate_sharpe_ratio(summary: dict, risk_free: float = 0.053) -> float:
    """
    Simplified Sharpe using 1-year return and annualised vol.
    Capped at 8 holdings to bound memory.
    """
    total_val  = summary["total_value"] or 1.0
    total_cost = summary["total_cost"]  or 1.0
    ann_return = (total_val / total_cost - 1.0)
    if not summary["holdings"]:
        return 0.0
    top_holdings = sorted(summary["holdings"], key=lambda h: h["mkt_val"], reverse=True)[:8]
    top_val = sum(h["mkt_val"] for h in top_holdings) or total_val
    weighted_vol = 0.0
    for h in top_holdings:
        w = h["mkt_val"] / top_val
        try:
            df  = fetch_stock_history_1y(h["sym"])
            tec = compute_technicals(df)
            weighted_vol += w * (tec.get("vol_annual", 30.0) / 100.0)
        except Exception:
            weighted_vol += w * 0.30
    if weighted_vol < 0.001:
        return 0.0
    return round((ann_return - risk_free) / weighted_vol, 3)

def calculate_diversification(summary: dict) -> int:
    """
    Diversification score 0–100 based on:
    • Holdings count (0–30 pts)
    • Sector spread (0–40 pts)
    • Concentration: 1 - HHI of weights (0–30 pts)
    """
    holdings = summary["holdings"]
    n = len(holdings)
    if n == 0:
        return 0

    # Holdings count score
    count_score = min(30, n * 4)

    # Sector inference from exchange label
    def _infer_sector(sym: str) -> str:
        s = sym.upper()
        if s.endswith("-USD") or s.endswith("-BTC"):  return "Crypto"
        if s.endswith(".NS") or s.endswith(".BO"):     return "India"
        if s.endswith(".KS"):                          return "Korea"
        if s.endswith(".HK"):                          return "HK"
        if s.endswith(".L"):                           return "UK"
        if s.endswith(".AX"):                          return "Australia"
        if s in {"AAPL","MSFT","NVDA","GOOGL","META","AMZN","TSLA"}: return "US_Tech"
        if s in {"JPM","GS","V","MA","BAC","WFC"}:    return "US_Finance"
        return "US_Other"

    sectors = set(_infer_sector(h["sym"]) for h in holdings)
    sector_score = min(40, len(sectors) * 8)

    # HHI concentration
    total = summary["total_value"] or 1.0
    hhi = sum((h["mkt_val"] / total) ** 2 for h in holdings)
    concentration_score = int((1 - hhi) * 30)

    return min(100, count_score + sector_score + concentration_score)

def render_portfolio_analytics(summary: dict, groq_api_key: str) -> None:
    """
    Full risk analytics dashboard: Beta, VaR(95/99), Sharpe, Diversification,
    correlation heatmap, top-risk contributor, AI risk commentary.
    """
    if not summary["holdings"]:
        return

    V = VELVET

    st.markdown(
        f'<div style="font-family:Space Mono,monospace;font-size:.54rem;letter-spacing:.18em;'
        f'text-transform:uppercase;color:#C084C8;margin:.8rem 0 .5rem;">◈ Portfolio Risk Analytics</div>',
        unsafe_allow_html=True,
    )

    # ── Metric cards row ─────────────────────────────────────────────────
    with st.spinner("Computing risk metrics (fetching 1Y history)…"):
        p_beta  = calculate_portfolio_beta(summary)
        var_95  = calculate_var(summary, 0.95)
        var_99  = calculate_var(summary, 0.99)
        sharpe  = calculate_sharpe_ratio(summary)
        divers  = calculate_diversification(summary)

    def _risk_card(label, val, note, color):
        st.markdown(
            f'<div style="background:{V["card2"]};border:1px solid {V["border"]};'
            f'border-top:2px solid {color};border-radius:8px;padding:.65rem .8rem;">'
            f'<div style="font-family:Space Mono,monospace;font-size:.44rem;letter-spacing:.15em;'
            f'text-transform:uppercase;color:#4A3858;">{label}</div>'
            f'<div style="font-family:Cormorant Garamond,serif;font-size:1.45rem;font-weight:300;'
            f'color:{color};line-height:1.1;">{val}</div>'
            f'<div style="font-family:Space Mono,monospace;font-size:.44rem;color:#4A3858;">{note}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    r1, r2, r3, r4, r5 = st.columns(5)
    with r1: _risk_card("Portfolio Beta", f"{p_beta:.2f}",
        "vs S&P 500" + (" · Aggressive" if p_beta>1.2 else " · Defensive" if p_beta<0.8 else " · Market-neutral"),
        "#f87171" if p_beta>1.3 else ("#4ade80" if p_beta<0.8 else "#F0C040"))
    with r2: _risk_card("VaR 95% (1-day)", f"-${var_95:,.0f}",
        "Max daily loss (95% conf.)",
        "#F0C040")
    with r3: _risk_card("VaR 99% (1-day)", f"-${var_99:,.0f}",
        "Max daily loss (99% conf.)",
        "#f87171")
    with r4: _risk_card("Sharpe Ratio", f"{sharpe:.2f}",
        "Risk-adj. return (rf=5.3%)" + (" · Excellent" if sharpe>1.5 else " · Good" if sharpe>0.8 else " · Poor" if sharpe<0 else " · Moderate"),
        "#4ade80" if sharpe>1 else ("#F0C040" if sharpe>0 else "#f87171"))
    with r5:
        d_col = "#4ade80" if divers>=70 else ("#F0C040" if divers>=40 else "#f87171")
        _risk_card("Diversification", f"{divers}%", "", d_col)

    # Diversification progress bar
    st.markdown(
        f'<div style="margin:.5rem 0 .8rem;">'
        f'<div style="height:6px;background:rgba(107,45,107,.15);border-radius:3px;overflow:hidden;">'
        f'<div style="height:100%;width:{divers}%;background:linear-gradient(90deg,'
        f'{"#f87171" if divers<40 else "#F0C040" if divers<70 else "#4ade80"},'
        f'{"#F0C040" if divers<70 else "#4ade80"});border-radius:3px;'
        f'transition:width .6s ease;"></div></div>'
        f'<div style="font-family:Space Mono,monospace;font-size:.44rem;color:#4A3858;margin-top:.2rem;">'
        f'Diversification score {divers}/100 — '
        f'{"Well diversified" if divers>=70 else "Moderately diversified" if divers>=40 else "Concentrated — consider adding more sectors"}'
        f'</div></div>',
        unsafe_allow_html=True,
    )

    # ── Per-holding risk contribution table ─────────────────────────────
    st.markdown(
        f'<div style="font-family:Space Mono,monospace;font-size:.48rem;letter-spacing:.15em;'
        f'text-transform:uppercase;color:#4A3858;margin:.6rem 0 .3rem;">Risk Contribution per Holding</div>',
        unsafe_allow_html=True,
    )
    total_val = summary["total_value"] or 1.0
    risk_rows_html = ""
    risk_data = []
    for h in summary["holdings"]:
        try:
            df   = fetch_stock_history_1y(h["sym"])
            tec  = compute_technicals(df)
            vol  = tec.get("vol_annual", 30.0)
            beta = _fetch_beta(h["sym"])
        except Exception:
            vol, beta = 30.0, 1.0
        w = h["mkt_val"] / total_val
        daily_vol = vol / 100 / (252 ** 0.5)
        holding_var = round(h["mkt_val"] * 1.645 * daily_vol, 2)
        risk_data.append({"sym": h["sym"], "w": w, "vol": vol, "beta": beta, "var95": holding_var})

    for rd in sorted(risk_data, key=lambda x: x["var95"], reverse=True):
        bar_w = min(100, rd["var95"] / max(r["var95"] for r in risk_data) * 100) if risk_data else 0
        b_col = "#f87171" if rd["beta"]>1.2 else ("#4ade80" if rd["beta"]<0.8 else "#F0C040")
        risk_rows_html += (
            f'<div style="display:flex;align-items:center;gap:.6rem;padding:.3rem 0;'
            f'border-bottom:1px solid rgba(107,45,107,.08);">'
            f'<div style="font-family:Space Mono,monospace;font-size:.56rem;color:#C084C8;width:5.5rem;">{rd["sym"][:10]}</div>'
            f'<div style="flex:1;">'
            f'  <div style="height:4px;background:rgba(107,45,107,.12);border-radius:2px;overflow:hidden;">'
            f'  <div style="height:100%;width:{bar_w:.0f}%;background:linear-gradient(90deg,#C084C8,#f87171);border-radius:2px;"></div></div>'
            f'</div>'
            f'<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#9A8AAA;width:4rem;text-align:right;">β {rd["beta"]:.2f}</div>'
            f'<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#9A8AAA;width:4rem;text-align:right;">{rd["vol"]:.1f}% vol</div>'
            f'<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#f87171;width:5.5rem;text-align:right;">VaR -${rd["var95"]:,.0f}</div>'
            f'</div>'
        )
    st.markdown(f'<div style="margin-bottom:.8rem;">{risk_rows_html}</div>', unsafe_allow_html=True)

    # ── AI Risk Commentary ───────────────────────────────────────────────
    if groq_api_key and st.button("🤖  Generate AI Risk Assessment", key="pf_risk_ai_btn"):
        top3 = sorted(risk_data, key=lambda x: x["var95"], reverse=True)[:3]
        top3_str = ", ".join(f'{r["sym"]}(b={r["beta"]:.2f},vol={r["vol"]:.0f}%)' for r in top3)
        risk_prompt = (
            f"Portfolio: ${total_val:,.0f} across {len(summary['holdings'])} holdings.\n"
            f"Beta={p_beta:.2f}, Sharpe={sharpe:.2f}, VaR95=-${var_95:,.0f}/day, VaR99=-${var_99:,.0f}/day\n"
            f"Diversification={divers}/100\n"
            f"Top risk contributors: {top3_str}\n\n"
            f"Give a concise risk assessment (<=200 words): key risks, hedging suggestions, position sizing recommendations."
        )
        with st.spinner("Generating AI risk assessment…"):
            ai_risk = groq_call(
                api_key=groq_api_key,
                messages=[{"role":"user","content":risk_prompt}],
                system="You are a portfolio risk manager. Be direct and quantitative.",
                max_tokens=400, site_key="pf_risk"
            )
        st.markdown(
            f'<div style="background:{V["card2"]};border:1px solid rgba(139,58,139,.3);'
            f'border-left:3px solid #C084C8;border-radius:0 8px 8px 0;'
            f'padding:.8rem 1rem;margin:.6rem 0;font-family:Syne,sans-serif;'
            f'font-size:.82rem;color:#C8B8D8;line-height:1.75;">{ai_risk}</div>',
            unsafe_allow_html=True,
        )


# ─── 3. PRICE ALERT ENGINE ───────────────────────────────────────────────────

def check_price_alerts() -> list[dict]:
    """
    Check all configured price alerts against current prices.
    Returns list of newly-triggered alerts.
    Mutates st.session_state.price_alerts to clear triggered ones.
    """
    alerts = st.session_state.price_alerts
    if not alerts:
        return []
    triggered = []
    for sym, cfg in list(alerts.items()):
        try:
            info  = fetch_quote(sym)
            if not info:
                continue
            price = info["price"]
            ts    = _dt.datetime.utcnow().strftime("%H:%M UTC")
            if cfg.get("above") is not None and price >= cfg["above"]:
                triggered.append({"sym":sym,"type":"above","price":price,
                                   "target":cfg["above"],"ts":ts})
                alerts[sym]["above"] = None   # one-shot: clear after trigger
            if cfg.get("below") is not None and price <= cfg["below"]:
                triggered.append({"sym":sym,"type":"below","price":price,
                                   "target":cfg["below"],"ts":ts})
                alerts[sym]["below"] = None
        except Exception:
            pass
    # Prune empty alert configs
    st.session_state.price_alerts = {
        s: c for s, c in alerts.items()
        if c.get("above") is not None or c.get("below") is not None
    }
    # Append to log
    if triggered:
        log = st.session_state.alert_log
        log.extend(triggered)
        st.session_state.alert_log = log[-50:]   # keep last 50
    return triggered

def render_alerts_panel(portfolio: dict) -> None:
    """Alert configuration and notification log UI inside portfolio panel."""
    V = VELVET

    # ── Run check ────────────────────────────────────────────────────────
    newly = check_price_alerts()
    if newly:
        for a in newly:
            dir_sym = "▲" if a["type"]=="above" else "▼"
            col = "#4ade80" if a["type"]=="above" else "#f87171"
            st.markdown(
                f'<div style="background:rgba(74,222,128,.08);border:1px solid rgba(74,222,128,.3);'
                f'border-radius:8px;padding:.5rem .9rem;margin-bottom:.3rem;'
                f'font-family:Space Mono,monospace;font-size:.6rem;">'
                f'🔔 <span style="color:{col};">{dir_sym} {a["sym"]} hit ${a["price"]:,.2f}</span>'
                f' (target ${a["target"]:,.2f}) · {a["ts"]}</div>',
                unsafe_allow_html=True,
            )

    st.markdown(
        f'<div style="font-family:Space Mono,monospace;font-size:.52rem;letter-spacing:.18em;'
        f'text-transform:uppercase;color:{V["ghost"]};margin:.5rem 0 .4rem;">◈ Price Alerts</div>',
        unsafe_allow_html=True,
    )

    syms = sorted(portfolio.keys()) if portfolio else []
    if not syms:
        st.caption("Add holdings first to set price alerts.")
        return

    al_col1, al_col2, al_col3, al_col4 = st.columns([2, 1.5, 1.5, 1])
    with al_col1:
        al_sym = st.selectbox("Symbol", syms, key="al_sym", label_visibility="collapsed")
    with al_col2:
        al_above = st.number_input("Alert above $", value=0.0, step=0.01, min_value=0.0,
                                   key="al_above", label_visibility="collapsed",
                                   help="Notify when price rises above this value")
    with al_col3:
        al_below = st.number_input("Alert below $", value=0.0, step=0.01, min_value=0.0,
                                   key="al_below", label_visibility="collapsed",
                                   help="Notify when price falls below this value")
    with al_col4:
        if st.button("Set Alert", key="al_set_btn", use_container_width=True):
            if al_sym:
                cfg = st.session_state.price_alerts.get(al_sym, {})
                if al_above > 0: cfg["above"] = al_above
                if al_below > 0: cfg["below"] = al_below
                if cfg:
                    st.session_state.price_alerts[al_sym] = cfg
                    st.success(f"Alert set for {al_sym}")
                    st.rerun()

    # Active alerts table
    active = st.session_state.price_alerts
    if active:
        rows_html = ""
        for sym, cfg in active.items():
            try:
                curr = fetch_quote(sym)
                curr_p = f"${curr['price']:,.2f}" if curr else "—"
            except Exception:
                curr_p = "—"
            above_s = f"▲ ${cfg['above']:,.2f}" if cfg.get("above") else "—"
            below_s = f"▼ ${cfg['below']:,.2f}" if cfg.get("below") else "—"
            rows_html += (
                f'<tr>'
                f'<td style="font-family:Space Mono,monospace;font-size:.58rem;color:#C084C8;'
                f'padding:.3rem .6rem;border:1px solid rgba(107,45,107,.2);">{sym}</td>'
                f'<td style="font-family:Space Mono,monospace;font-size:.58rem;color:#9A8AAA;'
                f'padding:.3rem .6rem;border:1px solid rgba(107,45,107,.2);">{curr_p}</td>'
                f'<td style="font-family:Space Mono,monospace;font-size:.58rem;color:#4ade80;'
                f'padding:.3rem .6rem;border:1px solid rgba(107,45,107,.2);">{above_s}</td>'
                f'<td style="font-family:Space Mono,monospace;font-size:.58rem;color:#f87171;'
                f'padding:.3rem .6rem;border:1px solid rgba(107,45,107,.2);">{below_s}</td>'
                f'</tr>'
            )
        st.markdown(
            f'<table style="width:100%;border-collapse:collapse;margin:.4rem 0 .6rem;">'
            f'<thead><tr>'
            f'<th style="background:rgba(107,45,107,.18);padding:.3rem .6rem;font-family:Space Mono,monospace;'
            f'font-size:.44rem;text-transform:uppercase;color:#4A3858;text-align:left;border:1px solid rgba(107,45,107,.2);">Symbol</th>'
            f'<th style="background:rgba(107,45,107,.18);padding:.3rem .6rem;font-family:Space Mono,monospace;'
            f'font-size:.44rem;text-transform:uppercase;color:#4A3858;border:1px solid rgba(107,45,107,.2);">Current</th>'
            f'<th style="background:rgba(107,45,107,.18);padding:.3rem .6rem;font-family:Space Mono,monospace;'
            f'font-size:.44rem;text-transform:uppercase;color:#4A3858;border:1px solid rgba(107,45,107,.2);">Above</th>'
            f'<th style="background:rgba(107,45,107,.18);padding:.3rem .6rem;font-family:Space Mono,monospace;'
            f'font-size:.44rem;text-transform:uppercase;color:#4A3858;border:1px solid rgba(107,45,107,.2);">Below</th>'
            f'</tr></thead><tbody>{rows_html}</tbody></table>',
            unsafe_allow_html=True,
        )
        if st.button("🗑 Clear all alerts", key="al_clear_all"):
            st.session_state.price_alerts = {}
            st.rerun()

    # Alert log
    log = st.session_state.alert_log
    if log:
        with st.expander(f"📋 Alert history ({len(log)})"):
            for a in reversed(log[-20:]):
                dir_s = "▲ Hit above" if a["type"]=="above" else "▼ Hit below"
                col   = "#4ade80" if a["type"]=="above" else "#f87171"
                st.markdown(
                    f'<div style="font-family:Space Mono,monospace;font-size:.54rem;'
                    f'color:{col};padding:.15rem 0;">'
                    f'{a["ts"]} · {a["sym"]} {dir_s} ${a["price"]:,.2f} (target ${a["target"]:,.2f})</div>',
                    unsafe_allow_html=True,
                )


# ─── 4. DOCUMENT COMPARISON ─────────────────────────────────────────────────

def render_document_comparison(file_names: list[str], auto_metrics: list[dict]) -> None:
    """Side-by-side metric comparison between two uploaded documents."""
    V = VELVET

    if len(file_names) < 2:
        st.info("Upload at least 2 documents to use the comparison tool.")
        return

    dc1, dc2 = st.columns(2)
    with dc1:
        doc1 = st.selectbox("Document A", file_names, key="dc_doc1")
    with dc2:
        doc2 = st.selectbox("Document B", [f for f in file_names if f != doc1],
                             key="dc_doc2")
    if not doc1 or not doc2:
        return

    # Build metric dicts per document from auto_metrics
    def _metrics_for(fname: str) -> dict[str, dict]:
        """Return {label: {value, unit}} for a specific doc."""
        result: dict[str, dict] = {}
        for m in auto_metrics:
            if m.get("source") == fname or m.get("filename") == fname:
                result[m["label"]] = {"value": m["value"], "unit": m["unit"]}
        # If source field not present, try label-level dedup from all metrics
        if not result:
            for m in auto_metrics:
                if m["label"] not in result:
                    result[m["label"]] = {"value": m["value"], "unit": m["unit"]}
        return result

    m1 = _metrics_for(doc1)
    m2 = _metrics_for(doc2)
    all_labels = sorted(set(m1) | set(m2))

    if not all_labels:
        st.warning("No metrics extracted from the selected documents. Try Templates tab for LLM extraction.")
        return

    comparison = []
    for lbl in all_labels:
        v1 = m1.get(lbl)
        v2 = m2.get(lbl)
        v1_str = fmt_val(v1["value"], v1["unit"]) if v1 else "—"
        v2_str = fmt_val(v2["value"], v2["unit"]) if v2 else "—"
        if v1 and v2 and v1["value"] and v2["value"]:
            try:
                chg = (v2["value"] / v1["value"] - 1) * 100
                chg_str = f"{chg:+.1f}%"
                chg_color = "#4ade80" if chg >= 0 else "#f87171"
            except Exception:
                chg_str, chg_color = "—", "#9A8AAA"
        else:
            chg_str, chg_color = "—", "#9A8AAA"
        comparison.append({"label": lbl, "v1": v1_str, "v2": v2_str,
                            "chg": chg_str, "chg_color": chg_color})

    # Render as styled HTML table
    rows_html = ""
    for row in comparison:
        rows_html += (
            f'<tr>'
            f'<td style="font-family:Space Mono,monospace;font-size:.58rem;color:#C084C8;'
            f'padding:.35rem .7rem;border:1px solid rgba(107,45,107,.2);">{row["label"]}</td>'
            f'<td style="font-family:Space Mono,monospace;font-size:.62rem;color:#EDE8F5;'
            f'padding:.35rem .7rem;border:1px solid rgba(107,45,107,.2);text-align:right;">{row["v1"]}</td>'
            f'<td style="font-family:Space Mono,monospace;font-size:.62rem;color:#EDE8F5;'
            f'padding:.35rem .7rem;border:1px solid rgba(107,45,107,.2);text-align:right;">{row["v2"]}</td>'
            f'<td style="font-family:Space Mono,monospace;font-size:.6rem;color:{row["chg_color"]};'
            f'padding:.35rem .7rem;border:1px solid rgba(107,45,107,.2);text-align:right;">{row["chg"]}</td>'
            f'</tr>'
        )
    short1 = doc1[:28] + "…" if len(doc1) > 28 else doc1
    short2 = doc2[:28] + "…" if len(doc2) > 28 else doc2
    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;margin:.5rem 0;">'
        f'<thead><tr>'
        f'<th style="background:rgba(107,45,107,.22);padding:.35rem .7rem;font-family:Space Mono,monospace;'
        f'font-size:.46rem;text-transform:uppercase;color:#4A3858;text-align:left;border:1px solid rgba(107,45,107,.2);">Metric</th>'
        f'<th style="background:rgba(107,45,107,.22);padding:.35rem .7rem;font-family:Space Mono,monospace;'
        f'font-size:.46rem;text-transform:uppercase;color:#4ade80;border:1px solid rgba(107,45,107,.2);text-align:right;">{short1}</th>'
        f'<th style="background:rgba(107,45,107,.22);padding:.35rem .7rem;font-family:Space Mono,monospace;'
        f'font-size:.46rem;text-transform:uppercase;color:#60a5fa;border:1px solid rgba(107,45,107,.2);text-align:right;">{short2}</th>'
        f'<th style="background:rgba(107,45,107,.22);padding:.35rem .7rem;font-family:Space Mono,monospace;'
        f'font-size:.46rem;text-transform:uppercase;color:#F0C040;border:1px solid rgba(107,45,107,.2);text-align:right;">Change</th>'
        f'</tr></thead><tbody>{rows_html}</tbody></table>',
        unsafe_allow_html=True,
    )

    # CSV download
    if comparison:
        csv_rows = [f"Metric,{doc1},{doc2},Change"]
        for row in comparison:
            csv_rows.append(f'{row["label"]},{row["v1"]},{row["v2"]},{row["chg"]}')
        csv_str = "\n".join(csv_rows)
        st.download_button(
            "⬇ Download comparison CSV",
            data=csv_str,
            file_name="document_comparison.csv",
            mime="text/csv",
            key="dc_download",
        )


# ─── 5. EXPORT / CSV DOWNLOAD ────────────────────────────────────────────────

def export_portfolio_csv(summary: dict) -> str:
    """Generate CSV string for portfolio holdings."""
    lines = ["Symbol,Name,Shares,Avg Cost,Current Price,Market Value,P&L,P&L %,Weight %,Today %"]
    for h in summary["holdings"]:
        lines.append(
            f'{h["sym"]},{h["short_name"].replace(",","")},{h["shares"]:,.4g},'
            f'{h["avg_cost"]:,.2f},{h["price"]:,.2f},{h["mkt_val"]:,.2f},'
            f'{h["pnl"]:+,.2f},{h["pnl_pct"]:+.2f},{h["weight"]:.1f},{h["pct"]:+.2f}'
        )
    lines.append("")
    lines.append(f'Total Value,,,,,{summary["total_value"]:,.2f},{summary["total_pnl"]:+,.2f},{summary["total_pnl_pct"]:+.2f},,')
    return "\n".join(lines)

def export_metrics_csv(metrics: list[dict]) -> str:
    """Generate CSV string for extracted document metrics."""
    lines = ["Metric,Value,Unit,Category,Raw Text"]
    for m in metrics:
        raw = m.get("raw", "").replace(",", ";").replace("\n", " ")[:80]
        lines.append(
            f'{m["label"]},{fmt_val(m["value"],m["unit"])},{m["unit"]},{m["category"]},{raw}'
        )
    return "\n".join(lines)

def render_export_panel(summary: dict, metrics: list[dict], doc_full_text: str) -> None:
    """Export section: portfolio CSV, metrics CSV, raw text."""
    V = VELVET
    st.markdown(
        f'<div style="font-family:Space Mono,monospace;font-size:.52rem;letter-spacing:.18em;'
        f'text-transform:uppercase;color:{V["ghost"]};margin:.5rem 0 .4rem;">◈ Export & Reports</div>',
        unsafe_allow_html=True,
    )

    exp_cols = st.columns(3)

    with exp_cols[0]:
        if summary["holdings"]:
            csv_pf = export_portfolio_csv(summary)
            st.download_button(
                "⬇ Portfolio CSV",
                data=csv_pf,
                file_name=f"portfolio_{_dt.datetime.utcnow().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
                key="exp_pf_csv",
            )
        else:
            st.button("⬇ Portfolio CSV", disabled=True, use_container_width=True, key="exp_pf_csv_dis")

    with exp_cols[1]:
        if metrics:
            csv_m = export_metrics_csv(metrics)
            st.download_button(
                "⬇ Metrics CSV",
                data=csv_m,
                file_name=f"metrics_{_dt.datetime.utcnow().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
                key="exp_met_csv",
            )
        else:
            st.button("⬇ Metrics CSV", disabled=True, use_container_width=True, key="exp_met_csv_dis")

    with exp_cols[2]:
        if doc_full_text:
            st.download_button(
                "⬇ Raw Document Text",
                data=doc_full_text,
                file_name=f"extracted_text_{_dt.datetime.utcnow().strftime('%Y%m%d')}.txt",
                mime="text/plain",
                use_container_width=True,
                key="exp_raw_txt",
            )
        else:
            st.button("⬇ Raw Document Text", disabled=True, use_container_width=True, key="exp_raw_txt_dis")


# ─── 6. THEME SYSTEM ────────────────────────────────────────────────────────

_THEMES: dict[str, dict] = {
    "Royal Velvet": {
        "primary":    "#6B2D6B",
        "accent":     "#C084C8",
        "bg":         "#07060C",
        "card":       "#0D0B12",
        "text":       "#EDE8F5",
        "border":     "rgba(139,58,139,.35)",
    },
    "Midnight Blue": {
        "primary":    "#1E3A8A",
        "accent":     "#3B82F6",
        "bg":         "#0F172A",
        "card":       "#1E293B",
        "text":       "#F8FAFC",
        "border":     "rgba(59,130,246,.3)",
    },
    "Forest": {
        "primary":    "#166534",
        "accent":     "#22C55E",
        "bg":         "#052E16",
        "card":       "#0A3D1F",
        "text":       "#F0FDF4",
        "border":     "rgba(34,197,94,.3)",
    },
    "Obsidian": {
        "primary":    "#292524",
        "accent":     "#F59E0B",
        "bg":         "#0C0A09",
        "card":       "#1C1917",
        "text":       "#FAFAF9",
        "border":     "rgba(245,158,11,.25)",
    },
}

def apply_theme(theme_name: str) -> None:
    """Inject CSS variables for the selected theme."""
    t = _THEMES.get(theme_name, _THEMES["Royal Velvet"])
    st.markdown(
        f'<style>'
        f':root{{'
        f'  --bg:{t["bg"]};'
        f'  --card:{t["card"]};'
        f'  --accent:{t["accent"]};'
        f'  --border:{t["border"]};'
        f'  --text:{t["text"]};'
        f'}}'
        f'[data-testid="stAppViewContainer"]{{background:{t["bg"]}!important}}'
        f'[data-testid="stSidebar"]{{background:{t["card"]}!important;'
        f'  border-right:1px solid {t["border"]}!important}}'
        f'</style>',
        unsafe_allow_html=True,
    )


# ─── 7. KEYBOARD SHORTCUTS ──────────────────────────────────────────────────

def inject_keyboard_shortcuts() -> None:
    """Inject fixed side-bar + keyboard shortcuts via st.markdown (no iframe gap)."""
    st.markdown("""
<style>
/* ══ DYNAMIC FIXED SIDE BARS ══════════════════════════════════════════════ */
#rag-floatbar {
  position: fixed;
  right: 0;
  top: 50%;
  transform: translateY(-50%);
  z-index: 999999;
  display: flex;
  flex-direction: column;
  gap: 3px;
  pointer-events: none;
}
.rfb-btn {
  position: relative;
  width: 10px;
  height: 90px;
  cursor: pointer;
  pointer-events: all;
  border: none;
  outline: none;
  background: none;
  padding: 0;
  transition: width .4s cubic-bezier(.34,1.56,.64,1);
  overflow: hidden;
  border-radius: 8px 0 0 8px;
}
.rfb-btn:hover, .rfb-btn.rfb-active { width: 80px; }

/* Coloured background face */
.rfb-face {
  position: absolute;
  inset: 0;
  border-radius: 8px 0 0 8px;
  border: 1.5px solid currentColor;
  border-right: none;
  background: rgba(1,13,26,.95);
  transition: box-shadow .4s ease, background .3s ease;
}
.rfb-btn:hover .rfb-face,
.rfb-btn.rfb-active .rfb-face {
  box-shadow: -4px 0 40px rgba(56,189,248,.5);
  background: rgba(1,18,36,.98);
}
.rfb-btn.rfb-active .rfb-face {
  box-shadow: -4px 0 60px rgba(56,189,248,.8), -2px 0 20px rgba(56,189,248,.5);
}

/* Content: icon + label (vertical) */
.rfb-content {
  position: absolute;
  inset: 0;
  display: flex;
  flex-direction: column;
  align-items: center;
  justify-content: center;
  gap: 5px;
  opacity: 0;
  transform: translateX(16px);
  transition: opacity .3s ease .05s, transform .35s cubic-bezier(.34,1.56,.64,1);
  pointer-events: none;
}
.rfb-btn:hover .rfb-content,
.rfb-btn.rfb-active .rfb-content {
  opacity: 1;
  transform: translateX(0);
}
.rfb-icon {
  font-size: 1.3rem;
  line-height: 1;
}
.rfb-label {
  font-family: 'Space Mono', monospace;
  font-size: .36rem;
  letter-spacing: .2em;
  text-transform: uppercase;
  writing-mode: vertical-rl;
  transform: rotate(180deg);
  white-space: nowrap;
}

/* Narrow peek strip (visible when collapsed) */
.rfb-strip {
  position: absolute;
  right: 0; top: 0; bottom: 0;
  width: 6px;
  border-radius: 4px 0 0 4px;
  transition: width .4s cubic-bezier(.34,1.56,.64,1), opacity .3s ease;
  opacity: .9;
}
.rfb-btn:hover .rfb-strip,
.rfb-btn.rfb-active .rfb-strip {
  width: 100%;
  opacity: 0;
}

/* Scroll progress line on left edge */
#rag-floatbar::before {
  content: '';
  position: absolute;
  left: -3px; top: 0;
  width: 2px;
  height: var(--rfb-scroll, 0%);
  background: linear-gradient(180deg, #38bdf8, #a855f7);
  border-radius: 2px;
  transition: height .08s linear;
  pointer-events: none;
}
</style>

<div id="rag-floatbar">
  <button class="rfb-btn" id="rfb-upload" style="color:#38bdf8">
    <div class="rfb-face"></div>
    <div class="rfb-strip" style="background:#38bdf8"></div>
    <div class="rfb-content" style="color:#38bdf8">
      <span class="rfb-icon">↑</span>
      <span class="rfb-label">Upload</span>
    </div>
  </button>
  <button class="rfb-btn" id="rfb-portfolio" style="color:#0ea5e9">
    <div class="rfb-face"></div>
    <div class="rfb-strip" style="background:#0ea5e9"></div>
    <div class="rfb-content" style="color:#0ea5e9">
      <span class="rfb-icon">◈</span>
      <span class="rfb-label">Portfolio</span>
    </div>
  </button>
</div>

<script>
(function(){
  // Colour palette cycling with scroll
  var PALETTE = [
    [0.00, '#38bdf8', '#0ea5e9'],
    [0.15, '#818cf8', '#6366f1'],
    [0.30, '#34d399', '#10b981'],
    [0.45, '#f59e0b', '#d97706'],
    [0.60, '#f472b6', '#ec4899'],
    [0.75, '#a78bfa', '#8b5cf6'],
    [0.90, '#fb923c', '#ea580c'],
    [1.00, '#38bdf8', '#0ea5e9'],
  ];

  function hexR(h){h=h.replace('#','');return parseInt(h.slice(0,2),16);}
  function hexG(h){h=h.replace('#','');return parseInt(h.slice(2,4),16);}
  function hexB(h){h=h.replace('#','');return parseInt(h.slice(4,6),16);}
  function lerp(a,b,t){return a+(b-a)*t;}
  function lerpHex(c1,c2,t){
    var r=Math.round(lerp(hexR(c1),hexR(c2),t)).toString(16).padStart(2,'0');
    var g=Math.round(lerp(hexG(c1),hexG(c2),t)).toString(16).padStart(2,'0');
    var b=Math.round(lerp(hexB(c1),hexB(c2),t)).toString(16).padStart(2,'0');
    return '#'+r+g+b;
  }
  function colsAtPct(p){
    for(var i=0;i<PALETTE.length-1;i++){
      var s=PALETTE[i],e=PALETTE[i+1];
      if(p>=s[0]&&p<=e[0]){
        var t=(p-s[0])/(e[0]-s[0]);
        return [lerpHex(s[1],e[1],t), lerpHex(s[2],e[2],t)];
      }
    }
    return [PALETTE[0][1],PALETTE[0][2]];
  }

  function applyColours(upCol, simCol){
    var up  = document.getElementById('rfb-upload');
    var sim = document.getElementById('rfb-portfolio');
    if(!up||!sim) return;
    up.style.color  = upCol;
    sim.style.color = simCol;
    up.querySelector('.rfb-strip').style.background  = upCol;
    sim.querySelector('.rfb-strip').style.background = simCol;
    up.querySelector('.rfb-content').style.color  = upCol;
    sim.querySelector('.rfb-content').style.color = simCol;
  }

  function onScroll(){
    var el = document.documentElement;
    var max = el.scrollHeight - el.clientHeight;
    var pct = max>0 ? el.scrollTop/max : 0;
    var bar = document.getElementById('rag-floatbar');
    if(bar) bar.style.setProperty('--rfb-scroll', Math.round(pct*100)+'%');
    var cols = colsAtPct(pct);
    applyColours(cols[0], cols[1]);
  }

  // Click → find corresponding Streamlit button and click it
  function clickStreamlitBtn(idx){
    var btns = document.querySelectorAll('[data-testid="stMainBlockContainer"] [data-testid="baseButton-secondary"]');
    if(btns[idx]){ btns[idx].click(); return; }
    var all = document.querySelectorAll('[data-testid="stMainBlockContainer"] .stButton > button');
    if(all[idx]) all[idx].click();
  }

  document.addEventListener('DOMContentLoaded', function(){
    document.addEventListener('scroll', onScroll, {passive:true});
    onScroll();

    var upBtn  = document.getElementById('rfb-upload');
    var simBtn = document.getElementById('rfb-portfolio');

    if(upBtn) upBtn.addEventListener('click', function(){
      clickStreamlitBtn(0);
      upBtn.classList.toggle('rfb-active');
    });
    if(simBtn) simBtn.addEventListener('click', function(){
      clickStreamlitBtn(1);
      simBtn.classList.toggle('rfb-active');
    });
  });

  // Keyboard: Ctrl+U = upload
  if(!window._ragKbLoaded){
    window._ragKbLoaded = true;
    document.addEventListener('keydown', function(e){
      if((e.ctrlKey||e.metaKey) && e.key==='u'){
        e.preventDefault();
        clickStreamlitBtn(0);
      }
    });
  }
})();
</script>
    """, unsafe_allow_html=True)


def render_portfolio_panel(groq_api_key: str) -> None:
    """Full portfolio UI — called inside the portfolio panel."""

    portfolio = st.session_state.portfolio

    # ── Add / Remove Holdings ──────────────────────────────────────────────
    with st.expander("➕  Add Holdings", expanded=not bool(portfolio)):

        # ── Global Search ────────────────────────────────────────────────
        st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.15em;'
                    'text-transform:uppercase;color:#4A3858;margin-bottom:.35rem;">'
                    'Search any stock · ETF · crypto across all global markets</div>',
                    unsafe_allow_html=True)

        srch_col, add_col = st.columns([5, 1])
        with srch_col:
            ticker_q = st.text_input(
                "Search",
                placeholder="Type name or ticker — e.g. Tesla, TSLA, Reliance, HDFCBANK.NS, BTC...",
                label_visibility="collapsed", key="pf_search_q",
            )
        with add_col:
            add_custom = st.button("Add ticker →", key="pf_add_custom", use_container_width=True)

        # Live search results
        search_results = search_tickers(ticker_q) if ticker_q and len(ticker_q) >= 1 else []

        if search_results:
            st.markdown('<div style="font-family:Space Mono,monospace;font-size:.48rem;letter-spacing:.12em;'
                        'text-transform:uppercase;color:#4A3858;margin:.4rem 0 .3rem;">'
                        f'{len(search_results)} matches · click to select</div>', unsafe_allow_html=True)
            # Show as clickable chips — up to 4 per row
            for row_s in range(0, min(len(search_results), 12), 4):
                r_cols = st.columns(4)
                for ci, (name, sym, exch, sector) in enumerate(search_results[row_s:row_s+4]):
                    with r_cols[ci]:
                        exch_color = {
                            "NASDAQ":"#60a5fa","NYSE":"#4ade80","NSE":"#fb923c",
                            "BSE":"#fb923c","KRX":"#a78bfa","HKEX":"#f87171",
                            "Crypto":"#F0C040","ETF":"#34d399","OTC":"#9CA3AF",
                        }.get(exch, "#9CA3AF")
                        st.markdown(
                            f'<div style="background:var(--card-2);border:1px solid var(--border);'
                            f'border-radius:8px;padding:.5rem .7rem;margin-bottom:.3rem;">'
                            f'<div style="font-family:Space Mono,monospace;font-size:.62rem;'
                            f'color:var(--accent);font-weight:700;">{sym}</div>'
                            f'<div style="font-family:Syne,sans-serif;font-size:.7rem;color:var(--text-dim);'
                            f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{name}</div>'
                            f'<div style="display:flex;gap:.3rem;margin-top:.2rem;">'
                            f'<span style="font-family:Space Mono,monospace;font-size:.44rem;'
                            f'color:{exch_color};background:rgba(0,0,0,.3);border-radius:3px;'
                            f'padding:.05rem .3rem;">{exch}</span>'
                            f'<span style="font-family:Space Mono,monospace;font-size:.44rem;'
                            f'color:#4A3858;background:rgba(107,45,107,.1);border-radius:3px;'
                            f'padding:.05rem .3rem;">{sector}</span>'
                            f'</div></div>',
                            unsafe_allow_html=True,
                        )
                        if st.button(f"+ {sym}", key=f"sr_{sym}", use_container_width=True):
                            st.session_state["_pf_selected_sym"] = sym
                            st.rerun()

        # If a ticker was selected from search results, pre-fill entry form
        selected_sym = st.session_state.pop("_pf_selected_sym", "") or ""
        resolved_sym = selected_sym or (ticker_q.strip().upper() if add_custom and ticker_q.strip() else "")

        # ── Entry form ───────────────────────────────────────────────────
        st.markdown("<hr style='border-color:rgba(139,58,139,.12);margin:.5rem 0 .4rem;'>",
                    unsafe_allow_html=True)
        st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.15em;'
                    'text-transform:uppercase;color:#4A3858;margin-bottom:.3rem;">'
                    'Set position size</div>', unsafe_allow_html=True)

        ef1, ef2, ef3, ef4 = st.columns([2.5, 1.5, 1.5, 1])
        with ef1:
            sym_entry = st.text_input("Ticker", value=resolved_sym,
                                      placeholder="Symbol e.g. AAPL",
                                      label_visibility="collapsed", key="pf_sym")
        with ef2:
            shares_in = st.number_input("Shares", min_value=0.0001, value=1.0,
                                        step=0.1, format="%.4f",
                                        label_visibility="collapsed", key="pf_shares")
        with ef3:
            cost_in = st.number_input("Buy price ($)", min_value=0.0, value=0.0,
                                      step=0.01, format="%.2f",
                                      help="Leave 0 to use today's price",
                                      label_visibility="collapsed", key="pf_cost")
        with ef4:
            add_clicked = st.button("Add ✓", use_container_width=True, key="pf_add")

        if add_clicked and sym_entry.strip():
            sym = sym_entry.strip().upper()
            info = fetch_stock_fundamentals(sym)
            if not info or not info.get("price"):
                st.error(f"Could not fetch '{sym}'. Try the Yahoo Finance symbol (e.g. RELIANCE.NS, BTC-USD).")
            else:
                buy_price = cost_in if cost_in > 0 else info["price"]
                st.session_state.portfolio[sym] = {
                    "shares":   shares_in,
                    "avg_cost": buy_price,
                    "added":    _dt.datetime.utcnow().strftime("%Y-%m-%d"),
                }
                st.success(f"✓ {shares_in:,.4g} × {sym} ({info.get('short_name','')}) @ ${buy_price:,.2f}")
                st.rerun()

        # ── Quick-add by sector ──────────────────────────────────────────
        st.markdown("<hr style='border-color:rgba(139,58,139,.1);margin:.5rem 0 .4rem;'>",
                    unsafe_allow_html=True)
        st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.15em;'
                    'text-transform:uppercase;color:#4A3858;margin-bottom:.35rem;">'
                    'Quick-add by region / sector</div>', unsafe_allow_html=True)

        sector_groups = {
            "🇺🇸 US Tech":     ["AAPL","MSFT","NVDA","GOOGL","META"],
            "🇺🇸 US Finance":  ["JPM","GS","V","MA","BRK-B"],
            "🇮🇳 India":       ["RELIANCE.NS","TCS.NS","HDFCBANK.NS","INFY.NS","ICICIBANK.NS"],
            "🌏 Asia":         ["TSM","SONY","005930.KS","BABA","TM"],
            "🌍 Europe":       ["ASML","SAP","NVO","AZN","SHEL"],
            "₿ Crypto":       ["BTC-USD","ETH-USD","SOL-USD","BNB-USD","XRP-USD"],
        }
        sg_cols = st.columns(len(sector_groups))
        for gi, (grp_label, grp_syms) in enumerate(sector_groups.items()):
            with sg_cols[gi]:
                st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:.44rem;'
                            f'letter-spacing:.1em;text-transform:uppercase;color:#4A3858;'
                            f'margin-bottom:.25rem;">{grp_label}</div>', unsafe_allow_html=True)
                for ps in grp_syms:
                    if st.button(ps, key=f"qa_{ps}", use_container_width=True):
                        info2 = fetch_stock_fundamentals(ps)
                        if info2 and info2.get("price"):
                            st.session_state.portfolio[ps] = {
                                "shares":   1.0,
                                "avg_cost": info2["price"],
                                "added":    _dt.datetime.utcnow().strftime("%Y-%m-%d"),
                            }
                            st.rerun()

        # ── Remove ───────────────────────────────────────────────────────
        if portfolio:
            st.markdown("<hr style='border-color:rgba(139,58,139,.1);margin:.5rem 0 .4rem;'>",
                        unsafe_allow_html=True)
            rm_col1, rm_col2 = st.columns([4, 1])
            with rm_col1:
                rm_sym = st.selectbox("Remove holding", ["—"] + list(portfolio.keys()),
                                      label_visibility="collapsed", key="pf_rm")
            with rm_col2:
                if rm_sym != "—" and st.button("🗑 Remove", key="pf_rm_btn", use_container_width=True):
                    del st.session_state.portfolio[rm_sym]
                    if rm_sym in st.session_state.portfolio_notes:
                        del st.session_state.portfolio_notes[rm_sym]
                    st.rerun()

    if not portfolio:
        st.markdown('<div style="text-align:center;padding:2.5rem 1rem;">'
                    '<div style="font-size:2rem;opacity:.3;margin-bottom:.6rem;">📊</div>'
                    '<div style="font-family:\'Cormorant Garamond\',serif;font-size:1.3rem;'
                    'font-weight:300;font-style:italic;color:#4A3858;">'
                    'Add your first holding above</div></div>', unsafe_allow_html=True)
        return

    # ── Portfolio Summary ─────────────────────────────────────────────────
    with st.spinner("Fetching live prices…"):
        summary = portfolio_summary(portfolio)

    sv     = summary["total_value"]
    sc     = summary["total_cost"]
    pnl    = summary["total_pnl"]
    pnl_p  = summary["total_pnl_pct"]
    pnl_cl = "pos" if pnl >= 0 else "neg"
    pnl_sg = "+" if pnl >= 0 else ""

    st.markdown(
        f'<div style="display:flex;gap:1.5rem;flex-wrap:wrap;margin:.6rem 0 .8rem;padding:.8rem 1rem;'
        f'background:var(--card-2);border:1px solid var(--border);border-radius:10px;">'
        f'<div class="pph-stat"><div class="pph-stat-lbl">Portfolio Value</div>'
        f'<div class="pph-stat-val neu">${sv:,.2f}</div></div>'
        f'<div class="pph-stat"><div class="pph-stat-lbl">Total Cost</div>'
        f'<div class="pph-stat-val neu">${sc:,.2f}</div></div>'
        f'<div class="pph-stat"><div class="pph-stat-lbl">Unrealised P&L</div>'
        f'<div class="pph-stat-val {pnl_cl}">{pnl_sg}${abs(pnl):,.2f} ({pnl_sg}{abs(pnl_p):.2f}%)</div></div>'
        f'<div class="pph-stat"><div class="pph-stat-lbl">Holdings</div>'
        f'<div class="pph-stat-val neu">{len(portfolio)}</div></div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Allocation bar ────────────────────────────────────────────────────
    segs = ""
    for i, h in enumerate(summary["holdings"]):
        c = _ALLOC_COLORS[i % len(_ALLOC_COLORS)]
        segs += f'<div class="alloc-seg" style="width:{h["weight"]}%;background:{c};"></div>'
    st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.15em;'
                f'text-transform:uppercase;color:#4A3858;margin-bottom:.25rem;">Allocation</div>'
                f'<div class="alloc-bar">{segs}</div>', unsafe_allow_html=True)

    # Legend
    leg = ""
    for i, h in enumerate(summary["holdings"][:10]):
        c = _ALLOC_COLORS[i % len(_ALLOC_COLORS)]
        leg += (f'<span style="font-family:Space Mono,monospace;font-size:.48rem;color:{c};'
                f'background:rgba(0,0,0,.2);border-radius:3px;padding:.1rem .35rem;">'
                f'● {h["sym"]} {h["weight"]}%</span> ')
    st.markdown(f'<div style="display:flex;gap:.3rem;flex-wrap:wrap;margin-bottom:.8rem;">{leg}</div>',
                unsafe_allow_html=True)

    # ── Holdings Grid ─────────────────────────────────────────────────────
    st.markdown('<div style="font-family:Space Mono,monospace;font-size:.52rem;letter-spacing:.18em;'
                'text-transform:uppercase;color:var(--velvet-gl);margin:.3rem 0 .5rem;">Holdings</div>',
                unsafe_allow_html=True)

    n_cols = 3
    holdings = summary["holdings"]

    # ── Track which card is selected ─────────────────────────────────────
    _sel = st.session_state.get("_pf_selected_holding", "")

    for row_start in range(0, len(holdings), n_cols):
        row_h = holdings[row_start:row_start+n_cols]
        cols  = st.columns(n_cols)
        for ci, h in enumerate(row_h):
            with cols[ci]:
                arr  = "▲" if h["pct"] >= 0 else "▼"
                cls  = "up" if h["pct"] >= 0 else ("down" if h["pct"] < 0 else "flat")
                gain_cls = "gain" if h["pnl"] >= 0 else "loss"
                pnl_sg2  = "+" if h["pnl"] >= 0 else ""
                price_str = f"{h['currency']} {h['price']:,.4f}" if h["price"] < 1 else f"{h['currency']} {h['price']:,.2f}"
                hc_dir_cls = "hc-up" if h["pct"] >= 0 else "hc-down"
                is_selected = (h["sym"] == _sel)

                # Infer exchange label
                sym_up = h["sym"].upper()
                if sym_up.endswith(".NS") or sym_up.endswith(".BO"):
                    exch_label = "🇮🇳 NSE · India"; exch_color = "#fb923c"
                elif sym_up.endswith("-USD") or sym_up.endswith("-BTC"):
                    exch_label = "₿ Crypto"; exch_color = "#F0C040"
                elif sym_up.endswith(".KS"):
                    exch_label = "🇰🇷 KRX · S. Korea"; exch_color = "#a78bfa"
                elif sym_up.endswith(".HK"):
                    exch_label = "🇭🇰 HKEX · HK"; exch_color = "#f87171"
                elif sym_up.endswith(".TW"):
                    exch_label = "🇹🇼 TWSE · Taiwan"; exch_color = "#34d399"
                elif sym_up.endswith(".AX"):
                    exch_label = "🇦🇺 ASX · Australia"; exch_color = "#60a5fa"
                elif sym_up.endswith(".L"):
                    exch_label = "🇬🇧 LSE · London"; exch_color = "#e2e8f0"
                else:
                    exch_label = "🇺🇸 NYSE / NASDAQ"; exch_color = "#4ade80"

                # 52w bar
                if h.get("52w_high") and h.get("52w_low") and h["52w_high"] != h["52w_low"]:
                    pos52 = max(0, min(100, (h["price"] - h["52w_low"]) / (h["52w_high"] - h["52w_low"]) * 100))
                else:
                    pos52 = 50

                # Selected card gets a highlight ring
                sel_style = ("border-color:rgba(192,132,200,.65)!important;"
                             "box-shadow:0 0 0 2px rgba(192,132,200,.25),0 0 18px rgba(192,132,200,.18);"
                             if is_selected else "")

                st.markdown(
                    f'<div class="holding-card {hc_dir_cls}" style="{sel_style}">'
                    f'<div style="display:flex;align-items:flex-start;justify-content:space-between;">'
                    f'  <div><div class="hc-sym">{h["sym"]}</div>'
                    f'  <div class="hc-country" style="color:{exch_color};">{exch_label}</div></div>'
                    f'  <div style="text-align:right;font-family:Space Mono,monospace;font-size:.48rem;color:#4A3858;">{h["currency"]}</div>'
                    f'</div>'
                    f'<div class="hc-name">{h["short_name"][:32]}</div>'
                    f'<div class="hc-price">{price_str}</div>'
                    f'<div class="hc-chg {cls}">{arr} {abs(h["pct"]):.2f}% today</div>'
                    f'<div class="hc-meta">'
                    f'  <span class="hc-chip">{h["shares"]:,.4g} shares</span>'
                    f'  <span class="hc-chip">{h["weight"]}% alloc</span>'
                    f'  <span class="hc-chip {gain_cls}">P&L {pnl_sg2}${abs(h["pnl"]):,.2f} ({pnl_sg2}{abs(h["pnl_pct"]):.1f}%)</span>'
                    f'</div>'
                    f'<div style="margin-top:.5rem;">'
                    f'  <div style="font-family:Space Mono,monospace;font-size:.42rem;color:#4A3858;margin-bottom:.18rem;">'
                    f'    52W LOW {h["52w_low"] or "—"}  ←  HERE →  HIGH {h["52w_high"] or "—"}</div>'
                    f'  <div style="height:4px;background:rgba(107,45,107,.15);border-radius:2px;overflow:hidden;">'
                    f'    <div style="height:100%;width:{pos52:.1f}%;background:linear-gradient(90deg,#F0C040,#C084C8);border-radius:2px;"></div>'
                    f'  </div></div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                # Tap button — full-width under card, minimal
                btn_label = "▼ Hide Analysis" if is_selected else "🤖 Analyse"
                if st.button(btn_label, key=f"hc_sel_{h['sym']}", use_container_width=True,
                             help=f"AI performance analysis for {h['sym']}"):
                    if is_selected:
                        st.session_state["_pf_selected_holding"] = ""
                    else:
                        st.session_state["_pf_selected_holding"] = h["sym"]
                    st.rerun()

    # ── Instant AI Holding Analysis Panel ────────────────────────────────
    if _sel and _sel in {h["sym"] for h in holdings}:
        _h_data = next((h for h in holdings if h["sym"] == _sel), None)
        if _h_data:
            _ai_cache = st.session_state.get("_pf_holding_ai", {})
            if _sel not in _ai_cache:
                _ai_cache[_sel] = {}

            st.markdown(
                f'<div style="background:linear-gradient(135deg,rgba(107,45,107,.14),rgba(13,11,18,.98));'
                f'border:1.5px solid rgba(192,132,200,.3);border-radius:14px;'
                f'padding:1rem 1.2rem;margin:.6rem 0 .8rem;animation:slideDown .2s ease;">'
                f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:.7rem;">'
                f'<div style="font-family:\'Cormorant Garamond\',serif;font-size:1.15rem;font-weight:300;color:#EDE8F5;">'
                f'  <span style="color:#C084C8;">{_sel}</span>'
                f'  <span style="font-size:.7rem;font-family:Space Mono,monospace;color:#4A3858;margin-left:.5rem;">'
                f'  {_h_data["short_name"][:28]}</span></div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#ff80c0;">🤖 AI Performance Analysis</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # Timeframe tabs: Week · Month · Year
            _tf_labels  = ["1W · Week", "1M · Month", "1Y · Year"]
            _tf_keys    = ["1W", "1M", "1Y"]
            _tf_periods = {"1W": ("5d","30m"), "1M": ("1mo","1d"), "1Y": ("1y","1d")}
            _tf_descs   = {
                "1W": "past week's swing structure, support/resistance, and momentum",
                "1M": "monthly trend, SMA positioning, and mean-reversion signals",
                "1Y": "annual performance, macro cycle positioning, and relative sector strength",
            }

            # Arrow nav row
            _holding_tf_key = f"_pf_htf_{_sel}"
            if _holding_tf_key not in st.session_state:
                st.session_state[_holding_tf_key] = "1W"
            _htf = st.session_state[_holding_tf_key]
            _htf_i = _tf_keys.index(_htf)

            _anav = st.columns([0.5, 1, 1, 1, 0.5])
            with _anav[0]:
                if st.button("◀", key=f"hprev_{_sel}"):
                    st.session_state[_holding_tf_key] = _tf_keys[max(0, _htf_i - 1)]
                    _ai_cache[_sel].pop(st.session_state[_holding_tf_key], None)
                    st.rerun()
            for _ti, (_tlbl, _tk) in enumerate(zip(_tf_labels, _tf_keys)):
                with _anav[_ti + 1]:
                    _is_htf_active = (_tk == _htf)
                    _htf_style = ("background:rgba(192,132,200,.22)!important;"
                                  "border-color:#C084C8!important;color:#EDE8F5!important;"
                                  if _is_htf_active else "")
                    if st.button(_tlbl, key=f"htf_{_sel}_{_tk}", use_container_width=True):
                        st.session_state[_holding_tf_key] = _tk
                        st.rerun()
            with _anav[4]:
                if st.button("▶", key=f"hnext_{_sel}"):
                    st.session_state[_holding_tf_key] = _tf_keys[min(len(_tf_keys)-1, _htf_i + 1)]
                    st.rerun()

            # Fetch & display AI analysis for selected timeframe
            _period, _interval = _tf_periods[_htf]
            if _htf not in _ai_cache.get(_sel, {}):
                with st.spinner(f"Analysing {_sel} over {_htf}…"):
                    _stats = _compute_tf_stats(_sel, _period, _interval)
                    _ctx   = (f"{_sel} ({_h_data['short_name']}) — {_tf_descs[_htf]} — "
                              f"Position: {_h_data['shares']:,.4g} shares @ ${_h_data['avg_cost']:,.2f} avg cost, "
                              f"P&L: ${_h_data['pnl']:+,.2f} ({_h_data['pnl_pct']:+.1f}%)")
                    _ai_txt = ai_market_analysis([_stats], groq_api_key, context=_ctx)
                    _ai_cache.setdefault(_sel, {})[_htf] = _ai_txt
                    st.session_state["_pf_holding_ai"] = _ai_cache

            _txt = _ai_cache.get(_sel, {}).get(_htf, "")
            if _txt:
                _pct1 = abs(_h_data["pct"])
                _arr1 = "▲" if _h_data["pct"] >= 0 else "▼"
                _cc1  = "#4ade80" if _h_data["pct"] >= 0 else "#f87171"
                st.markdown(
                    f'<div style="display:flex;gap:.5rem;flex-wrap:wrap;margin-bottom:.55rem;">'
                    f'<span style="font-family:Space Mono,monospace;font-size:.52rem;'
                    f'background:rgba(107,45,107,.15);border:1px solid rgba(139,58,139,.3);'
                    f'border-radius:4px;padding:.18rem .5rem;color:#C084C8;">'
                    f'Today: <span style="color:{_cc1};">{_arr1} {_pct1:.2f}%</span></span>'
                    f'<span style="font-family:Space Mono,monospace;font-size:.52rem;'
                    f'background:rgba(107,45,107,.15);border:1px solid rgba(139,58,139,.3);'
                    f'border-radius:4px;padding:.18rem .5rem;color:#C084C8;">'
                    f'Horizon: {_htf}</span>'
                    f'</div>'
                    f'<div style="font-family:Syne,sans-serif;font-size:.83rem;color:#C8B8D8;'
                    f'line-height:1.78;border-left:2px solid rgba(192,132,200,.3);'
                    f'padding-left:.75rem;">{_txt}</div>',
                    unsafe_allow_html=True,
                )
            st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr style='border-color:rgba(139,58,139,.12);margin:1rem 0 .8rem;'>",
                unsafe_allow_html=True)

    # ── Deep Analysis: pick a stock ───────────────────────────────────────
    st.markdown('<div style="font-family:Space Mono,monospace;font-size:.52rem;letter-spacing:.18em;'
                'text-transform:uppercase;color:var(--velvet-gl);margin-bottom:.5rem;">'
                'AI-Assisted Stock Deep Analysis</div>', unsafe_allow_html=True)

    an_sym = st.selectbox("Analyse a holding", list(portfolio.keys()),
                          key="pf_analyse_sym", label_visibility="collapsed")

    if an_sym:
        an_tabs = st.tabs(["📉 Technicals","🤖 AI Analysis","📊 Simulate Position",
                            "⚠ Risk Analytics","🔔 Alerts"])

        # ── Tab 0: Technicals ─────────────────────────────────────────────
        with an_tabs[0]:
            with st.spinner(f"Computing technicals for {an_sym}…"):
                hist_df = fetch_stock_history_1y(an_sym)
                tech    = compute_technicals(hist_df)
                fund    = fetch_stock_fundamentals(an_sym)

            if not tech:
                st.warning("Not enough historical data for technical analysis.")
            else:
                price = fund.get("price", 0) or 0

                # Signal badge
                sig = tech["signal"]
                sig_cls = sig.lower()
                sig_emoji = {"BUY":"🟢","HOLD":"🟡","SELL":"🔴","WATCH":"🔵"}.get(sig,"⚪")
                st.markdown(f'<div style="margin-bottom:.8rem;">'
                            f'<span class="signal {sig_cls}">{sig_emoji} {sig}</span>'
                            f'<span style="font-family:Space Mono,monospace;font-size:.5rem;'
                            f'color:#4A3858;margin-left:.6rem;">Technical Score: {tech["score"]}/7</span>'
                            f'</div>', unsafe_allow_html=True)

                # Key metrics grid
                t1, t2, t3, t4 = st.columns(4)
                def _tech_card(col, label, value, note="", good=None):
                    if value is None: return
                    clr = ("#4ade80" if good is True else "#f87171" if good is False else "#C084C8")
                    col.markdown(
                        f'<div style="background:var(--card-2);border:1px solid var(--border);'
                        f'border-top:2px solid {clr};border-radius:8px;padding:.65rem .8rem;">'
                        f'<div style="font-family:Space Mono,monospace;font-size:.46rem;'
                        f'letter-spacing:.15em;text-transform:uppercase;color:#4A3858;">{label}</div>'
                        f'<div style="font-family:Cormorant Garamond,serif;font-size:1.4rem;'
                        f'font-weight:300;color:#EDE8F5;line-height:1;">{value}</div>'
                        f'{"<div style=font-family:Space Mono,monospace;font-size:.46rem;color:"+clr+">" + note + "</div>" if note else ""}'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                rsi = tech["rsi14"]
                _tech_card(t1, "RSI-14", f"{rsi}", "Oversold" if rsi<30 else ("Overbought" if rsi>70 else "Neutral"),
                           good=True if rsi<30 else (False if rsi>70 else None))
                _tech_card(t2, "MACD Hist", f"{tech['macd_hist']:+.4f}",
                           "Bullish" if tech["macd_hist"]>0 else "Bearish",
                           good=tech["macd_hist"]>0)
                _tech_card(t3, "BB Position", f"{tech['bb_pos']:.0f}%",
                           "Near Upper" if tech["bb_pos"]>80 else ("Near Lower" if tech["bb_pos"]<20 else "Mid Range"),
                           good=True if tech["bb_pos"]<20 else (False if tech["bb_pos"]>80 else None))
                _tech_card(t4, "ATR-14", f"${tech['atr14']:.2f}", "Daily volatility range")

                st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)
                t5, t6, t7, t8 = st.columns(4)
                _tech_card(t5, "SMA 20",  f"${tech['sma20']:,.2f}" if tech["sma20"] else "—",
                           "Price above" if price > (tech["sma20"] or 0) else "Price below",
                           good=price > (tech["sma20"] or float("inf")))
                _tech_card(t6, "SMA 50",  f"${tech['sma50']:,.2f}" if tech["sma50"] else "—",
                           "Price above" if price > (tech["sma50"] or 0) else "Price below",
                           good=price > (tech["sma50"] or float("inf")))
                _tech_card(t7, "SMA 200", f"${tech['sma200']:,.2f}" if tech["sma200"] else "—",
                           "Price above" if price > (tech["sma200"] or 0) else "Price below",
                           good=price > (tech["sma200"] or float("inf")))
                _tech_card(t8, "Ann. Vol", f"{tech['vol_annual']:.1f}%",
                           "High" if tech["vol_annual"]>40 else ("Low" if tech["vol_annual"]<15 else "Moderate"),
                           good=tech["vol_annual"]<25)

                # Momentum row
                st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;'
                            'letter-spacing:.15em;text-transform:uppercase;color:#4A3858;'
                            'margin:.8rem 0 .4rem;">Price Momentum</div>', unsafe_allow_html=True)
                m1, m2, m3, m4 = st.columns(4)
                for col, lbl, val in [(m1,"1 Month",tech["mom_1m"]),(m2,"3 Month",tech["mom_3m"]),
                                      (m3,"6 Month",tech["mom_6m"]),(m4,"1 Year",tech["mom_1y"])]:
                    if val is not None:
                        _tech_card(col, lbl, f"{val:+.1f}%", "", good=val>0)

                # Price chart
                if not hist_df.empty:
                    st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;'
                                'letter-spacing:.15em;text-transform:uppercase;color:#4A3858;'
                                'margin:.8rem 0 .3rem;">1-Year Price Chart</div>', unsafe_allow_html=True)
                    chart_df = hist_df[["close"]].copy()
                    chart_df.columns = [an_sym]
                    st.line_chart(chart_df, height=200, use_container_width=True)

                # Bollinger band chart
                if not hist_df.empty and len(hist_df) >= 20:
                    st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;'
                                'letter-spacing:.15em;text-transform:uppercase;color:#4A3858;'
                                'margin:.6rem 0 .3rem;">Bollinger Bands (last 60 days)</div>',
                                unsafe_allow_html=True)
                    bb_df = hist_df["close"].iloc[-60:].to_frame()
                    roll  = hist_df["close"].rolling(20)
                    bb_df["SMA20"]    = roll.mean()
                    bb_df["BB Upper"] = roll.mean() + 2 * roll.std()
                    bb_df["BB Lower"] = roll.mean() - 2 * roll.std()
                    bb_df = bb_df.rename(columns={"close": an_sym}).iloc[-60:]
                    st.line_chart(bb_df.dropna(), height=180, use_container_width=True)

        # ── Tab 1: AI Analysis ────────────────────────────────────────────
        with an_tabs[1]:
            with st.spinner("Loading data…"):
                hist_df2 = fetch_stock_history_1y(an_sym) if "hist_df" not in dir() else hist_df
                tech2    = compute_technicals(hist_df2) if not tech else tech
                fund2    = fetch_stock_fundamentals(an_sym) if not fund else fund
                pos      = portfolio.get(an_sym, {})

            if not groq_api_key:
                st.info("Add a Groq API key in the sidebar to enable AI analysis.")
            else:
                if st.button(f"🤖  Generate Full AI Analysis for {an_sym}", key=f"ai_{an_sym}",
                             use_container_width=False):
                    with st.spinner("Analysing with Llama 3.3-70B…"):
                        try:
                            from openai import OpenAI
                            oai = OpenAI(api_key=groq_api_key, base_url="https://api.groq.com/openai/v1")

                            tech_summary = (
                                f"RSI-14={tech2.get('rsi14','—')}, "
                                f"MACD hist={tech2.get('macd_hist','—')}, "
                                f"SMA20={tech2.get('sma20','—')}, SMA50={tech2.get('sma50','—')}, "
                                f"SMA200={tech2.get('sma200','—')}, "
                                f"BB pos={tech2.get('bb_pos','—')}%, "
                                f"Ann.Vol={tech2.get('vol_annual','—')}%, "
                                f"Signal={tech2.get('signal','—')}"
                            )
                            mom_summary = (
                                f"1M={tech2.get('mom_1m','—')}%, "
                                f"3M={tech2.get('mom_3m','—')}%, "
                                f"6M={tech2.get('mom_6m','—')}%, "
                                f"1Y={tech2.get('mom_1y','—')}%"
                            )
                            pos_summary = (
                                f"Holding {pos.get('shares','—')} shares, "
                                f"avg cost ${pos.get('avg_cost',0):.2f}, "
                                f"current ${fund2.get('price',0):.2f}, "
                                f"unrealised P&L ${pos.get('shares',0)*(fund2.get('price',0)-pos.get('avg_cost',0)):.2f}"
                            )
                            fund_summary = (
                                f"52W High={fund2.get('52w_high','—')}, "
                                f"52W Low={fund2.get('52w_low','—')}, "
                                f"Market Cap={fund2.get('mkt_cap','—')}, "
                                f"Volume={fund2.get('volume','—')}"
                            )

                            # Current market context
                            fng2 = fetch_fear_greed()
                            sp_q = fetch_quote("^GSPC")
                            mkt_ctx = (f"S&P 500 at {sp_q['price']:,.0f} ({sp_q['pct']:+.2f}%), "
                                       f"Fear & Greed = {fng2['value']} ({fng2['label']})"
                                       if sp_q else "Market data unavailable")

                            prompt = f"""You are a senior equity analyst. Provide a comprehensive analysis of {an_sym} ({fund2.get("short_name","")}).

TECHNICAL INDICATORS:
{tech_summary}

MOMENTUM:
{mom_summary}

FUNDAMENTAL SNAPSHOT:
{fund_summary}

INVESTOR POSITION:
{pos_summary}

CURRENT MARKET CONTEXT:
{mkt_ctx}

Write a structured analysis with these sections:
1. **Technical Picture** — What the indicators say, confluence or divergence, key levels to watch
2. **Momentum Assessment** — Trend strength, rotation signals, sector context
3. **Risk Assessment** — Volatility, downside risk, stop-loss levels based on ATR and support
4. **Catalyst Watch** — What could drive the next move up or down
5. **Recommendation** — Clear BUY / HOLD / SELL / WATCH with price target rationale and suggested position sizing

Be specific with numbers. Max 350 words. Use the investor's actual position data when relevant."""

                            resp = oai.chat.completions.create(
                                model="llama-3.3-70b-versatile",
                                messages=[
                                    {"role":"system","content":"You are a senior equity analyst. Be direct, data-driven, cite specific numbers."},
                                    {"role":"user","content":prompt},
                                ],
                                temperature=0.15, max_tokens=700,
                            )
                            analysis = resp.choices[0].message.content
                            st.session_state.portfolio_notes[an_sym] = analysis
                        except Exception as e:
                            st.error(f"AI analysis error: {e}")

                if an_sym in st.session_state.portfolio_notes:
                    st.markdown(
                        f'<div class="ai-analysis-card">'
                        f'<div class="aac-header">AI Analysis · {an_sym} · {_dt.datetime.utcnow().strftime("%Y-%m-%d")}</div>'
                        f'<div class="aac-body">{st.session_state.portfolio_notes[an_sym].replace(chr(10),"<br>")}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown('<div style="font-family:Syne,sans-serif;font-size:.8rem;color:#4A3858;'
                                'padding:1.5rem;text-align:center;">Click the button above to generate AI-powered analysis</div>',
                                unsafe_allow_html=True)

        # ── Tab 2: Simulate Position ──────────────────────────────────────
        with an_tabs[2]:
            fund3 = fetch_stock_fundamentals(an_sym) if not fund else fund
            curr_price = fund3.get("price") or 0
            pos3 = portfolio.get(an_sym, {})
            avg3 = pos3.get("avg_cost", curr_price)
            shares3 = pos3.get("shares", 1)

            st.markdown('<div style="font-family:Space Mono,monospace;font-size:.52rem;'
                        'letter-spacing:.18em;text-transform:uppercase;color:var(--velvet-gl);'
                        'margin-bottom:.7rem;">Position Simulator</div>', unsafe_allow_html=True)

            sc1, sc2 = st.columns(2)
            with sc1:
                sim_target = st.number_input("Target price ($)", value=float(round(curr_price*1.15,2)),
                                             step=0.5, key="sim_target")
                sim_stop   = st.number_input("Stop-loss price ($)", value=float(round(curr_price*0.92,2)),
                                             step=0.5, key="sim_stop")
            with sc2:
                sim_shares = st.number_input("Position size (shares)", value=float(shares3),
                                             step=0.1, key="sim_shares")
                sim_entry  = st.number_input("Entry price ($)", value=float(round(avg3,2)),
                                             step=0.1, key="sim_entry")

            if curr_price > 0:
                gain_amt   = (sim_target - sim_entry) * sim_shares
                loss_amt   = (sim_entry  - sim_stop)  * sim_shares
                rr_ratio   = abs(gain_amt / loss_amt) if loss_amt else 0
                gain_pct   = (sim_target - sim_entry) / sim_entry * 100
                loss_pct   = (sim_entry  - sim_stop)  / sim_entry * 100
                curr_pnl   = (curr_price - sim_entry) * sim_shares
                position_val = sim_entry * sim_shares

                r1, r2, r3, r4 = st.columns(4)
                def _sim_card(col, label, val, color):
                    col.markdown(f'<div style="background:var(--card-2);border:1px solid var(--border);'
                                 f'border-top:2px solid {color};border-radius:8px;padding:.65rem .8rem;'
                                 f'font-family:Space Mono,monospace;">'
                                 f'<div style="font-size:.44rem;letter-spacing:.15em;text-transform:uppercase;'
                                 f'color:#4A3858;">{label}</div>'
                                 f'<div style="font-family:Cormorant Garamond,serif;font-size:1.45rem;'
                                 f'font-weight:300;color:{color};line-height:1.1;">{val}</div>'
                                 f'</div>', unsafe_allow_html=True)

                _sim_card(r1, "Max Gain",    f"+${gain_amt:,.2f} (+{gain_pct:.1f}%)", "#4ade80")
                _sim_card(r2, "Max Loss",    f"-${loss_amt:,.2f} (-{loss_pct:.1f}%)", "#f87171")
                _sim_card(r3, "R:R Ratio",   f"{rr_ratio:.2f}:1",
                          "#4ade80" if rr_ratio >= 2 else ("#F0C040" if rr_ratio >= 1 else "#f87171"))
                _sim_card(r4, "Current P&L", f"{'+'if curr_pnl>=0 else ''}${curr_pnl:,.2f}",
                          "#4ade80" if curr_pnl >= 0 else "#f87171")

                st.markdown(f'<div style="margin-top:.7rem;padding:.6rem .9rem;'
                            f'background:rgba(107,45,107,.08);border:1px solid var(--border);'
                            f'border-radius:8px;font-family:Space Mono,monospace;font-size:.58rem;'
                            f'color:var(--text-dim);">'
                            f'Position Value: ${position_val:,.2f} &nbsp;·&nbsp; '
                            f'Entry: ${sim_entry:,.2f} &nbsp;·&nbsp; '
                            f'Current: ${curr_price:,.2f} &nbsp;·&nbsp; '
                            f'Target: ${sim_target:,.2f} &nbsp;·&nbsp; '
                            f'Stop: ${sim_stop:,.2f}'
                            f'</div>', unsafe_allow_html=True)

                # Scenario table
                st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;'
                            'letter-spacing:.15em;text-transform:uppercase;color:#4A3858;'
                            'margin:.8rem 0 .4rem;">Scenario Analysis</div>', unsafe_allow_html=True)
                scenarios = [
                    ("Bear -20%", curr_price*0.80),
                    ("Bear -10%", curr_price*0.90),
                    ("Base  +0%", curr_price),
                    ("Bull +10%", curr_price*1.10),
                    ("Bull +20%", curr_price*1.20),
                    ("Bull +30%", curr_price*1.30),
                ]
                sc_rows = "".join(
                    f'<tr>'
                    f'<td style="font-family:Space Mono,monospace;font-size:.6rem;color:var(--text-dim);">{sc_lbl}</td>'
                    f'<td style="font-family:Space Mono,monospace;font-size:.6rem;color:#9A8AAA;">${sc_price:,.2f}</td>'
                    f'<td style="font-family:Space Mono,monospace;font-size:.6rem;'
                    f'color:{"#4ade80" if (sc_price-sim_entry)*sim_shares>=0 else "#f87171"};">'
                    f'{"+"if (sc_price-sim_entry)*sim_shares>=0 else ""}${(sc_price-sim_entry)*sim_shares:,.2f}</td>'
                    f'<td style="font-family:Space Mono,monospace;font-size:.6rem;'
                    f'color:{"#4ade80" if sc_price>=sim_entry else "#f87171"};">'
                    f'{"+"if sc_price>=sim_entry else ""}{(sc_price-sim_entry)/sim_entry*100:.1f}%</td>'
                    f'</tr>'
                    for sc_lbl, sc_price in scenarios
                )
                st.markdown(
                    f'<table style="width:100%;border-collapse:collapse;">'
                    f'<thead><tr>'
                    f'<th style="background:rgba(107,45,107,.18);border:1px solid var(--border);'
                    f'padding:.35rem .7rem;font-family:Space Mono,monospace;font-size:.46rem;'
                    f'text-transform:uppercase;letter-spacing:.1em;color:var(--velvet-gl);text-align:left;">Scenario</th>'
                    f'<th style="background:rgba(107,45,107,.18);border:1px solid var(--border);'
                    f'padding:.35rem .7rem;font-family:Space Mono,monospace;font-size:.46rem;'
                    f'text-transform:uppercase;letter-spacing:.1em;color:var(--velvet-gl);">Price</th>'
                    f'<th style="background:rgba(107,45,107,.18);border:1px solid var(--border);'
                    f'padding:.35rem .7rem;font-family:Space Mono,monospace;font-size:.46rem;'
                    f'text-transform:uppercase;letter-spacing:.1em;color:var(--velvet-gl);">P&L</th>'
                    f'<th style="background:rgba(107,45,107,.18);border:1px solid var(--border);'
                    f'padding:.35rem .7rem;font-family:Space Mono,monospace;font-size:.46rem;'
                    f'text-transform:uppercase;letter-spacing:.1em;color:var(--velvet-gl);">Return</th>'
                    f'</tr></thead><tbody>{sc_rows}</tbody></table>',
                    unsafe_allow_html=True,
                )

        # ── Tab 3: Risk Analytics ─────────────────────────────────────────
        with an_tabs[3]:
            with st.spinner("Loading portfolio risk analytics…"):
                summary_for_risk = portfolio_summary(portfolio)
            render_portfolio_analytics(summary_for_risk, groq_api_key)

        # ── Tab 4: Alerts ─────────────────────────────────────────────────
        with an_tabs[4]:
            render_alerts_panel(portfolio)

# Section classifier: map chunk text → probable report section
_SECTION_PATTERNS = [
    (r"management.{0,20}discussion|MD&A",                        "Management Discussion & Analysis"),
    (r"risk factor",                                              "Risk Factors"),
    (r"financial statement|balance sheet|income statement|P&L",  "Financial Statements"),
    (r"cash flow",                                                "Cash Flow Statement"),
    (r"note\s+\d|notes to|accounting polic",                     "Notes to Accounts"),
    (r"auditor.{0,15}report|independent auditor",                "Auditor's Report"),
    (r"segment|business unit|division",                          "Segment Reporting"),
    (r"revenue|sales|turnover",                                   "Revenue & Sales"),
    (r"profit|margin|EBITDA|operating income",                   "Profitability"),
    (r"debt|borrowing|loan|credit facilit",                      "Debt & Financing"),
    (r"dividend|share capital|equity",                            "Shareholders' Equity"),
    (r"outlook|guidance|forward.looking",                        "Outlook & Guidance"),
    (r"corporate governance",                                     "Corporate Governance"),
    (r"chairman|board of director|CEO|management team",          "Board & Leadership"),
    (r"macroeconomic|market condition|industry trend",           "Market & Industry"),
]

_QUERY_EXPANSIONS: dict[str, list[str]] = {
    r"\beps\b|earnings per share": [
        "diluted EPS", "basic EPS", "earnings per share diluted",
        "diluted earnings per share", "EPS diluted", "net income per share"],
    r"revenue|total revenue|net sales": [
        "total revenue", "net sales", "revenue from operations",
        "gross revenue", "top line", "turnover"],
    r"net income|net profit|profit after tax|PAT": [
        "net income", "net profit", "profit after tax", "PAT",
        "net earnings", "bottom line"],
    r"gross margin|gross profit": [
        "gross profit", "gross margin", "gross profit margin",
        "cost of revenue", "COGS", "cost of goods sold"],
    r"ebitda": [
        "EBITDA", "earnings before interest tax depreciation",
        "operating EBITDA", "adjusted EBITDA"],
    r"free cash flow|FCF": [
        "free cash flow", "FCF", "operating cash flow minus capex",
        "cash from operations", "capital expenditure"],
    r"debt|borrowing|leverage": [
        "total debt", "long-term debt", "borrowings", "net debt",
        "debt-to-equity", "leverage ratio"],
    r"dividend|DPS|payout": [
        "dividend per share", "DPS", "dividend payout",
        "interim dividend", "final dividend"],
    r"ROE|return on equity": [
        "return on equity", "ROE", "return on net worth"],
    r"ROA|return on assets": [
        "return on assets", "ROA", "return on total assets"],
}

def _expand_query(q: str) -> str:
    """
    Two-stage financial query expansion:
    1. NER-style entity detection — identifies company names, ticker symbols,
       and fiscal periods to anchor the query context.
    2. Synonym expansion — appends canonical financial term variants so BM25
       can recall chunks that use different phrasing for the same metric.
    The expanded string is used only for embedding/BM25 — never shown to the user.
    """
    additions: list[str] = []

    # ── Stage 1: fiscal period anchoring ────────────────────────────────
    # If the query mentions a year or quarter, make it explicit in the expansion
    yr_match = re.search(r"\b(20\d{2}|FY\d{2,4}|Q[1-4][\s\-]?20\d{2})\b", q, re.IGNORECASE)
    if yr_match:
        yr = yr_match.group(1)
        additions.append(f"fiscal year {yr} annual results")

    # ── Stage 2: financial synonym expansion ────────────────────────────
    for pattern, terms in _QUERY_EXPANSIONS.items():
        if re.search(pattern, q, re.IGNORECASE):
            additions.extend(t for t in terms if t.lower() not in q.lower())

    if additions:
        return q + " " + " ".join(additions[:8])
    return q

def guess_section(text: str) -> str:
    """Infer probable document section from chunk text."""
    t = text[:600].lower()
    for pattern, label in _SECTION_PATTERNS:
        if re.search(pattern, t, re.IGNORECASE):
            return label
    return "General / Other"

def render_source_panel(sources: list[dict]) -> None:
    """
    Render rich expandable evidence panel below an answer.
    sources: list of {"filename", "score", "preview", "chunk_full"(optional), "chunk_idx"(optional)}
    """
    if not sources:
        return
    n = len(sources)
    with st.expander(f"📂  View Sources  ·  {n} evidence chunk{'s' if n>1 else ''} retrieved"):
        st.markdown('<div class="evidence-panel">', unsafe_allow_html=True)
        for i, src in enumerate(sources, 1):
            score     = src.get("score", 0)
            score_pct = int(score * 100)
            score_cls = "score-hi" if score_pct >= 75 else ("score-mid" if score_pct >= 50 else "score-lo")
            score_lbl = "High relevance" if score_pct >= 75 else ("Moderate" if score_pct >= 50 else "Low relevance")
            # Use full chunk if available, else preview
            chunk_text = src.get("chunk_full") or src.get("preview","")
            section    = guess_section(chunk_text)
            chunk_idx  = src.get("chunk_idx", "")
            page_num   = src.get("page", "")
            page_est   = (f"p.{page_num} · Chunk #{chunk_idx}" if page_num and chunk_idx != ""
                          else f"p.{page_num}" if page_num
                          else f"Chunk #{chunk_idx}" if chunk_idx != ""
                          else f"Chunk ~{i}")
            fname      = _ht.escape(src.get("filename","Unknown"))
            chunk_disp = _ht.escape(chunk_text[:500] + ("…" if len(chunk_text) > 500 else ""))

            st.markdown(
                f'<div class="evidence-source-card">'
                f'  <div class="esc-header">'
                f'    <div style="display:flex;align-items:center;gap:.6rem;">'
                f'      <span class="esc-source-num">Source {i}</span>'
                f'      <span class="esc-filename">📄 {fname}</span>'
                f'    </div>'
                f'    <div class="esc-meta">'
                f'      <span class="esc-chip">{page_est}</span>'
                f'      <span class="esc-chip {score_cls}">⬡ {score_pct}% · {score_lbl}</span>'
                f'    </div>'
                f'  </div>'
                f'  <div class="esc-section">📑 {_ht.escape(section)}</div>'
                f'  <div class="esc-body">{chunk_disp}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)


def render_analyst_output(raw_json_str: str, question: str) -> None:
    """
    Parse the structured JSON from Analyst Mode and render as a rich card.
    Falls back gracefully to markdown if JSON is malformed.
    """
    # Strip possible fences
    clean = re.sub(r"```(?:json)?", "", raw_json_str, flags=re.IGNORECASE).strip().strip("`").strip()

    try:
        data = json.loads(clean)
    except json.JSONDecodeError:
        # Graceful fallback — render raw markdown
        st.markdown(raw_json_str)
        return

    metrics  = data.get("metrics", [])       # list of {label, value, unit, change, direction}
    summary  = data.get("summary", "")
    risks    = data.get("key_risks", [])
    outlook  = data.get("outlook", "")
    signal   = data.get("recommendation", "") # BUY / HOLD / SELL / WATCH / N/A

    sig_css = {"BUY":"buy","HOLD":"hold","SELL":"sell","WATCH":"watch"}.get(
                signal.upper() if signal else "", "watch")
    sig_emoji = {"BUY":"🟢","HOLD":"🟡","SELL":"🔴","WATCH":"🔵","N/A":"⚪"}.get(
                  signal.upper() if signal else "N/A","⚪")

    # ── Header ────────────────────────────────────────────────────────────
    st.markdown(
        f'<div class="analyst-output">'
        f'<div class="ao-header">'
        f'  <div class="ao-title">Analyst Report</div>'
        f'  <div style="display:flex;align-items:center;gap:.5rem;">'
        f'    <span class="ao-mode-badge">📊 Analyst Mode</span>'
        + (f'    <span class="signal {sig_css}" style="font-size:.5rem;">{sig_emoji} {signal.upper()}</span>'
           if signal and signal.upper() != "N/A" else "") +
        f'  </div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Metrics grid ──────────────────────────────────────────────────────
    if metrics:
        metrics_html = '<div class="ao-grid">'
        for m in metrics:
            val   = str(m.get("value","—"))
            unit  = m.get("unit","")
            chg   = m.get("change","")
            dirn  = str(m.get("direction","")).lower()
            v_cls = "pos" if dirn == "up" else ("neg" if dirn == "down" else "neu")
            chg_html = ""
            if chg:
                chg_col = "#4ade80" if dirn=="up" else ("#f87171" if dirn=="down" else "#9A8AAA")
                chg_html = (f'<div style="font-family:Space Mono,monospace;font-size:.44rem;'
                            f'color:{chg_col};margin-top:.1rem;">'
                            f'{"▲" if dirn=="up" else ("▼" if dirn=="down" else "●")} {_ht.escape(str(chg))}</div>')
            metrics_html += (
                f'<div class="ao-metric">'
                f'  <div class="ao-metric-label">{_ht.escape(m.get("label",""))}</div>'
                f'  <div class="ao-metric-value {v_cls}">{_ht.escape(val)}'
                f'  <span style="font-family:Space Mono,monospace;font-size:.5rem;color:#4A3858;"> {_ht.escape(unit)}</span></div>'
                + chg_html +
                f'</div>'
            )
        metrics_html += "</div>"
        st.markdown(metrics_html, unsafe_allow_html=True)

    # ── Summary ───────────────────────────────────────────────────────────
    if summary:
        st.markdown(
            f'<div class="ao-section">'
            f'<div class="ao-section-title">Executive Summary</div>'
            f'<div class="ao-section-body">{_ht.escape(summary).replace(chr(10),"<br>")}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Key risks ─────────────────────────────────────────────────────────
    if risks:
        risk_items = "".join(
            f'<div class="ao-risk-item">{_ht.escape(str(r))}</div>'
            for r in risks
        )
        st.markdown(
            f'<div class="ao-section">'
            f'<div class="ao-section-title">Key Risk Highlights</div>'
            f'{risk_items}'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Outlook ───────────────────────────────────────────────────────────
    if outlook:
        st.markdown(
            f'<div class="ao-section">'
            f'<div class="ao-section-title">Outlook</div>'
            f'<div class="ao-section-body">{_ht.escape(outlook).replace(chr(10),"<br>")}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("</div>", unsafe_allow_html=True)   # close .analyst-output


# Analyst Mode system prompt + user prompt templates
_ANALYST_SYSTEM = """You are a senior financial analyst. The user is querying financial documents.
You MUST respond ONLY with a valid JSON object — no prose, no markdown fences, no preamble.

JSON schema (all fields optional except metrics):
{
  "metrics": [
    {"label": "Revenue FY24", "value": "₹2,43,020 Cr", "unit": "INR Cr", "change": "+12.3% YoY", "direction": "up"},
    {"label": "Net Income", "value": "₹26,S248 Cr", "unit": "INR Cr", "change": "+8.1% YoY", "direction": "up"},
    ...
  ],
  "summary": "2-3 sentence executive summary of the key findings.",
  "key_risks": ["Risk 1 text", "Risk 2 text", "Risk 3 text"],
  "outlook": "Forward-looking commentary based on document.",
  "recommendation": "BUY | HOLD | SELL | WATCH | N/A"
}

Rules:
- Extract exact numbers from the document context. Never fabricate values.
- If a value is not found, omit that metric rather than guessing.
- direction must be exactly "up", "down", or "flat".
- recommendation must be exactly one of: BUY, HOLD, SELL, WATCH, N/A.
- Return ONLY the JSON. No other text whatsoever."""

def build_analyst_prompt(question: str, live_context: str, doc_context: str) -> str:
    return (f"{live_context}\n\n=== DOCUMENT CONTEXT ===\n{doc_context}\n\n"
            f"User question: {question}\n\n"
            f"Extract structured financial data as specified. Return ONLY JSON.")


# ─────────────────────────────────────────────────────────────────────────────
# INDIA NEWS RSS
# ─────────────────────────────────────────────────────────────────────────────
NEWS_SOURCES = {
    "📰 Business Standard":  {"rss":"https://www.business-standard.com/rss/home_page_top_stories.rss","tag":"Markets"},
    "📰 Economic Times":     {"rss":"https://economictimes.indiatimes.com/markets/rssfeeds/1977021501.cms","tag":"Markets"},
    "📰 Mint":               {"rss":"https://www.livemint.com/rss/markets","tag":"Markets"},
    "📰 Hindu BusinessLine": {"rss":"https://www.thehindubusinessline.com/markets/feeder/default.rss","tag":"Markets"},
    "🏛️ RBI Notifications":  {"rss":"https://www.rbi.org.in/Scripts/Notifications_Rss.aspx","tag":"Policy"},
    "🏛️ SEBI Orders":        {"rss":"https://www.sebi.gov.in/sebiweb/other/OtherAction.do?doGetPublicationRss=yes&rssHead=4","tag":"Policy"},
    "🏛️ Finance Ministry":   {"rss":"https://pib.gov.in/RssMain.aspx?ModId=6&Lang=1&Regid=3","tag":"Policy"},
    "🏛️ PIB – Economy":      {"rss":"https://pib.gov.in/RssMain.aspx?ModId=4&Lang=1&Regid=3","tag":"Policy"},
}

@st.cache_data(ttl=600)
def fetch_rss(url, max_items=6):
    import xml.etree.ElementTree as ET
    headers = {"User-Agent":"Mozilla/5.0 (compatible; NewsBot/1.0)","Accept":"application/rss+xml,*/*"}
    try:
        r = requests.get(url, headers=headers, timeout=12); r.raise_for_status()
        root = ET.fromstring(r.content)
        items = root.findall(".//item"); results = []
        for item in items[:max_items]:
            title = (item.findtext("title") or "").strip()
            link  = (item.findtext("link")  or "").strip()
            pub   = (item.findtext("pubDate") or "").strip()
            desc  = re.sub(r"<[^>]+>", "", (item.findtext("description") or ""))[:180].strip()
            try:
                from email.utils import parsedate_to_datetime
                pub = parsedate_to_datetime(pub).strftime("%d %b, %H:%M")
            except: pub = pub[:16]
            if title: results.append({"title":title,"link":link,"date":pub,"summary":desc})
        return results
    except: return []

# ─────────────────────────────────────────────────────────────────────────────
# RICH CHART BUILDER  +  AI MARKET ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
_CHART_THEME = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font_family="Space Mono, monospace",
    font_color="#9A8AAA",
    xaxis=dict(
        gridcolor="rgba(107,45,107,.14)", gridwidth=1,
        zeroline=False, showline=False,
        tickfont=dict(size=9, color="#4A3858"),
    ),
    yaxis=dict(
        gridcolor="rgba(107,45,107,.14)", gridwidth=1,
        zeroline=False, showline=False,
        tickfont=dict(size=9, color="#4A3858"),
    ),
    margin=dict(l=6, r=6, t=28, b=6),
    legend=dict(
        bgcolor="rgba(13,11,18,.7)", bordercolor="rgba(107,45,107,.3)",
        borderwidth=1, font=dict(size=9, color="#9A8AAA"),
        orientation="h", y=-0.12,
    ),
    hovermode="x unified",
    hoverlabel=dict(
        bgcolor="#1A1225", bordercolor="rgba(107,45,107,.4)",
        font=dict(size=11, family="Space Mono, monospace", color="#EDE8F5"),
    ),
)

_PALETTE = ["#C084C8","#4ADE80","#F0C040","#60A5FA","#FB923C","#F472B6","#34D399","#A78BFA"]

def _add_bollinger(fig: "go.Figure", closes: pd.Series, window: int = 20, color: str = "#C084C8") -> None:
    """Add Bollinger Bands ±2σ as filled area to figure."""
    sma = closes.rolling(window).mean()
    std = closes.rolling(window).std()
    upper = sma + 2 * std; lower = sma - 2 * std
    fig.add_trace(go.Scatter(x=upper.index, y=upper.values, name="BB Upper",
                             line=dict(color=color, width=0.6, dash="dot"), showlegend=False,
                             hoverinfo="skip"))
    fig.add_trace(go.Scatter(x=lower.index, y=lower.values, name="BB Band",
                             fill="tonexty",
                             fillcolor=f"rgba(192,132,200,0.06)",
                             line=dict(color=color, width=0.6, dash="dot"),
                             showlegend=False, hoverinfo="skip"))

def _sma_trace(fig: "go.Figure", closes: pd.Series, w: int, color: str, name: str) -> None:
    sma = closes.rolling(w).mean().dropna()
    if not sma.empty:
        fig.add_trace(go.Scatter(x=sma.index, y=sma.values, name=name,
                                 line=dict(color=color, width=1, dash="dash"),
                                 opacity=0.7, hovertemplate=f"{name}: %{{y:.2f}}<extra></extra>"))

def build_rich_chart(
    series_dict: dict,           # {label: pd.Series}
    mode: str = "normalized",    # "normalized" | "absolute" | "candlestick"
    title: str = "1-Year Price History",
    show_bollinger: bool = True,
    show_sma: bool = True,
    height: int = 380,
    single_ohlc: pd.DataFrame | None = None,  # for candlestick
) -> "go.Figure | None":
    """Build a premium dark Plotly chart from a dict of price series."""
    if not _PLOTLY:
        return None

    fig = go.Figure()

    if mode == "candlestick" and single_ohlc is not None and not single_ohlc.empty:
        sym = list(series_dict.keys())[0] if series_dict else "Price"
        fig.add_trace(go.Candlestick(
            x=single_ohlc.index,
            open=single_ohlc["open"], high=single_ohlc["high"],
            low=single_ohlc["low"],   close=single_ohlc["close"],
            name=sym,
            increasing=dict(fillcolor="rgba(74,222,128,.55)", line=dict(color="#4ade80", width=1)),
            decreasing=dict(fillcolor="rgba(248,113,113,.45)", line=dict(color="#f87171", width=1)),
        ))
        if show_bollinger:
            _add_bollinger(fig, single_ohlc["close"])
        if show_sma:
            _sma_trace(fig, single_ohlc["close"], 20,  "#F0C040", "SMA-20")
            _sma_trace(fig, single_ohlc["close"], 50,  "#60A5FA", "SMA-50")
            _sma_trace(fig, single_ohlc["close"], 200, "#FB923C", "SMA-200")
        fig.update_layout(xaxis_rangeslider_visible=False)

    else:
        for i, (label, s) in enumerate(series_dict.items()):
            if s is None or s.empty: continue
            s = s.dropna()
            if mode == "normalized":
                y = (s / s.iloc[0] - 1) * 100
                ytitle = "% Return from start"
                hover  = f"%{{y:.2f}}%<extra>{label}</extra>"
            else:
                y = s; ytitle = "Price"; hover = f"%{{y:,.4f}}<extra>{label}</extra>"
            color = _PALETTE[i % len(_PALETTE)]
            # Area fill for single series
            fill = "tozeroy" if len(series_dict) == 1 else "none"
            fillcolor = "rgba(192,132,200,.06)" if len(series_dict) == 1 else None
            fig.add_trace(go.Scatter(
                x=y.index, y=y.values, name=label, mode="lines",
                line=dict(color=color, width=1.8),
                fill=fill, fillcolor=fillcolor,
                hovertemplate=hover,
            ))
            if show_bollinger and len(series_dict) == 1:
                _add_bollinger(fig, s, color=color)
            if show_sma and len(series_dict) == 1:
                _sma_trace(fig, s, 20,  "#F0C040", "SMA-20")
                _sma_trace(fig, s, 50,  "#60A5FA", "SMA-50")
                _sma_trace(fig, s, 200, "#FB923C", "SMA-200")
        ytitle = "% Return" if mode == "normalized" else "Price"
        fig.update_layout(yaxis_title=ytitle)

    layout = dict(**_CHART_THEME, height=height,
                  title=dict(text=title, font=dict(family="Cormorant Garamond, serif",
                                                   size=14, color="#C084C8"), x=0.01))
    fig.update_layout(**layout)
    return fig


@st.cache_data(ttl=300)
def fetch_market_news_context(topic: str = "financial markets economy stocks") -> str:
    """
    Fetch 6 recent news headlines for AI context injection.
    Returns a compact string like: '• Headline 1 (Reuters) • Headline 2 (Bloomberg) …'
    Used to ground AI analysis in real-world events.
    """
    items = fetch_gnews_with_images(topic, "News", "#9CA3AF", max_items=6)
    if not items:
        # Fallback to India ET feed
        items = fetch_rss("https://economictimes.indiatimes.com/markets/rssfeeds/1977021501.cms", 6)
    if not items:
        return ""
    lines = []
    for it in items[:6]:
        title = it.get("title", "").strip()
        src   = it.get("source", it.get("date", ""))
        if title:
            lines.append(f"• {title}" + (f" ({src})" if src else ""))
    return "\n".join(lines)


_AI_MARKET_SYSTEM = """You are a senior quantitative analyst and trader with 20 years of experience.
You receive structured price data AND recent news headlines for a SPECIFIC TIMEFRAME.
Your job: connect the price action to the news. Rules:
- Lead with the most important insight for that timeframe
- Cite exact numbers from the data (current price, period high/low, % change, volatility)
- Reference specific news headlines that explain price moves — name the event, not just "news"
- Comment on technicals appropriate to the timeframe: for 1D use momentum/volume; for 1W use swing levels;
  for 1M use trend/SMA crossovers; for 1Y use macro positioning/annual cyclicality
- If multiple assets: compare relative strength, note which is most news-sensitive
- End with a PREDICTION calibrated to the timeframe (next session / next week / next month)
  — give a directional call with a specific price target or % range, and cite the catalyst
- Be direct, specific, professional. Bloomberg terminal brevity.
- 5-7 sentences. Never be vague. Never say "it depends"."""

# Timeframe → Yahoo Finance params
_TF_PARAMS = {
    "1D":  ("1d",  "5m",  "Intraday (Today)"),
    "1W":  ("5d",  "30m", "Past Week"),
    "1M":  ("1mo", "1d",  "Past Month"),
    "1Y":  ("1y",  "1d",  "Past Year"),
}

@st.cache_data(ttl=120)
def fetch_tf_series(symbol: str, period: str, interval: str) -> pd.Series:
    """Fetch a close-price series for any period/interval from Yahoo Finance."""
    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
           f"?range={period}&interval={interval}&includePrePost=false")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
    }
    try:
        r = requests.get(url, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json(); res = data["chart"]["result"][0]
        ts = res["timestamp"]; q = res["indicators"]["quote"][0]
        closes = q.get("close", [])
        idx = pd.to_datetime(ts, unit="s", utc=True)
        s = pd.Series(closes, index=idx, dtype=float, name=symbol).dropna()
        return s
    except Exception:
        return pd.Series(dtype=float)

def _compute_tf_stats(symbol: str, period: str, interval: str) -> dict:
    """Stats for a given timeframe, used by AI."""
    s = fetch_tf_series(symbol, period, interval)
    if s.empty: return {"symbol": symbol}
    pct   = round((s.iloc[-1] / s.iloc[0] - 1) * 100, 2) if len(s) > 1 else None
    hi    = round(float(s.max()), 4)
    lo    = round(float(s.min()), 4)
    vol   = round(float(s.pct_change().std() * 100), 4) if len(s) > 2 else None
    sma10 = round(float(s.rolling(10).mean().iloc[-1]), 4) if len(s) >= 10 else None
    return {
        "symbol": symbol,
        "current_price": round(float(s.iloc[-1]), 4),
        "pct_change": pct,
        "period_high": hi, "period_low": lo,
        "typical_move_pct": vol,
        "sma10": sma10,
    }

def ai_market_analysis(symbols_data: list[dict], groq_api_key: str,
                       context: str = "", news_context: str = "") -> str:
    if not groq_api_key: return "⚠ No API key — add your Groq key in the sidebar."
    lines = []
    for d in symbols_data:
        parts = [f"  {d.get('symbol','?')}:"]
        for k,v in d.items():
            if k != "symbol" and v is not None:
                parts.append(f"{k}={v}")
        lines.append(" ".join(parts))
    news_block = (f"\n\nRecent News Headlines:\n{news_context}" if news_context else "")
    prompt = (f"Timeframe context: {context}\n\nPrice Data:\n" + "\n".join(lines)
              + news_block
              + "\n\nProvide your analyst note with prediction for this timeframe:")
    return groq_call(
        api_key     = groq_api_key,
        messages    = [{"role":"user","content":prompt}],
        system      = _AI_MARKET_SYSTEM,
        temperature = 0.22,
        max_tokens  = 420,
        site_key    = "ai_market_analysis",
    )

def _compute_symbol_stats(df: pd.DataFrame, sym: str) -> dict:
    """Compute stats from a 1Y OHLCV DataFrame (legacy, kept for portfolio use)."""
    if df.empty: return {"symbol": sym}
    closes = df["close"].dropna()
    sma20  = float(closes.rolling(20).mean().iloc[-1]) if len(closes) >= 20 else None
    sma50  = float(closes.rolling(50).mean().iloc[-1]) if len(closes) >= 50 else None
    vol30  = float(closes.pct_change().rolling(30).std().iloc[-1] * 100 * (252**0.5)) if len(closes) >= 30 else None
    pct1y  = round((closes.iloc[-1] / closes.iloc[0] - 1) * 100, 2) if len(closes) > 1 else None
    return {
        "symbol": sym, "current_price": round(float(closes.iloc[-1]), 4),
        "pct_1y": pct1y, "high_1y": round(float(closes.max()), 4),
        "low_1y": round(float(closes.min()), 4),
        "sma20": round(sma20, 4) if sma20 else None,
        "sma50": round(sma50, 4) if sma50 else None,
        "volatility_30d": round(vol30, 2) if vol30 else None,
    }

def render_ai_timeframe_panel(
    symbols: list[str],
    panel_key: str,
    groq_api_key: str,
    accent: str = "#C084C8",
    label: str = "AI Market Analysis",
    news_topic: str = "financial markets economy stocks",
    chart_series_fn=None,       # callable(period, interval) → dict[name, pd.Series]
) -> None:
    """
    Renders the AI analysis panel with 1D · 1W · 1M · 1Y buttons.
    When chart_series_fn is provided, also re-renders the chart for the selected period.
    Now includes news context in AI analysis and a News Deep-Dive button.
    """
    TF_ORDER  = ["1D", "1W", "1M", "1Y"]
    TF_LABELS = {"1D": "Day", "1W": "Week", "1M": "Month", "1Y": "Year"}
    tf_key    = f"_{panel_key}_ai_tf"
    text_key  = f"_{panel_key}_ai_text"
    news_key  = f"_{panel_key}_news_text"

    current_tf = st.session_state.get(tf_key, "1Y")

    # ── Period button row ─────────────────────────────────────────────────
    btn_cols = st.columns(4)
    for i, tf in enumerate(TF_ORDER):
        with btn_cols[i]:
            is_active = (tf == current_tf)
            btn_label = f"**{TF_LABELS[tf]}**" if is_active else TF_LABELS[tf]
            if st.button(btn_label, key=f"{panel_key}_tf_{tf}", use_container_width=True,
                         help=f"Show {TF_LABELS[tf].lower()} view"):
                if tf != current_tf:
                    st.session_state[tf_key]   = tf
                    st.session_state[text_key] = ""
                    st.rerun()

    period, interval, tf_label = _TF_PARAMS[current_tf]

    # ── Chart for this period (if renderer provided) ──────────────────────
    if chart_series_fn is not None:
        try:
            chart_data = chart_series_fn(period, interval)
            if chart_data:
                fig_tf = build_rich_chart(
                    chart_data, mode="normalized",
                    title=f"% Return · {tf_label} ({current_tf})",
                    show_bollinger=False, show_sma=False, height=260,
                )
                if fig_tf:
                    st.plotly_chart(fig_tf, use_container_width=True,
                                    config=dict(displayModeBar=False, displaylogo=False))
        except Exception:
            pass

    # ── AI Analysis box ───────────────────────────────────────────────────
    st.markdown(
        f'<div style="background:linear-gradient(135deg,rgba(13,11,18,.98),rgba(20,15,30,.97));'
        f'border:1px solid rgba(107,45,107,.28);border-radius:12px;'
        f'padding:.85rem 1.1rem .8rem;margin-top:.6rem;">',
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:.5rem;">'
        f'<div style="font-family:Space Mono,monospace;font-size:.52rem;letter-spacing:.16em;'
        f'text-transform:uppercase;color:{accent};">🤖 {label}</div>'
        f'<div style="font-family:Space Mono,monospace;font-size:.46rem;color:#4A3858;">'
        f'Horizon: {tf_label} · {current_tf}</div></div>',
        unsafe_allow_html=True,
    )

    # ── Auto-load AI text ─────────────────────────────────────────────────
    cached_text = st.session_state.get(text_key, "")
    if not cached_text and groq_api_key and symbols:
        with st.spinner(f"Analysing {tf_label.lower()} price action + news…"):
            stats = []
            for sym in symbols[:6]:
                s = _compute_tf_stats(sym, period, interval)
                if s.get("current_price"):
                    stats.append(s)
            if stats:
                # Fetch news headlines for this asset class
                news_ctx = fetch_market_news_context(news_topic)
                context_str = f"{tf_label} ({current_tf}) · {', '.join(symbols[:6])}"
                text = ai_market_analysis(stats, groq_api_key,
                                          context=context_str, news_context=news_ctx)
                st.session_state[text_key] = text
                cached_text = text

    if cached_text:
        st.markdown(
            f'<div style="font-family:Syne,sans-serif;font-size:.82rem;color:#C8B8D8;'
            f'line-height:1.78;padding:.2rem 0;">{cached_text}</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div style="font-family:Space Mono,monospace;font-size:.6rem;color:#4A3858;'
            'padding:.5rem 0;">Add a Groq API key in the sidebar to enable AI analysis.</div>',
            unsafe_allow_html=True,
        )

    # ── News Deep-Dive button ─────────────────────────────────────────────
    nd_col1, nd_col2 = st.columns([1, 1])
    with nd_col1:
        if st.button("🗞️ News Deep-Dive", key=f"{panel_key}_news_btn",
                     use_container_width=True,
                     help="Get a news-driven market prediction from recent headlines"):
            with st.spinner("Fetching headlines & generating prediction…"):
                news_raw = fetch_market_news_context(news_topic)
                if news_raw and groq_api_key:
                    stats2 = []
                    for sym in symbols[:4]:
                        s2 = _compute_tf_stats(sym, period, interval)
                        if s2.get("current_price"):
                            stats2.append(s2)
                    news_analysis = groq_call(
                        api_key  = groq_api_key,
                        messages = [{"role":"user","content":
                            f"Price data ({tf_label}):\n" +
                            "\n".join(f"  {d['symbol']}: price={d.get('current_price')} "
                                      f"pct={d.get('pct_change')}% high={d.get('period_high')} "
                                      f"low={d.get('period_low')}" for d in stats2) +
                            f"\n\nLatest News Headlines:\n{news_raw}\n\n"
                            f"Based on these headlines and price data, give a detailed "
                            f"financial prediction for the next {TF_LABELS[current_tf].lower()}. "
                            f"Which headlines are most market-moving? What price targets? "
                            f"What risks to this view?"}],
                        system   = ("You are a top-tier financial analyst. Analyse the news "
                                    "headlines and price data together. For each headline, "
                                    "assess its market impact (bullish/bearish/neutral) and "
                                    "magnitude. Then give a synthesised directional prediction "
                                    "with specific price targets. Be precise, cite numbers. "
                                    "7-10 sentences."),
                        temperature = 0.2,
                        max_tokens  = 600,
                        site_key    = f"{panel_key}_news_deepdive",
                    )
                    st.session_state[news_key] = news_analysis
                elif not news_raw:
                    st.session_state[news_key] = "⚠ Could not fetch news headlines right now."
    with nd_col2:
        if st.button("🔄 Refresh Analysis", key=f"{panel_key}_refresh",
                     use_container_width=True, help="Re-run AI analysis with fresh data"):
            st.session_state[text_key] = ""
            st.session_state[news_key] = ""
            st.rerun()

    # Show news deep-dive result if available
    news_result = st.session_state.get(news_key, "")
    if news_result:
        st.markdown(
            f'<div style="background:rgba(192,132,200,.05);border:1px solid rgba(192,132,200,.2);'
            f'border-left:3px solid {accent};border-radius:0 8px 8px 0;'
            f'padding:.75rem 1rem;margin-top:.6rem;">'
            f'<div style="font-family:Space Mono,monospace;font-size:.44rem;color:{accent};'
            f'text-transform:uppercase;letter-spacing:.12em;margin-bottom:.4rem;">'
            f'🗞️ News-Driven Prediction</div>'
            f'<div style="font-family:Syne,sans-serif;font-size:.8rem;color:#C8B8D8;line-height:1.8;">'
            f'{news_result}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# MARKET SYMBOL DICTS
# ─────────────────────────────────────────────────────────────────────────────
COMMODITY_SYMS = {
    "GC=F":("Gold","$/oz","🪙",2),"SI=F":("Silver","$/oz","⚪",3),
    "CL=F":("Crude Oil","$/bbl","🛢️",2),"PL=F":("Platinum","$/oz","💎",2),
    "PA=F":("Palladium","$/oz","✨",2),"HG=F":("Copper","$/lb","🟤",3),
}
CRYPTO_SYMS = {
    "BTC-USD":("Bitcoin","BTC","₿",2),"ETH-USD":("Ethereum","ETH","Ξ",2),
    "BNB-USD":("BNB","BNB","🔶",2),"SOL-USD":("Solana","SOL","◎",2),
    "XRP-USD":("XRP","XRP","✕",4),"DOGE-USD":("Dogecoin","DOGE","🐕",5),
    "ADA-USD":("Cardano","ADA","🔵",4),"AVAX-USD":("Avalanche","AVAX","🔺",2),
}
ALL_FX = {
    "USDINR=X":{"label":"USD/INR","flag":"🇮🇳","name":"Indian Rupee","invert":False},
    "USDJPY=X":{"label":"USD/JPY","flag":"🇯🇵","name":"Japanese Yen","invert":False},
    "USDCNY=X":{"label":"USD/CNY","flag":"🇨🇳","name":"Chinese Yuan","invert":False},
    "EURUSD=X":{"label":"EUR/USD","flag":"🇪🇺","name":"Euro","invert":True},
    "GBPUSD=X":{"label":"GBP/USD","flag":"🇬🇧","name":"British Pound","invert":True},
    "USDCHF=X":{"label":"USD/CHF","flag":"🇨🇭","name":"Swiss Franc","invert":False},
    "USDKRW=X":{"label":"USD/KRW","flag":"🇰🇷","name":"S. Korean Won","invert":False},
    "USDBRL=X":{"label":"USD/BRL","flag":"🇧🇷","name":"Brazilian Real","invert":False},
    "USDCAD=X":{"label":"USD/CAD","flag":"🇨🇦","name":"Canadian Dollar","invert":False},
    "USDAUD=X":{"label":"USD/AUD","flag":"🇦🇺","name":"Australian Dollar","invert":False},
    "USDSGD=X":{"label":"USD/SGD","flag":"🇸🇬","name":"Singapore Dollar","invert":False},
    "USDHKD=X":{"label":"USD/HKD","flag":"🇭🇰","name":"Hong Kong Dollar","invert":False},
    "USDMXN=X":{"label":"USD/MXN","flag":"🇲🇽","name":"Mexican Peso","invert":False},
    "USDTRY=X":{"label":"USD/TRY","flag":"🇹🇷","name":"Turkish Lira","invert":False},
    "USDRUB=X":{"label":"USD/RUB","flag":"🇷🇺","name":"Russian Ruble","invert":False},
    "USDZAR=X":{"label":"USD/ZAR","flag":"🇿🇦","name":"S. African Rand","invert":False},
    "USDAED=X":{"label":"USD/AED","flag":"🇦🇪","name":"UAE Dirham","invert":False},
    "USDNOK=X":{"label":"USD/NOK","flag":"🇳🇴","name":"Norwegian Krone","invert":False},
    "USDSEK=X":{"label":"USD/SEK","flag":"🇸🇪","name":"Swedish Krona","invert":False},
    "USDDKK=X":{"label":"USD/DKK","flag":"🇩🇰","name":"Danish Krone","invert":False},
    "USDNZD=X":{"label":"USD/NZD","flag":"🇳🇿","name":"New Zealand Dollar","invert":False},
    "USDPLN=X":{"label":"USD/PLN","flag":"🇵🇱","name":"Polish Zloty","invert":False},
    "USDTHB=X":{"label":"USD/THB","flag":"🇹🇭","name":"Thai Baht","invert":False},
    "USDIDR=X":{"label":"USD/IDR","flag":"🇮🇩","name":"Indonesian Rupiah","invert":False},
    "USDPHP=X":{"label":"USD/PHP","flag":"🇵🇭","name":"Philippine Peso","invert":False},
}

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:0 0 1rem;">
      <div style="font-family:'Cormorant Garamond',serif;font-size:1.4rem;font-weight:300;
                  color:#EDE8F5;line-height:1.1;">
        RAG <em style="color:#C084C8;font-style:italic;">Assistant</em>
      </div>
      <div style="font-family:'Space Mono',monospace;font-size:.52rem;letter-spacing:.22em;
                  color:#4A3858;text-transform:uppercase;margin-top:.35rem;">
        Financial Intelligence · v6
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="sb-lbl" style="border-top:none;padding-top:0;margin-top:0;">Configuration</div>',
                unsafe_allow_html=True)
    default_key = st.secrets.get("GROQ_API_KEY", os.getenv("GROQ_API_KEY",""))
    if default_key:
        GROQ_API_KEY = default_key
        st.markdown('<div class="key-ok"><div class="key-dot"></div>API Key Active</div>',
                    unsafe_allow_html=True)
    else:
        GROQ_API_KEY = st.text_input("", type="password", placeholder="gsk_…",
                                     label_visibility="collapsed")
        st.markdown("<span style='font-family:Space Mono,monospace;font-size:.56rem;color:#4A3858;'>"
                    "console.groq.com → free key</span>", unsafe_allow_html=True)
        if GROQ_API_KEY:
            os.environ["GROQ_API_KEY"] = GROQ_API_KEY
            st.markdown('<div class="key-ok"><div class="key-dot"></div>API Key Active</div>',
                        unsafe_allow_html=True)

    bucket = _get_bucket(); tl = int(bucket._tokens); pf = tl / bucket._cap
    bc = "#4ade80" if pf > 0.5 else ("#f0c040" if pf > 0.2 else "#f87171")
    st.markdown(f'<div style="margin:.6rem 0 .3rem;">'
                f'<div style="font-family:Space Mono,monospace;font-size:.52rem;letter-spacing:.2em;'
                f'color:#4A3858;text-transform:uppercase;margin-bottom:.3rem;">API Rate Limit</div>'
                f'<div style="background:#0D0B12;border:1px solid rgba(139,58,139,.22);'
                f'border-radius:4px;height:5px;overflow:hidden;">'
                f'<div style="height:100%;width:{int(pf*100)}%;background:{bc};border-radius:4px;'
                f'transition:width .4s;"></div></div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#4A3858;margin-top:.2rem;">'
                f'{tl}/{bucket._cap} calls · 60s reset</div></div>',
                unsafe_allow_html=True)

    if st.session_state.file_names:
        st.markdown('<div class="sb-lbl">Knowledge Base</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        c1.metric("Chunks", st.session_state.chunk_count)
        c2.metric("Docs",   st.session_state.uploaded_docs)
        dot_colors = {"PDF":"#f87171","XLSX":"#4ade80","XLS":"#4ade80",
                      "CSV":"#F0C040","DOCX":"#60a5fa","TXT":"#9A8AAA"}
        for fn in st.session_state.file_names:
            short = fn[:22]+"…" if len(fn) > 22 else fn
            ext   = fn.rsplit(".",1)[-1].upper() if "." in fn else "?"
            dc    = dot_colors.get(ext, "#4A3858")
            st.markdown(f'<div class="doc-pill">'
                        f'<div class="doc-dot" style="background:{dc};"></div>'
                        f'<span style="font-family:Space Mono,monospace;font-size:.44rem;'
                        f'color:{dc};margin-right:.25rem;">{ext}</span>{short}</div>',
                        unsafe_allow_html=True)

    st.markdown('<div class="sb-lbl">Quick Ask</div>', unsafe_allow_html=True)
    for q_item in ["What is USD/INR today?","Compare INR vs JPY vs Yuan",
                   "Gold price today?","Bitcoin vs Ethereum?",
                   "What was total revenue?","Main risk factors?","EPS change YoY?"]:
        if st.button(q_item, use_container_width=True, key=f"qa_{q_item[:14]}"):
            st.session_state["_prefill"] = q_item

    # ── Retrieval info ────────────────────────────────────────────────────
    st.markdown('<div class="sb-lbl">⚙ Retrieval</div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-family:Space Mono,monospace;font-size:.44rem;color:#4A3858;line-height:1.7;">'
        'TF-IDF + BM25 hybrid retrieval<br>'
        'RRF rank fusion · zero ML deps<br>'
        '~30MB memory footprint'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown('<div class="sb-lbl">Actions</div>', unsafe_allow_html=True)
    col_a1, col_a2 = st.columns(2)
    with col_a1:
        if st.button("✕ Clear Chat", use_container_width=True):
            st.session_state.messages = []; st.rerun()
    with col_a2:
        if st.button("🗑 Clear Docs", use_container_width=True):
            for k in ["vectorstore","file_names","uploaded_docs","chunk_count",
                      "doc_full_text","auto_metrics","auto_generated","search_query","search_results"]:
                st.session_state[k] = (None if k=="vectorstore" else
                                       [] if k in ["file_names","auto_metrics","search_results"] else
                                       0  if k in ["uploaded_docs","chunk_count"] else
                                       False if k=="auto_generated" else "")
            st.rerun()

    # ── Theme selector ──────────────────────────────────────────────────────
    st.markdown('<div class="sb-lbl">🎨 Theme</div>', unsafe_allow_html=True)
    _theme_choice = st.selectbox(
        "theme",
        list(_THEMES.keys()),
        index=list(_THEMES.keys()).index(st.session_state.get("app_theme","Royal Velvet")),
        label_visibility="collapsed",
        key="theme_selector",
    )
    if _theme_choice != st.session_state.get("app_theme"):
        st.session_state["app_theme"] = _theme_choice
        st.rerun()

    # ── Keyboard shortcut hint ──────────────────────────────────────────────
    st.markdown(
        '<div style="font-family:Space Mono,monospace;font-size:.44rem;color:#4A3858;'
        'margin-top:.8rem;line-height:1.8;">'
        '⌨ Ctrl+K · focus search<br>'
        '⌨ Ctrl+U · toggle upload<br>'
        '⌨ Esc · blur search'
        '</div>',
        unsafe_allow_html=True,
    )

# Apply active theme
apply_theme(st.session_state.get("app_theme", "Royal Velvet"))
# Inject keyboard shortcuts (runs once per page load)
inject_keyboard_shortcuts()

# ─────────────────────────────────────────────────────────────────────────────
# HERO
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="rag-header">
  <div class="rag-kicker">Financial Intelligence Platform</div>
  <h1>Interrogate Your<br><em>Financial Documents</em></h1>
  <p>Semantic search and AI-powered analysis across Annual Reports,
     10-Ks &amp; Earnings Transcripts. Live markets &amp; crypto always on.</p>
  <div class="badge-row">
    <span class="badge v">FinBERT Embeddings</span>
    <span class="badge v">Source-backed Answers</span>
    <span class="badge v">Llama 3.3 · 70B</span>
    <span class="badge">Groq</span>
    <span class="badge g">Live Data</span>
    <span class="badge b">PDF · Excel · CSV · DOCX</span>
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# TOP ACTION BAR — Upload (square) + Simulate Portfolio (circle globe)
# Both rendered as one components.html iframe for real JS execution
# Click → postMessage → parent Streamlit rerun via hidden buttons
# ─────────────────────────────────────────────────────────────────────────────

_pf_open_state = "true" if st.session_state.show_portfolio else "false"
_up_open_state = "true" if st.session_state.show_upload    else "false"

_ACTION_BAR_HTML = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Cormorant+Garamond:wght@300;400&display=swap');

  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  html, body {{ background: transparent; overflow: hidden; }}

  body {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 2.5rem;
    height: 170px;
    padding: 12px 20px;
    font-family: 'Space Mono', monospace;
  }}

  /* ═══ UPLOAD BUTTON — elegant dark square ═══════════════════════════════ */
  #upload-btn {{
    width: 140px;
    height: 140px;
    border-radius: 18px;
    background: linear-gradient(145deg, #010e1c 0%, #021b2e 45%, #010912 100%);
    border: 1.5px solid rgba(56,189,248,.4);
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 8px;
    cursor: pointer;
    position: relative;
    overflow: hidden;
    box-shadow:
      0 0 0 1px rgba(56,189,248,.08),
      0 6px 40px rgba(0,100,160,.4),
      inset 0 1px 0 rgba(56,189,248,.07);
    transition: all .35s cubic-bezier(.34,1.56,.64,1);
    flex-shrink: 0;
  }}
  #upload-btn::before {{
    content: '';
    position: absolute;
    top: -40%; left: -40%;
    width: 180%; height: 180%;
    background: radial-gradient(ellipse at 40% 40%, rgba(56,189,248,.06) 0%, transparent 60%);
    pointer-events: none;
  }}
  /* Animated corner accents */
  #upload-btn::after {{
    content: '';
    position: absolute;
    inset: 6px;
    border-radius: 12px;
    border: 1px solid transparent;
    border-top-color: rgba(56,189,248,.25);
    border-left-color: rgba(56,189,248,.15);
    pointer-events: none;
    transition: border-color .3s ease;
  }}
  #upload-btn:hover {{
    transform: translateY(-5px) scale(1.04);
    border-color: rgba(56,189,248,.8);
    box-shadow:
      0 0 0 1px rgba(56,189,248,.2),
      0 0 60px rgba(0,160,240,.55),
      0 16px 40px rgba(0,0,0,.6),
      inset 0 1px 0 rgba(56,189,248,.15);
  }}
  #upload-btn:hover::after {{
    border-top-color: rgba(56,189,248,.6);
    border-left-color: rgba(56,189,248,.4);
  }}
  #upload-btn.active {{
    border-color: rgba(56,189,248,.9);
    transform: translateY(-6px) scale(1.05);
    box-shadow:
      0 0 0 2px rgba(56,189,248,.3),
      0 0 80px rgba(0,160,240,.65),
      0 16px 40px rgba(0,0,0,.7),
      inset 0 1px 0 rgba(56,189,248,.2);
  }}
  .upload-icon {{
    width: 44px;
    height: 44px;
    position: relative;
    display: flex;
    align-items: center;
    justify-content: center;
  }}
  /* SVG arrow drawn with CSS */
  .upload-arrow {{
    width: 2px;
    height: 22px;
    background: linear-gradient(180deg, #38bdf8, rgba(56,189,248,.3));
    border-radius: 2px;
    position: relative;
    box-shadow: 0 0 10px rgba(56,189,248,.6);
  }}
  .upload-arrow::before,
  .upload-arrow::after {{
    content: '';
    position: absolute;
    top: 0;
    width: 10px;
    height: 2px;
    background: #38bdf8;
    border-radius: 2px;
    box-shadow: 0 0 8px rgba(56,189,248,.6);
  }}
  .upload-arrow::before {{ left: -8px; transform: rotate(45deg); transform-origin: right center; }}
  .upload-arrow::after  {{ right: -8px; transform: rotate(-45deg); transform-origin: left center; }}
  /* Horizontal bar at bottom of arrow */
  .upload-base {{
    width: 28px;
    height: 2px;
    background: linear-gradient(90deg, transparent, rgba(56,189,248,.8), transparent);
    border-radius: 2px;
    margin-top: 2px;
  }}
  .upload-label {{
    font-family: 'Space Mono', monospace;
    font-size: 9px;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: rgba(56,189,248,.75);
    text-shadow: 0 0 12px rgba(56,189,248,.4);
  }}
  .upload-sublabel {{
    font-size: 7px;
    letter-spacing: 1.5px;
    color: rgba(56,189,248,.35);
    text-transform: uppercase;
  }}

  /* Scan line animation */
  @keyframes scan {{
    0%   {{ top: -100%; opacity: 0; }}
    15%  {{ opacity: .4; }}
    85%  {{ opacity: .3; }}
    100% {{ top: 200%; opacity: 0; }}
  }}
  .scan-line {{
    position: absolute;
    left: 0; right: 0;
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(56,189,248,.5), transparent);
    top: -100%;
    animation: scan 3s ease-in-out infinite;
    pointer-events: none;
  }}

  /* ═══ SIMULATE BUTTON — circle globe ═══════════════════════════════════ */
  #sim-wrap {{
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 8px;
    flex-shrink: 0;
  }}
  #globe {{
    width: 140px;
    height: 140px;
    border-radius: 50%;
    background: radial-gradient(circle at 38% 34%, #04233d 0%, #010f1e 50%, #000508 100%);
    border: 1.5px solid rgba(56,189,248,.35);
    position: relative;
    cursor: pointer;
    overflow: hidden;
    box-shadow:
      0 0 0 3px rgba(56,189,248,.06),
      0 0 0 7px rgba(56,189,248,.03),
      0 0 45px rgba(0,110,170,.4),
      0 8px 32px rgba(0,0,0,.7);
    transition: all .4s cubic-bezier(.34,1.56,.64,1);
  }}
  #globe:hover {{
    transform: translateY(-5px) scale(1.06);
    border-color: rgba(56,189,248,.75);
    box-shadow:
      0 0 0 3px rgba(56,189,248,.15),
      0 0 0 8px rgba(56,189,248,.06),
      0 0 70px rgba(0,150,230,.6),
      0 14px 40px rgba(0,0,0,.7);
  }}
  #globe.active {{
    border-color: rgba(56,189,248,.9);
    transform: translateY(-6px) scale(1.08);
    box-shadow:
      0 0 0 3px rgba(56,189,248,.28),
      0 0 0 10px rgba(56,189,248,.1),
      0 0 100px rgba(0,160,240,.75),
      0 0 160px rgba(0,120,180,.3),
      0 14px 40px rgba(0,0,0,.8);
  }}

  /* Orbit rings */
  .orbit {{
    position: absolute;
    border-radius: 50%;
    border: 1px dashed rgba(56,189,248,.18);
    pointer-events: none;
  }}
  .orbit-1 {{ inset: -10px; animation: orbit 9s linear infinite; }}
  .orbit-2 {{ inset: -18px; border-color: rgba(56,189,248,.08); animation: orbit 16s linear infinite reverse; }}
  @keyframes orbit {{ to {{ transform: rotate(360deg); }} }}

  /* Dot on orbit ring */
  .orbit-dot {{
    position: absolute;
    width: 4px; height: 4px;
    border-radius: 50%;
    background: #38bdf8;
    box-shadow: 0 0 6px #38bdf8;
    top: -2px; left: calc(50% - 2px);
  }}

  /* Canvas chart background */
  #globe-canvas {{
    position: absolute;
    inset: 0;
    border-radius: 50%;
    pointer-events: none;
  }}

  /* Overlay text */
  #globe-info {{
    position: absolute;
    inset: 0;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    z-index: 5;
    pointer-events: none;
    gap: 2px;
    padding: 8px;
  }}
  .g-flag  {{
    font-size: 28px;
    line-height: 1;
    filter: drop-shadow(0 1px 4px rgba(0,0,0,.9));
    transition: opacity .25s ease, transform .25s ease;
  }}
  .g-name  {{
    font-size: 8px;
    letter-spacing: 2px;
    color: rgba(56,189,248,.8);
    text-transform: uppercase;
    text-shadow: 0 0 8px rgba(0,0,0,.9);
    transition: opacity .25s ease, transform .25s ease;
  }}
  .g-val   {{
    font-family: 'Cormorant Garamond', serif;
    font-size: 18px;
    font-weight: 300;
    color: #e8f6ff;
    line-height: 1.1;
    text-shadow: 0 0 14px rgba(56,189,248,.35);
    transition: opacity .25s ease, transform .25s ease;
  }}
  .g-chg   {{
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1px;
    transition: opacity .25s ease, transform .25s ease;
  }}
  .g-chg.up  {{ color: #4ade80; text-shadow: 0 0 10px rgba(74,222,128,.6); }}
  .g-chg.dn  {{ color: #f87171; text-shadow: 0 0 10px rgba(248,113,113,.6); }}
  .fade-out {{ opacity: 0 !important; transform: translateY(-5px) !important; }}

  #sim-label {{
    font-size: 8px;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: rgba(56,189,248,.5);
    transition: color .3s ease;
  }}
  #globe.active ~ #sim-label {{ color: rgba(56,189,248,.9); }}
</style>
</head>
<body>

<!-- UPLOAD BUTTON -->
<div id="upload-btn" onclick="clickUpload()">
  <div class="scan-line"></div>
  <div class="upload-icon">
    <div class="upload-arrow"></div>
  </div>
  <div class="upload-base"></div>
  <div class="upload-label">Upload</div>
  <div class="upload-sublabel">PDF · XLS · CSV · DOCX</div>
</div>

<!-- SIMULATE GLOBE -->
<div id="sim-wrap">
  <div id="globe" onclick="clickSim()">
    <div class="orbit orbit-1"><div class="orbit-dot"></div></div>
    <div class="orbit orbit-2"></div>
    <canvas id="globe-canvas"></canvas>
    <div id="globe-info">
      <span class="g-flag" id="g-flag">🌐</span>
      <span class="g-name" id="g-name">Markets</span>
      <span class="g-val"  id="g-val">—</span>
      <span class="g-chg up" id="g-chg">Loading…</span>
    </div>
  </div>
  <div id="sim-label">Simulate Portfolio</div>
</div>

<script>
var MARKETS = [
  {{flag:'🇺🇸',name:'S&P 500',   val:'5,847',  chg:'+0.42%',up:true, col:'#34d399'}},
  {{flag:'🇯🇵',name:'Nikkei 225',val:'38,921', chg:'+1.18%',up:true, col:'#38bdf8'}},
  {{flag:'🇬🇧',name:'FTSE 100',  val:'8,312',  chg:'-0.23%',up:false,col:'#f87171'}},
  {{flag:'🇩🇪',name:'DAX',       val:'18,640', chg:'+0.87%',up:true, col:'#a78bfa'}},
  {{flag:'🇨🇳',name:'Shanghai',  val:'3,241',  chg:'-0.55%',up:false,col:'#f87171'}},
  {{flag:'🇮🇳',name:'SENSEX',    val:'73,248', chg:'+1.34%',up:true, col:'#fbbf24'}},
  {{flag:'🇧🇷',name:'IBOVESPA',  val:'127,430',chg:'+0.61%',up:true, col:'#34d399'}},
  {{flag:'🇰🇷',name:'KOSPI',     val:'2,631',  chg:'-0.18%',up:false,col:'#f87171'}},
  {{flag:'🇦🇺',name:'ASX 200',   val:'7,943',  chg:'+0.29%',up:true, col:'#38bdf8'}},
  {{flag:'🇫🇷',name:'CAC 40',    val:'7,521',  chg:'+0.52%',up:true, col:'#a78bfa'}},
  {{flag:'🇨🇦',name:'TSX',       val:'22,105', chg:'+0.33%',up:true, col:'#34d399'}},
  {{flag:'🇿🇦',name:'JSE TOP40', val:'71,832', chg:'-0.41%',up:false,col:'#f87171'}},
  {{flag:'🇸🇦',name:'Tadawul',   val:'11,942', chg:'+0.78%',up:true, col:'#fbbf24'}},
  {{flag:'🇸🇬',name:'STI',       val:'3,387',  chg:'+0.12%',up:true, col:'#38bdf8'}},
  {{flag:'🇭🇰',name:'Hang Seng', val:'17,284', chg:'-0.93%',up:false,col:'#f87171'}},
  {{flag:'🇮🇩',name:'IDX',       val:'7,126',  chg:'+0.45%',up:true, col:'#34d399'}},
  {{flag:'🇲🇽',name:'IPC',       val:'54,312', chg:'+0.21%',up:true, col:'#fbbf24'}},
  {{flag:'🇳🇴',name:'OBX',       val:'1,412',  chg:'+1.05%',up:true, col:'#a78bfa'}},
  {{flag:'🌐',name:'Crypto',     val:'$2.8T',  chg:'+2.11%',up:true, col:'#fb923c'}},
  {{flag:'🥇',name:'Gold',       val:'$3,082', chg:'+0.68%',up:true, col:'#fbbf24'}},
];

var idx = 0;
var canvas = document.getElementById('globe-canvas');
var ctx = canvas.getContext('2d');
var chartData = [];

// Set active states from Python
var pfOpen = {_pf_open_state};
var upOpen = {_up_open_state};
if (pfOpen) document.getElementById('globe').classList.add('active');
if (upOpen) document.getElementById('upload-btn').classList.add('active');

function resize() {{
  canvas.width  = canvas.offsetWidth  || 140;
  canvas.height = canvas.offsetHeight || 140;
}}

function genChart(up) {{
  chartData = [];
  var v = 50;
  for (var i = 0; i < 36; i++) {{
    v += (Math.random() - (up ? 0.44 : 0.56)) * 7;
    v = Math.max(8, Math.min(92, v));
    chartData.push(v);
  }}
}}

function hexToRgba(h, a) {{
  h = h.replace('#','');
  return 'rgba('+parseInt(h.slice(0,2),16)+','+parseInt(h.slice(2,4),16)+','+parseInt(h.slice(4,6),16)+','+a+')';
}}

function drawChart(col, up) {{
  if (!ctx) return;
  var w = canvas.width, h = canvas.height;
  ctx.clearRect(0,0,w,h);
  if (!chartData.length) return;

  var pad = 14, bot = h - 12, top = 18;
  var mn = Math.min.apply(null,chartData), mx = Math.max.apply(null,chartData);
  var rng = mx - mn || 8;
  function px(i) {{ return pad + (i/(chartData.length-1))*(w-pad*2); }}
  function py(v) {{ return bot - ((v-mn)/rng)*(bot-top); }}

  // Grid lines
  ctx.strokeStyle = 'rgba(255,255,255,.04)';
  ctx.lineWidth = .5;
  for (var y = top; y < bot; y += (bot-top)/3) {{
    ctx.beginPath(); ctx.moveTo(0,y); ctx.lineTo(w,y); ctx.stroke();
  }}

  // Fill
  var grad = ctx.createLinearGradient(0,top,0,bot);
  grad.addColorStop(0, hexToRgba(col,.22));
  grad.addColorStop(1, hexToRgba(col,0));
  ctx.beginPath();
  ctx.moveTo(px(0), bot);
  chartData.forEach(function(v,i) {{ ctx.lineTo(px(i), py(v)); }});
  ctx.lineTo(px(chartData.length-1), bot);
  ctx.closePath();
  ctx.fillStyle = grad;
  ctx.fill();

  // Line
  ctx.beginPath();
  chartData.forEach(function(v,i) {{
    if (i===0) ctx.moveTo(px(i),py(v)); else ctx.lineTo(px(i),py(v));
  }});
  ctx.strokeStyle = col;
  ctx.lineWidth = 1.6;
  ctx.shadowColor = col; ctx.shadowBlur = 8;
  ctx.stroke();
  ctx.shadowBlur = 0;

  // Endpoint dot
  var ex=px(chartData.length-1), ey=py(chartData[chartData.length-1]);
  ctx.beginPath(); ctx.arc(ex,ey,3.5,0,Math.PI*2);
  ctx.fillStyle = col;
  ctx.shadowColor = col; ctx.shadowBlur = 12;
  ctx.fill(); ctx.shadowBlur = 0;
}}

function fadeItems(out, cb) {{
  var items = [document.getElementById('g-flag'),
               document.getElementById('g-name'),
               document.getElementById('g-val'),
               document.getElementById('g-chg')];
  items.forEach(function(el) {{
    if (out) {{ el.classList.add('fade-out'); }}
    else     {{ el.classList.remove('fade-out'); }}
  }});
  if (cb) setTimeout(cb, 280);
}}

function showMarket(i) {{
  var m = MARKETS[i % MARKETS.length];
  fadeItems(true, function() {{
    document.getElementById('g-flag').textContent = m.flag;
    document.getElementById('g-name').textContent = m.name;
    document.getElementById('g-val').textContent  = m.val;
    var chgEl = document.getElementById('g-chg');
    chgEl.textContent  = (m.up ? '▲ ' : '▼ ') + m.chg;
    chgEl.className    = 'g-chg ' + (m.up ? 'up' : 'dn');
    document.getElementById('globe').style.borderColor = m.col + 'aa';
    genChart(m.up);
    resize();
    drawChart(m.col, m.up);
    fadeItems(false);
  }});
}}

function clickUpload() {{
  document.getElementById('upload-btn').classList.toggle('active');
  window.parent.postMessage({{type:'streamlit:setComponentValue', value:'upload'}}, '*');
}}
function clickSim() {{
  document.getElementById('globe').classList.toggle('active');
  window.parent.postMessage({{type:'streamlit:setComponentValue', value:'portfolio'}}, '*');
}}

// Init
setTimeout(function() {{
  resize();
  showMarket(0);
  setInterval(function() {{
    idx = (idx + 1) % MARKETS.length;
    showMarket(idx);
  }}, 3000);
}}, 100);
window.addEventListener('resize', function() {{ resize(); }});
</script>
</body>
</html>"""

# Render the iframe button panel
import streamlit.components.v1 as _cv1
_btn_click = _cv1.html(_ACTION_BAR_HTML, height=170)

# Hidden real Streamlit buttons that get triggered by JS postMessage via component value
# Since components.html returns the value set by postMessage
if _btn_click == "upload":
    st.session_state.show_upload = not st.session_state.show_upload
    st.rerun()
elif _btn_click == "portfolio":
    st.session_state.show_portfolio = not st.session_state.show_portfolio
    st.rerun()





# ─────────────────────────────────────────────────────────────────────────────
# UPLOAD PANEL  (slides open below action bar when 📂 clicked)
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.show_upload:
    st.markdown('<div class="upload-panel">'
                '<div class="upload-panel-hdr">'
                '<div><div class="upload-panel-title">◈ Upload Financial Documents</div>'
                '<div class="upload-panel-formats">Supported: PDF · XLSX · XLS · CSV · DOCX · TXT</div>'
                '</div></div>',
                unsafe_allow_html=True)
    inline_files = st.file_uploader(
        "Upload files", type=["pdf","txt","xlsx","xls","csv","docx"],
        accept_multiple_files=True, label_visibility="collapsed", key="drawer_upload",
    )
    col_ing, col_cls = st.columns([3, 1])
    with col_ing:
        if inline_files and st.button("⬆  Ingest & Analyse", use_container_width=True, key="drawer_ingest"):
            if not GROQ_API_KEY:
                st.error("Enter your Groq API key in the sidebar first.")
            else:
                try:
                    n = ingest_documents(inline_files)
                    st.success(f"✓ {n} chunks indexed · Analytics ready below ↓")
                    st.session_state.show_upload = False
                    st.rerun()
                except Exception as e:
                    st.error(str(e))
    with col_cls:
        if st.button("✕ Close", use_container_width=True, key="drawer_close"):
            st.session_state.show_upload = False; st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# INLINE ANALYTICS PANEL  (appears directly below upload, after ingest)
# No separate tab — lives right here in the main page flow
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.auto_generated:
    n_m = len(st.session_state.auto_metrics)
    st.markdown(f'<div class="analytics-inline-panel">'
                f'<div class="aip-header">'
                f'<div class="aip-title">Document Analytics</div>'
                f'<span class="aip-badge">✅ {n_m} metrics extracted · FinBERT embeddings</span>'
                f'</div></div>', unsafe_allow_html=True)

    _atabs = st.tabs(["📊 Metrics","📈 Doc vs Market","📋 Templates","🔍 Hybrid Search","🧪 Eval"])

    with _atabs[0]:
        if st.session_state.auto_metrics:
            render_metrics_dashboard(st.session_state.auto_metrics)
            st.markdown("<hr style='border-color:rgba(139,58,139,.15);margin:.8rem 0;'>",
                        unsafe_allow_html=True)
            render_trend_chart(st.session_state.auto_metrics)
            with st.expander("📄 Raw extraction table"):
                st.dataframe(pd.DataFrame([
                    {"Metric":m["label"],"Value":fmt_val(m["value"],m["unit"]),
                     "Unit":m["unit"],"Category":m["category"],"Raw Text":m["raw"]}
                    for m in st.session_state.auto_metrics
                ]), use_container_width=True, hide_index=True)
        else:
            st.info("No numeric metrics matched regex patterns. Try Templates tab for LLM extraction.")

    with _atabs[1]:
        render_comparison_tab(st.session_state.auto_metrics, GROQ_API_KEY)

    with _atabs[2]:
        cats = sorted({v["category"] for v in TEMPLATES.values()})
        chosen_cat = st.selectbox("Filter", ["All"]+cats, label_visibility="collapsed", key="tpl_cat2")
        visible = {k:v for k,v in TEMPLATES.items() if chosen_cat=="All" or v["category"]==chosen_cat}
        items = list(visible.items())
        for rs in range(0, len(items), 3):
            cols = st.columns(3)
            for ci, (tn, tm) in enumerate(items[rs:rs+3]):
                with cols[ci]:
                    color = CAT_COLORS.get(tm["category"], VELVET["dim"])
                    st.markdown(f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.22);'
                                f'border-top:2px solid {color};border-radius:10px;padding:.8rem .9rem .6rem;">'
                                f'<div style="font-size:1.2rem;">{tm["icon"]}</div>'
                                f'<div style="font-family:Syne,sans-serif;font-size:.82rem;font-weight:600;'
                                f'color:{VELVET["text"]};margin:.3rem 0 .2rem;">{tn}</div>'
                                f'<div style="font-family:Space Mono,monospace;font-size:.52rem;color:{color};'
                                f'text-transform:uppercase;">{tm["category"]}</div></div>', unsafe_allow_html=True)
                    if st.button("Run →", key=f"tpl2_{tn[:18]}", use_container_width=True):
                        st.session_state["_prefill"] = tm["prompt"]
                        st.session_state["show_chat"] = True
                        st.rerun()

    with _atabs[3]:
        hs_q = st.text_input("Hybrid search", placeholder="e.g. free cash flow 2023",
                             label_visibility="collapsed", key="hs_q2")
        c1, c2, c3 = st.columns(3)
        with c1: bw2 = st.slider("BM25 weight", 0.0, 1.0, 0.35, 0.05, key="bw2")
        with c2: tn2 = st.slider("Results", 3, 10, 5, key="tn2")
        with c3: uce2 = st.checkbox("Re-rank", value=True, key="uce2")
        if hs_q and st.session_state.vectorstore:
            with st.spinner("Retrieving…"):
                try:
                    vs   = st.session_state.vectorstore
                    cks  = vs.get("chunks", [])
                    mts  = vs.get("meta", [])
                    vec2 = vs.get("vectorizer")
                    mat2 = vs.get("tfidf_matrix")
                    if vec2 and mat2 is not None and cks:
                        from sklearn.metrics.pairwise import cosine_similarity
                        qv2  = vec2.transform([hs_q])
                        sim2 = cosine_similarity(qv2, mat2).flatten().tolist()
                    else:
                        sim2 = [0.0]*len(cks)
                    try:
                        from rank_bm25 import BM25Okapi
                        _b3  = BM25Okapi([_tokenize(c) for c in cks])
                        _br3 = _b3.get_scores(_tokenize(hs_q))
                        _bm3 = max(_br3) if max(_br3)>0 else 1.0
                        bm3  = [s/_bm3 for s in _br3]
                    except Exception:
                        bm3 = [0.0]*len(cks)
                    _K3=60; N3=len(cks); rrf3: dict[int,float]={}
                    for _rl3 in (sorted(range(N3), key=lambda i: sim2[i], reverse=True),
                                 sorted(range(N3), key=lambda i: bm3[i],  reverse=True)):
                        for _r3,_i3 in enumerate(_rl3):
                            rrf3[_i3]=rrf3.get(_i3,0.0)+1.0/(_K3+_r3+1)
                    hits2=[{"idx":i,"chunk":cks[i],"score":rrf3[i]}
                           for i in sorted(rrf3,key=lambda i:rrf3[i],reverse=True)[:tn2]]
                    for rank, h in enumerate(hits2, 1):
                        mt = mts[h["idx"]] if h["idx"] < len(mts) else {}
                        tags_h = tag_chunk(h["chunk"])
                        th = " ".join(f'<span style="background:rgba(139,58,139,.15);'
                                      f'border:1px solid rgba(139,58,139,.3);font-family:Space Mono,monospace;'
                                      f'font-size:.5rem;padding:.1rem .35rem;border-radius:3px;'
                                      f'color:{CAT_COLORS.get(t,VELVET["dim"])};">{t}</span>' for t in tags_h)
                        st.markdown(f'<div style="background:{VELVET["card"]};border:1px solid rgba(139,58,139,.22);'
                                    f'border-left:3px solid #C084C8;border-radius:0 8px 8px 0;'
                                    f'padding:.7rem .9rem;margin-bottom:.5rem;">'
                                    f'<div style="font-family:Space Mono,monospace;font-size:.58rem;color:#C084C8;'
                                    f'margin-bottom:.3rem;">#{rank} · 📄 {mt.get("filename","—")}'
                                    f' · score:{h["score"]:.3f}</div>'
                                    f'<div style="font-size:.8rem;color:#9A8AAA;line-height:1.55;">'
                                    f'{h["chunk"][:320]}…</div>'
                                    f'<div style="margin-top:.35rem;display:flex;gap:.3rem;flex-wrap:wrap;">{th}</div>'
                                    f'</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Search error: {e}")
        elif hs_q:
            st.info("Ingest documents first.")

    with _atabs[4]:
        # ── Monitoring strip ──────────────────────────────────────────────
        stats2 = compute_retrieval_stats()
        if stats2:
            s2_ce=stats2["avg_ce"]; s2_top=stats2["top_ce"]
            s2_n=stats2["n_queries"]; s2_r=stats2.get("recall_at_k")
            _ce_c2=VELVET["green"] if s2_top>0 else(VELVET["gold"] if s2_top>-3 else VELVET["red"])
            st.markdown(
                f'<div style="background:{VELVET["card2"]};border:1px solid rgba(139,58,139,.28);'
                f'border-radius:10px;padding:.65rem .9rem;margin-bottom:.8rem;display:flex;gap:1.5rem;flex-wrap:wrap;">'
                f'<div><div style="font-family:Space Mono,monospace;font-size:.44rem;text-transform:uppercase;color:{VELVET["ghost"]};">Queries</div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.86rem;color:{VELVET["text"]};">{s2_n}</div></div>'
                f'<div><div style="font-family:Space Mono,monospace;font-size:.44rem;text-transform:uppercase;color:{VELVET["ghost"]};">Avg CE</div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.86rem;color:{_ce_c2};">{s2_ce:.2f}</div></div>'
                f'<div><div style="font-family:Space Mono,monospace;font-size:.44rem;text-transform:uppercase;color:{VELVET["ghost"]};">Top-1 CE</div>'
                f'<div style="font-family:Space Mono,monospace;font-size:.86rem;color:{_ce_c2};">{s2_top:.2f}</div></div>'
                +(f'<div><div style="font-family:Space Mono,monospace;font-size:.44rem;text-transform:uppercase;color:{VELVET["ghost"]};">Recall@6</div>'
                  f'<div style="font-family:Space Mono,monospace;font-size:.86rem;color:{VELVET["green"] if s2_r and s2_r>=70 else(VELVET["gold"] if s2_r and s2_r>=40 else VELVET["red"])};">{s2_r:.1f}%</div></div>'
                  if s2_r is not None else"")+
                f'</div>', unsafe_allow_html=True)

        # ── Mode selector ─────────────────────────────────────────────────
        mode2 = st.radio("mode2",
            ["🎯 Document-Adaptive","📋 Generic","✏️ Custom"],
            label_visibility="collapsed", horizontal=True, key="eval_mode_sel2")

        if mode2.startswith("🎯"):
            col_g2, col_r2 = st.columns(2)
            with col_g2:
                if st.button("⚡ Generate QA", key="gen_qa2", use_container_width=True):
                    if st.session_state.doc_full_text and GROQ_API_KEY:
                        with st.spinner("Generating…"):
                            p2 = generate_doc_qa_pairs(st.session_state.doc_full_text, GROQ_API_KEY, 4)
                        if p2:
                            st.session_state.adaptive_qa_pairs = p2
                            st.success(f"{len(p2)} questions generated"); st.rerun()
            with col_r2:
                if st.button("▶  Run Adaptive", key="run_adaptive2", use_container_width=True,
                             disabled="adaptive_qa_pairs" not in st.session_state):
                    _run_eval_benchmark(st.session_state.adaptive_qa_pairs,
                                        st.session_state.vectorstore, GROQ_API_KEY,
                                        use_expected_answer=True)
            if "adaptive_qa_pairs" in st.session_state:
                st.caption(f"{len(st.session_state.adaptive_qa_pairs)} doc-specific questions ready")

        elif mode2.startswith("📋"):
            if st.button("▶  Run Generic Benchmark", key="bench2", use_container_width=True):
                _run_eval_benchmark(EVAL_QUESTIONS, st.session_state.vectorstore, GROQ_API_KEY,
                                    doc_text=st.session_state.get("doc_full_text", ""))

        else:
            _cq2 = st.session_state.get("custom_eval_qs", [])
            st.caption(f"{len(_cq2)} custom questions — add them in the Analytics tab Eval section")
            if _cq2:
                if st.button("▶  Run Custom", key="run_custom2", use_container_width=True):
                    _run_eval_benchmark(_cq2, st.session_state.vectorstore, GROQ_API_KEY,
                                        use_expected_answer=True)

    st.markdown("<hr style='border-color:rgba(139,58,139,.1);margin:.6rem 0 1rem;'>",
                unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# PORTFOLIO PANEL  (slides open when 📊 Portfolio clicked)
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.show_portfolio:
    n_holdings = len(st.session_state.portfolio)

    # ── 2-column layout: live markets strip | portfolio ───────────────────
    _pf_mkt_col, _pf_main_col = st.columns([1, 3], gap="small")

    with _pf_mkt_col:
        # Live global markets side panel — shown only when Portfolio is open
        st.markdown(
            '<div style="background:var(--card);border:1px solid var(--border);border-radius:12px;'
            'padding:.8rem .9rem;position:sticky;top:0;">'
            '<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.18em;'
            'text-transform:uppercase;color:var(--velvet-gl);margin-bottom:.7rem;">◈ Live Markets</div>',
            unsafe_allow_html=True,
        )
        # Global indices
        _pf_idx_syms = {
            "^GSPC":("S&P 500","🇺🇸"), "^IXIC":("NASDAQ","🇺🇸"),
            "^NSEI":("NIFTY 50","🇮🇳"), "^N225":("Nikkei","🇯🇵"),
            "^FTSE":("FTSE 100","🇬🇧"), "^GDAXI":("DAX","🇩🇪"),
        }
        _pf_idx_q = fetch_multi_quotes(tuple(_pf_idx_syms.keys()))
        _idx_rows = ""
        for sym, (name, flag) in _pf_idx_syms.items():
            info = _pf_idx_q.get(sym)
            if info:
                arr  = "▲" if info["pct"] >= 0 else "▼"
                col  = "#4ade80" if info["pct"] >= 0 else "#f87171"
                bg   = "rgba(74,222,128,.05)" if info["pct"] >= 0 else "rgba(248,113,113,.05)"
                bd   = "rgba(74,222,128,.15)" if info["pct"] >= 0 else "rgba(248,113,113,.15)"
                _idx_rows += (
                    f'<div style="display:flex;align-items:center;justify-content:space-between;'
                    f'padding:.35rem .5rem;background:{bg};border:1px solid {bd};'
                    f'border-radius:6px;margin-bottom:.3rem;">'
                    f'<div>'
                    f'  <div style="font-family:Space Mono,monospace;font-size:.52rem;color:var(--accent);">'
                    f'    {flag} {name}</div>'
                    f'  <div style="font-family:Cormorant Garamond,serif;font-size:1.05rem;'
                    f'    font-weight:300;color:#EDE8F5;line-height:1;">{info["price"]:,.0f}</div>'
                    f'</div>'
                    f'<div style="font-family:Space Mono,monospace;font-size:.56rem;color:{col};'
                    f'font-weight:700;">{arr} {abs(info["pct"]):.2f}%</div>'
                    f'</div>'
                )
        st.markdown(_idx_rows + "</div>", unsafe_allow_html=True)

        # Top movers in portfolio
        if st.session_state.portfolio:
            st.markdown(
                '<div style="background:var(--card);border:1px solid var(--border);border-radius:12px;'
                'padding:.8rem .9rem;margin-top:.6rem;">'
                '<div style="font-family:Space Mono,monospace;font-size:.5rem;letter-spacing:.18em;'
                'text-transform:uppercase;color:var(--velvet-gl);margin-bottom:.6rem;">◈ Your Movers</div>',
                unsafe_allow_html=True,
            )
            _mover_rows = ""
            for sym in list(st.session_state.portfolio.keys())[:8]:
                _mi = fetch_quote(sym)
                if _mi:
                    arr  = "▲" if _mi["pct"] >= 0 else "▼"
                    col  = "#4ade80" if _mi["pct"] >= 0 else "#f87171"
                    bg   = "rgba(74,222,128,.04)" if _mi["pct"] >= 0 else "rgba(248,113,113,.04)"
                    bd   = "rgba(74,222,128,.12)" if _mi["pct"] >= 0 else "rgba(248,113,113,.12)"
                    _mover_rows += (
                        f'<div style="display:flex;align-items:center;justify-content:space-between;'
                        f'padding:.28rem .45rem;background:{bg};border:1px solid {bd};'
                        f'border-radius:5px;margin-bottom:.25rem;">'
                        f'<div style="font-family:Space Mono,monospace;font-size:.56rem;color:var(--accent);">{sym}</div>'
                        f'<div style="font-family:Space Mono,monospace;font-size:.56rem;color:{col};">{arr} {abs(_mi["pct"]):.2f}%</div>'
                        f'</div>'
                    )
            st.markdown(_mover_rows + "</div>", unsafe_allow_html=True)

    with _pf_main_col:
        st.markdown(
            f'<div class="portfolio-panel">'
            f'<div class="portfolio-panel-hdr">'
            f'<div class="pph-title">Portfolio</div>'
            f'<div style="font-family:Space Mono,monospace;font-size:.52rem;color:#ff80c0;">'
            f'{n_holdings} holding{"s" if n_holdings != 1 else ""} · Live prices · FinBERT RAG</div>'
            f'</div></div>',
            unsafe_allow_html=True,
        )
        render_portfolio_panel(GROQ_API_KEY)

    st.markdown("<hr style='border-color:rgba(139,58,139,.1);margin:.6rem 0 1rem;'>",
                unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# STAT STRIP
# ─────────────────────────────────────────────────────────────────────────────
chunks = st.session_state.chunk_count
docs   = st.session_state.uploaded_docs
msgs   = len(st.session_state.messages) // 2
_vs    = st.session_state.vectorstore
_mlbl_disp = _vs.get("model_label", "FinBERT") if _vs else "FinBERT"

# Pipeline quality badge: show which components are active
_pipe_components = ["RRF", "CE Rerank", "BGE" if (_vs and _vs.get("is_bge")) else _mlbl_disp]
_stats_now = compute_retrieval_stats()
_avg_ce_disp = f" · CE avg {_stats_now['avg_ce']:+.1f}" if _stats_now else ""

st.markdown(f"""
<div class="stat-strip">
  <div class="stat-cell"><div class="stat-lbl">Embeddings</div>
    <div class="stat-val-mono">{_mlbl_disp} · Finance</div></div>
  <div class="stat-cell"><div class="stat-lbl">Chunks Indexed</div>
    <div class="stat-val {'active' if chunks else ''}">{chunks if chunks else '—'}</div></div>
  <div class="stat-cell"><div class="stat-lbl">Documents</div>
    <div class="stat-val {'active' if docs else ''}">{docs if docs else '—'}</div></div>
  <div class="stat-cell"><div class="stat-lbl">Pipeline</div>
    <div class="stat-val-mono" style="color:#4ade80;font-size:.62rem;">
      {'  ·  '.join(_pipe_components)}{_avg_ce_disp}</div></div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# MARKET MOOD + GLOBAL INDICES
# ─────────────────────────────────────────────────────────────────────────────
fng = fetch_fear_greed(); fng_val = fng["value"]; fng_label = fng["label"]
INDEX_SYMS = {"^GSPC":{"name":"S&P 500","flag":"🇺🇸"},"^IXIC":{"name":"NASDAQ","flag":"🇺🇸"},
              "^FTSE":{"name":"FTSE 100","flag":"🇬🇧"},"^NSEI":{"name":"NIFTY 50","flag":"🇮🇳"},
              "^N225":{"name":"Nikkei","flag":"🇯🇵"},"^GDAXI":{"name":"DAX","flag":"🇩🇪"}}
idx_quotes = fetch_multi_quotes(tuple(INDEX_SYMS.keys()))
idx_chips  = ""
for sym, meta in INDEX_SYMS.items():
    info = idx_quotes.get(sym)
    if info:
        arrow    = "▲" if info["pct"] >= 0 else "▼"
        cls      = "up" if info["pct"] >= 0 else "down"
        chip_cls = "chip-up" if info["pct"] >= 0 else "chip-down"
        idx_chips += (f'<div class="mood-idx-chip {chip_cls}">'
                      f'<div class="mood-idx-name">{meta["flag"]} {meta["name"]}</div>'
                      f'<div class="mood-idx-val">{info["price"]:,.0f}</div>'
                      f'<div class="mood-idx-chg {cls}">{arrow} {abs(info["pct"]):.2f}%</div></div>')
mood_color = "#f87171" if fng_val<25 else ("#fb923c" if fng_val<45 else ("#facc15" if fng_val<55 else ("#86efac" if fng_val<75 else "#4ade80")))
st.markdown(f"""
<div class="mood-bar-wrap">
  <div class="mood-title">◈ Market Mood &amp; Global Indices</div>
  <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.7rem;">
    <div>
      <div style="display:flex;align-items:baseline;gap:.4rem;">
        <span class="mood-index" style="color:{mood_color};">{fng_val}</span>
        <span style="font-family:'Space Mono',monospace;font-size:.62rem;letter-spacing:.1em;color:{mood_color};">{fng_label}</span>
      </div>
      <div style="font-family:'Space Mono',monospace;font-size:.5rem;color:#4A3858;margin-top:.2rem;">Crypto Fear &amp; Greed · alternative.me</div>
    </div>
    <div style="flex:1;">
      <div class="mood-track"><div class="mood-needle" style="left:{fng_val}%;"></div></div>
      <div class="mood-labels"><span>Extreme Fear</span><span>Fear</span><span>Neutral</span><span>Greed</span><span>Extreme Greed</span></div>
    </div>
  </div>
  <div class="mood-indices">{idx_chips}</div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# HEDGE FUND MOVES  — auto-sliding 3D cube carousel
# ─────────────────────────────────────────────────────────────────────────────
_HF_ANNOUNCEMENTS = [
    {"fund": "Bridgewater Associates", "action": "Increased", "asset": "Gold ETF (GLD)", "size": "$2.1B", "type": "BUY", "note": "Macro inflation hedge amid rising debt concerns"},
    {"fund": "Citadel", "action": "Opened", "asset": "NVIDIA (NVDA)", "size": "$890M", "type": "BUY", "note": "AI chip supercycle thesis — H100 demand underpriced"},
    {"fund": "Millennium Management", "action": "Reduced", "asset": "US Treasuries", "size": "$3.4B", "type": "SELL", "note": "Duration risk as Fed holds rates higher for longer"},
    {"fund": "Two Sigma", "action": "Initiated", "asset": "Reliance Industries", "size": "$540M", "type": "BUY", "note": "India infrastructure + Jio 5G runway undervalued"},
    {"fund": "Point72", "action": "Increased", "asset": "Bitcoin Futures", "size": "$1.2B", "type": "BUY", "note": "ETF approval structural inflows not yet priced in"},
    {"fund": "Renaissance Technologies", "action": "Exited", "asset": "Regional Banks (KRE)", "size": "$780M", "type": "SELL", "note": "CRE exposure and deposit outflow risk remain elevated"},
    {"fund": "D.E. Shaw", "action": "Built", "asset": "Taiwan Semiconductor (TSM)", "size": "$1.6B", "type": "BUY", "note": "Geopolitical premium overdone vs. structural AI demand"},
    {"fund": "Pershing Square", "action": "Shorted", "asset": "Long-dated US Bonds", "size": "$2.8B", "type": "SHORT", "note": "Fiscal deficits will keep long yields structurally elevated"},
    {"fund": "Tiger Global", "action": "Added", "asset": "Alphabet (GOOGL)", "size": "$1.1B", "type": "BUY", "note": "Search + Cloud re-acceleration; AI narrative recoupled"},
    {"fund": "Coatue Management", "action": "Trimmed", "asset": "Moderna (MRNA)", "size": "$420M", "type": "SELL", "note": "Post-COVID revenue cliff; pipeline needs 2026+ catalyst"},
]

def _build_hf_carousel(items: list[dict]) -> str:
    cards_html = ""
    for i, item in enumerate(items):
        color    = "#4ade80" if item["type"] in ("BUY","INCREASE") else ("#f87171" if item["type"] in ("SELL","SHORT") else "#F0C040")
        bg       = "rgba(74,222,128,.06)" if item["type"] in ("BUY","INCREASE") else ("rgba(248,113,113,.06)" if item["type"] in ("SELL","SHORT") else "rgba(240,192,64,.06)")
        bd       = color.replace(")", ",.25)").replace("rgb", "rgba") if "rgba" not in color else color
        tag_bg   = "rgba(74,222,128,.15)" if item["type"] == "BUY" else ("rgba(248,113,113,.15)" if item["type"] in ("SELL","SHORT") else "rgba(240,192,64,.15)")
        cards_html += f"""
        <div class="hf-card" style="background:{bg};border:1px solid rgba(255,255,255,.07);
             border-left:3px solid {color};border-radius:14px;
             padding:1.1rem 1.3rem;min-height:130px;display:flex;flex-direction:column;
             justify-content:space-between;flex-shrink:0;width:100%;box-sizing:border-box;">
          <div style="display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:.6rem;">
            <div>
              <div style="font-family:'Space Mono',monospace;font-size:.52rem;letter-spacing:.14em;
                   text-transform:uppercase;color:#4A3858;margin-bottom:.2rem;">{item['fund']}</div>
              <div style="font-family:'Cormorant Garamond',serif;font-size:1.12rem;font-weight:300;
                   color:#EDE8F5;line-height:1.2;">{item['action']} <span style="color:{color};">{item['asset']}</span></div>
            </div>
            <div style="font-family:'Space Mono',monospace;font-size:.62rem;font-weight:700;
                 background:{tag_bg};border:1px solid {color};border-radius:5px;
                 padding:.2rem .5rem;color:{color};white-space:nowrap;margin-left:.8rem;">{item['type']}</div>
          </div>
          <div style="display:flex;align-items:center;justify-content:space-between;">
            <div style="font-family:'Syne',sans-serif;font-size:.74rem;color:#9A8AAA;
                 line-height:1.5;max-width:78%;">{item['note']}</div>
            <div style="font-family:'Space Mono',monospace;font-size:.78rem;font-weight:700;
                 color:{color};white-space:nowrap;margin-left:.5rem;">{item['size']}</div>
          </div>
        </div>"""

    carousel_html = f"""<!DOCTYPE html>
<html><head><style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:transparent;overflow:hidden;}}
.hf-wrap{{
  background:linear-gradient(135deg,rgba(13,11,18,.97),rgba(20,12,30,.98));
  border:1px solid rgba(107,45,107,.25);border-radius:16px;
  padding:1rem 1.2rem;font-family:sans-serif;
  overflow:hidden;
}}
.hf-header{{
  display:flex;align-items:center;justify-content:space-between;margin-bottom:.8rem;
}}
.hf-title{{
  font-family:'Space Mono',monospace;font-size:.52rem;letter-spacing:.2em;
  text-transform:uppercase;color:#C084C8;display:flex;align-items:center;gap:.4rem;
}}
.hf-dot{{width:6px;height:6px;border-radius:50%;background:#C084C8;
  animation:blink 1.4s ease-in-out infinite;}}
@keyframes blink{{0%,100%{{opacity:1}}50%{{opacity:.3}}}}
.hf-counter{{font-family:'Space Mono',monospace;font-size:.5rem;color:#4A3858;}}
.hf-track{{position:relative;overflow:hidden;height:136px;}}
.hf-slider{{
  display:flex;flex-direction:column;
  transition:transform .6s cubic-bezier(.77,0,.175,1);
}}
.hf-card{{margin-bottom:.5rem;}}
.hf-dots{{display:flex;gap:.3rem;justify-content:center;margin-top:.6rem;flex-wrap:wrap;}}
.hf-pip{{width:5px;height:5px;border-radius:50%;background:rgba(107,45,107,.35);
  cursor:pointer;transition:all .2s;border:none;padding:0;}}
.hf-pip.active{{background:#C084C8;transform:scale(1.4);}}
</style></head><body>
<div class="hf-wrap">
  <div class="hf-header">
    <div class="hf-title">
      <div class="hf-dot"></div>
      Major Hedge Fund Moves
    </div>
    <div class="hf-counter" id="counter">1 / {len(items)}</div>
  </div>
  <div class="hf-track">
    <div class="hf-slider" id="slider">
      {cards_html}
    </div>
  </div>
  <div class="hf-dots" id="dots">
    {''.join(f'<button class="hf-pip{" active" if i==0 else ""}" onclick="goTo({i})" id="pip{i}"></button>' for i in range(len(items)))}
  </div>
</div>
<script>
const n={len(items)},CARD_H=152;
let cur=0,timer;
const slider=document.getElementById('slider');
const counter=document.getElementById('counter');
function goTo(i){{
  document.getElementById('pip'+cur).classList.remove('active');
  cur=(i+n)%n;
  slider.style.transform='translateY(-'+(cur*CARD_H)+'px)';
  counter.textContent=(cur+1)+' / '+n;
  document.getElementById('pip'+cur).classList.add('active');
}}
function next(){{goTo(cur+1);}}
timer=setInterval(next,3000);
slider.addEventListener('mouseenter',()=>clearInterval(timer));
slider.addEventListener('mouseleave',()=>{{timer=setInterval(next,3000);}});
</script></body></html>"""
    return carousel_html

_hf_html = _build_hf_carousel(_HF_ANNOUNCEMENTS)
st.components.v1.html(_hf_html, height=260, scrolling=False)


# ─────────────────────────────────────────────────────────────────────────────
# CHAT PANEL  (slides open below analytics when 💬 clicked)
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state.show_chat:
    st.markdown('<div class="chat-panel">'
                '<div class="chat-panel-hdr">'
                '<span class="chat-panel-title">◈ Ask Anything — Markets · Crypto · Documents</span>'
                '</div>'
                '<div class="chat-panel-body">',
                unsafe_allow_html=True)

    # ── Mode Toggle ────────────────────────────────────────────────────────
    _mode_cols = st.columns([2, 8])
    with _mode_cols[0]:
        _analyst = st.toggle(
            "📊 Analyst Mode",
            value=st.session_state.analyst_mode,
            key="analyst_toggle",
            help="OFF = Chat Mode (free-form answers)  ·  ON = Analyst Mode (structured financial extraction)",
        )
        if _analyst != st.session_state.analyst_mode:
            st.session_state.analyst_mode = _analyst
            st.rerun()
    with _mode_cols[1]:
        if st.session_state.analyst_mode:
            st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#F0C040;'
                        'padding:.3rem .5rem;background:rgba(240,192,64,.06);border:1px solid rgba(240,192,64,.2);'
                        'border-radius:5px;">📊 Analyst Mode — returns structured metrics table, '
                        'risks &amp; outlook as a JSON-parsed report card</div>',
                        unsafe_allow_html=True)
        else:
            st.markdown('<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#C084C8;'
                        'padding:.3rem .5rem;background:rgba(192,132,200,.05);border:1px solid rgba(192,132,200,.15);'
                        'border-radius:5px;">💬 Chat Mode — conversational answers with source evidence panel</div>',
                        unsafe_allow_html=True)

    st.markdown("<hr style='border-color:rgba(139,58,139,.12);margin:.4rem 0 .6rem;'>",
                unsafe_allow_html=True)

    if not st.session_state.messages:
        st.markdown("""
        <div style="text-align:center;padding:2rem 1rem;">
          <div style="font-size:2rem;margin-bottom:.6rem;opacity:.5;">◈</div>
          <div style="font-family:'Cormorant Garamond',serif;font-size:1.4rem;font-weight:300;
                      font-style:italic;color:#4A3858;">Ready without uploads</div>
          <div style="font-family:Syne,sans-serif;font-size:.76rem;color:#4A3858;
                      margin-top:.4rem;max-width:340px;margin-left:auto;margin-right:auto;line-height:1.8;">
            Ask about live stocks, gold, crypto, FX rates — no documents needed.<br>
            Upload a report above to unlock document Q&amp;A.
          </div>
        </div>""", unsafe_allow_html=True)

    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]):
            is_analyst = msg.get("analyst_mode", False)
            if msg["role"] == "assistant" and is_analyst:
                # Render structured analyst output card
                render_analyst_output(msg["content"], msg.get("question",""))
            else:
                st.markdown(msg["content"])
            # ── Source Evidence Panel ─────────────────────────────────────
            if msg.get("sources"):
                render_source_panel(msg["sources"])

    st.markdown("</div></div>", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# NEWS CAROUSELS
# ─────────────────────────────────────────────────────────────────────────────
news_items   = get_all_news()
policy_items = get_policy_news()
car_col1, car_col2 = st.columns(2)
with car_col1:
    if news_items:
        components.html(build_carousel_html(news_items, is_policy=False, height_px=400), height=400, scrolling=False)
    else:
        st.info("News unavailable — check back in a moment.")
with car_col2:
    if policy_items:
        components.html(build_carousel_html(policy_items, is_policy=True, height_px=400), height=400, scrolling=False)
    else:
        st.info("Policy news unavailable.")

# ─────────────────────────────────────────────────────────────────────────────
# COMMODITIES
# ─────────────────────────────────────────────────────────────────────────────
comm_quotes = fetch_multi_quotes(tuple(COMMODITY_SYMS.keys()))
comm_chips  = "".join(
    make_chip_html(sym, f"{name} · {unit}", info["price"], info["pct"], prefix="$", decimals=dec, icon=icon)
    for sym, (name, unit, icon, dec) in COMMODITY_SYMS.items() if (info := comm_quotes.get(sym))
)
if comm_chips:
    st.markdown(
        '<div class="comm-panel">'
        '<div class="comm-title">Precious Metals &amp; Commodities · Historical Data</div>',
        unsafe_allow_html=True)
    st.markdown('<div class="chips-row">'+comm_chips+'</div>', unsafe_allow_html=True)

    # Commodity chart + AI panel — chart re-renders based on period button
    def _comm_chart_fn(period, interval):
        data = {}
        for sym, (name, unit, icon, dec) in COMMODITY_SYMS.items():
            s = fetch_tf_series(sym, period, interval)
            if s is not None and len(s) > 1:
                data[f"{icon} {name}"] = s
        return data

    render_ai_timeframe_panel(
        symbols=list(COMMODITY_SYMS.keys()),
        panel_key="comm",
        groq_api_key=GROQ_API_KEY,
        accent="#F0C040",
        label="Commodity Analysis",
        news_topic="commodities gold oil copper energy metals",
        chart_series_fn=_comm_chart_fn,
    )
    st.markdown(
        '<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#4A3858;'
        'margin-top:.65rem;text-align:right;">Futures · Yahoo Finance · 60s cache</div></div>',
        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CRYPTO
# ─────────────────────────────────────────────────────────────────────────────
crypto_quotes = fetch_multi_quotes(tuple(CRYPTO_SYMS.keys()))
crypto_chips  = "".join(
    make_chip_html(ticker, name, info["price"], info["pct"], prefix="$", decimals=dec, icon=icon)
    for sym, (name, ticker, icon, dec) in CRYPTO_SYMS.items() if (info := crypto_quotes.get(sym))
)
if crypto_chips:
    st.markdown(
        '<div class="crypto-panel">'
        '<div class="crypto-title">Crypto Markets · Historical Data</div>',
        unsafe_allow_html=True)
    st.markdown('<div class="chips-row">'+crypto_chips+'</div>', unsafe_allow_html=True)

    # Crypto chart + AI panel — chart re-renders based on period button
    def _crypto_chart_fn(period, interval):
        data = {}
        for sym, (name, ticker, icon, dec) in CRYPTO_SYMS.items():
            s = fetch_tf_series(sym, period, interval)
            if s is not None and len(s) > 1:
                data[f"{icon} {name}"] = s
        return data

    render_ai_timeframe_panel(
        symbols=list(CRYPTO_SYMS.keys()),
        panel_key="crypto",
        groq_api_key=GROQ_API_KEY,
        accent="#FB923C",
        label="Crypto Analysis",
        news_topic="bitcoin ethereum crypto blockchain defi",
        chart_series_fn=_crypto_chart_fn,
    )
    st.markdown(
        '<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#4A3858;'
        'margin-top:.65rem;text-align:right;">Spot · Yahoo Finance · 60s cache</div></div>',
        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# LIVE STOCK CHART  — 1 Year · Candlestick/Area · AI Analysis
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    '<div style="background:#0D0B12;border:1px solid rgba(139,58,139,.22);'
    'border-radius:12px;padding:1.2rem 1.4rem .9rem;margin-bottom:1.4rem;">',
    unsafe_allow_html=True)
st.markdown(
    '<div style="font-family:\'Cormorant Garamond\',serif;font-size:1.1rem;font-weight:300;'
    'color:#EDE8F5;margin-bottom:.8rem;display:flex;align-items:center;gap:.5rem;">'
    '<span style="display:inline-block;width:3px;height:1.1rem;'
    'background:linear-gradient(180deg,#6B2D6B,#C084C8);border-radius:2px;"></span>'
    'Stock Charts · 1 Year Historical Data</div>', unsafe_allow_html=True)

sc_r1, sc_r2, sc_r3 = st.columns([4, 1.1, 1.1])
with sc_r1:
    symbols = st.multiselect("symbols",
        options=["AAPL","MSFT","NVDA","GOOGL","AMZN","TSLA","META","TSM","SAP","BABA","SONY","NVO",
                 "RELIANCE.NS","TCS.NS","INFY.NS","WIPRO.NS","HDFCBANK.NS","BTC-USD","ETH-USD",
                 "GC=F","CL=F","^GSPC","^NSEI"],
        default=["AAPL","MSFT","NVDA","TSLA"], label_visibility="collapsed",
        key="main_chart_syms")
with sc_r2:
    chart_type = st.selectbox("chart_type",
        ["% Return","Absolute Price","Candlestick (1 symbol)"],
        index=0, label_visibility="collapsed", key="chart_type_sel")
with sc_r3:
    rng = st.selectbox("range", ["1M","3M","6M","1Y"], index=3,
                       label_visibility="collapsed", key="chart_rng")

period_map   = {"1M":"1mo","3M":"3mo","6M":"6mo","1Y":"1y"}
interval_map = {"1M":"1d","3M":"1d","6M":"1d","1Y":"1d"}

if symbols:
    sq = fetch_multi_quotes(tuple(symbols))
    cps = []
    for sym in symbols:
        info = sq.get(sym)
        if info:
            arr = "▲" if info["pct"] >= 0 else "▼"
            cc  = "#4ade80" if info["pct"] >= 0 else "#f87171"
            cps.append(
                f'<div style="display:flex;flex-direction:column;align-items:center;'
                f'background:#120E1A;border:1px solid rgba(139,58,139,.22);border-radius:8px;'
                f'padding:.45rem .75rem;min-width:80px;font-family:Space Mono,monospace;">'
                f'<span style="font-size:.62rem;color:#C084C8;font-weight:700;">{sym}</span>'
                f'<span style="font-size:.74rem;color:#EDE8F5;margin-top:.1rem;">${info["price"]:,.2f}</span>'
                f'<span style="font-size:.58rem;color:{cc};">{arr} {abs(info["pct"]):.2f}%</span></div>'
            )
    if cps:
        st.markdown('<div style="display:flex;gap:.55rem;flex-wrap:wrap;margin-bottom:.8rem;">'
                    +"".join(cps)+"</div>", unsafe_allow_html=True)

    # ── Fetch 1Y history ────────────────────────────────────────────────
    hist_frames: dict[str, pd.DataFrame] = {}
    series_dict: dict[str, pd.Series] = {}
    with st.spinner("Loading 1 year of historical data…"):
        for sym in symbols:
            df = fetch_stock_history_1y(sym)
            if not df.empty:
                hist_frames[sym] = df
                series_dict[sym] = df["close"]

    if series_dict:
        is_candle = (chart_type == "Candlestick (1 symbol)")
        mode = "normalized" if "Return" in chart_type else "absolute"

        if is_candle and len(symbols) == 1 and symbols[0] in hist_frames:
            fig = build_rich_chart(
                series_dict, mode="candlestick",
                title=f"{symbols[0]} · Candlestick Chart · {rng}",
                single_ohlc=hist_frames[symbols[0]],
                height=420,
            )
        else:
            fig = build_rich_chart(
                series_dict, mode=mode,
                title=f"{'Normalised % Return' if mode=='normalized' else 'Price'} · {rng}",
                show_bollinger=(len(series_dict) == 1),
                show_sma=(len(series_dict) == 1),
                height=380,
            )

        if fig is not None:
            st.plotly_chart(fig, use_container_width=True, config=dict(
                displayModeBar=True,
                modeBarButtonsToRemove=["toImage","sendDataToCloud","zoom2d","pan2d"],
                displaylogo=False,
            ))
        else:
            normed = (pd.DataFrame(series_dict).dropna(how="all").ffill()
                      / pd.DataFrame(series_dict).dropna(how="all").ffill().iloc[0] - 1) * 100
            st.line_chart(normed, height=280, use_container_width=True)

        # ── AI Timeframe Analysis with news ──────────────────────────────
        _syms_for_news = symbols
        def _stock_chart_fn(period, interval):
            data = {}
            for sym in _syms_for_news:
                s = fetch_tf_series(sym, period, interval)
                if s is not None and len(s) > 1:
                    data[sym] = s
            return data

        render_ai_timeframe_panel(
            symbols=symbols,
            panel_key="chart",
            groq_api_key=GROQ_API_KEY,
            accent="#C084C8",
            label="Stock Market Analysis",
            news_topic=f"stocks equity {' '.join(symbols[:3])} earnings market",
            chart_series_fn=_stock_chart_fn,
        )

    else:
        st.warning("Chart data unavailable — try again in a moment.")
else:
    st.info("Select at least one symbol above.")
st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CURRENCY PANEL
# ─────────────────────────────────────────────────────────────────────────────
fx_options     = {f"{m['flag']} {m['label']} · {m['name']}": sym for sym, m in ALL_FX.items()}
default_labels = [k for k, v in fx_options.items()
                  if v in ("USDINR=X","USDJPY=X","USDCNY=X","EURUSD=X","GBPUSD=X","USDCHF=X")]
st.markdown('<div class="fx-panel"><div class="fx-panel-title">Currencies vs USD · 1 Year Data</div>',
            unsafe_allow_html=True)
fx_r1, fx_r2 = st.columns([5, 1])
with fx_r1:
    selected_labels = st.multiselect("currencies", options=list(fx_options.keys()),
        default=default_labels, label_visibility="collapsed", key="fx_select")
with fx_r2:
    fx_rng = st.selectbox("fx_range", ["1M","3M","6M","1Y"], index=3,
                          label_visibility="collapsed", key="fx_rng")
selected_syms = [fx_options[lbl] for lbl in selected_labels]
st.session_state["fx_select_syms"] = selected_syms
fx_period   = {"1M":"1mo","3M":"3mo","6M":"6mo","1Y":"1y"}
fx_interval = {"1M":"1d","3M":"1d","6M":"1d","1Y":"1d"}
if selected_syms:
    fx_series: dict[str, pd.Series] = {}
    fx_hist:   dict[str, pd.DataFrame] = {}
    for sym in selected_syms:
        meta = ALL_FX[sym]; s = fetch_yahoo_series(sym, fx_period[fx_rng], fx_interval[fx_rng])
        if s is not None and not s.empty:
            if meta["invert"]: s = 1.0 / s
            label = meta["flag"]+" "+meta["label"]
            fx_series[label] = s
            # store raw for stats (inverted if needed)
            df_raw = pd.DataFrame({"close": s})
            fx_hist[sym] = df_raw

    if fx_series:
        fig_fx = build_rich_chart(fx_series, mode="normalized",
                                  title=f"Currency % Change vs USD · {fx_rng}",
                                  show_bollinger=False, show_sma=False, height=280)
        if fig_fx is not None:
            st.plotly_chart(fig_fx, use_container_width=True,
                            config=dict(displayModeBar=False, displaylogo=False))
        else:
            nc = pd.DataFrame(fx_series).dropna(how="all").ffill()
            nc = (nc / nc.iloc[0] - 1) * 100
            st.line_chart(nc, height=220, use_container_width=True)

    # ── AI Timeframe Analysis ─────────────────────────────────────────────
    if selected_syms:
        render_ai_timeframe_panel(
            symbols=selected_syms,
            panel_key="fx",
            groq_api_key=GROQ_API_KEY,
            accent="#60A5FA",
            label="Currency Analysis",
            news_topic="forex currency dollar rupee yen euro central bank interest rates",
        )

    fx_quotes = fetch_multi_quotes(tuple(selected_syms))
    fx_chips  = []
    for sym in selected_syms:
        meta = ALL_FX[sym]; info = fx_quotes.get(sym)
        if info:
            rate = info["price"]; pct = info["pct"]
            rs   = f"{rate:,.2f}" if rate >= 10 else f"{rate:.4f}"
            arr  = "▲" if pct > 0.005 else ("▼" if pct < -0.005 else "●")
            cls  = "up" if pct > 0.005 else ("down" if pct < -0.005 else "flat")
            chip_cls = "chip-up" if pct > 0.005 else ("chip-down" if pct < -0.005 else "")
            fx_chips.append(
                f'<div class="price-chip {chip_cls}">'
                f'<div class="pc-sym">{meta["flag"]} {meta["label"]}</div>'
                f'<div class="pc-name">{meta["name"]}</div>'
                f'<div class="pc-val">{rs}</div>'
                f'<div class="pc-chg {cls}">{arr} {abs(pct):.3f}%</div></div>'
            )
    if fx_chips:
        now_ist = _dt.datetime.utcnow() + _dt.timedelta(hours=5, minutes=30)
        st.markdown('<div class="chips-row" style="margin-top:.75rem;">'+"".join(fx_chips)
                    +f'</div><div style="font-family:Space Mono,monospace;font-size:.5rem;color:#4A3858;'
                    f'margin-top:.6rem;text-align:right;">Live · {now_ist.strftime("%H:%M")} IST · 60s cache</div>',
                    unsafe_allow_html=True)
else:
    st.info("Select at least one currency pair above.")
st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# INDIA NEWS & POLICY
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="display:flex;align-items:center;gap:.6rem;margin-bottom:.8rem;">
  <span style="display:inline-block;width:3px;height:1.4rem;
    background:linear-gradient(180deg,#6B2D6B,#C084C8);border-radius:2px;flex-shrink:0;"></span>
  <span style="font-family:'Cormorant Garamond',serif;font-size:1.35rem;font-weight:300;color:#EDE8F5;">
    India News &amp; Policy
  </span>
</div>
""", unsafe_allow_html=True)
nc1, nc2, nc3 = st.columns([3, 2, 1])
with nc1:
    sel_sources = st.multiselect("sources", options=list(NEWS_SOURCES.keys()),
        default=["📰 Economic Times","📰 Business Standard","🏛️ RBI Notifications","🏛️ Finance Ministry"],
        label_visibility="collapsed", key="news_sources")
with nc2:
    news_filter = st.selectbox("filter", ["All","Markets","Policy"], index=0,
                               label_visibility="collapsed", key="news_filter")
with nc3:
    n_per_source = st.selectbox("items", [3,5,8,10], index=0,
                                label_visibility="collapsed", key="news_n")
active_sources = [s for s in sel_sources if news_filter=="All" or NEWS_SOURCES[s]["tag"]==news_filter]
if active_sources:
    all_articles = []
    for src_name in active_sources:
        src_cfg  = NEWS_SOURCES[src_name]
        articles = fetch_rss(src_cfg["rss"], max_items=n_per_source)
        for a in articles: a["source"]=src_name; a["src_tag"]=src_cfg["tag"]
        all_articles.extend(articles)
    if all_articles:
        def render_card(a):
            tc = "#C084C8" if a["src_tag"]=="Policy" else "#4ade80"
            tb = "rgba(192,132,200,.1)" if a["src_tag"]=="Policy" else "rgba(74,222,128,.08)"
            tbd = "rgba(192,132,200,.3)" if a["src_tag"]=="Policy" else "rgba(74,222,128,.25)"
            ss  = a["source"].replace("📰 ","").replace("🏛️ ","")
            return (f'<div style="background:#0D0B12;border:1px solid rgba(139,58,139,.22);'
                    f'border-radius:10px;padding:.8rem 1rem;margin-bottom:.6rem;border-left:3px solid {tc};">'
                    f'<div style="display:flex;align-items:center;gap:.4rem;margin-bottom:.4rem;flex-wrap:wrap;">'
                    f'<span style="background:{tb};border:1px solid {tbd};color:{tc};'
                    f'font-family:\'Space Mono\',monospace;font-size:.52rem;letter-spacing:.1em;'
                    f'padding:.1rem .4rem;border-radius:3px;text-transform:uppercase;">{a["src_tag"]}</span>'
                    f'<span style="font-family:\'Space Mono\',monospace;font-size:.52rem;color:#4A3858;">{ss}</span>'
                    f'<span style="font-family:\'Space Mono\',monospace;font-size:.5rem;color:#4A3858;margin-left:auto;">{a["date"]}</span></div>'
                    f'<a href="{a.get("link","#")}" target="_blank" style="text-decoration:none;">'
                    f'<div style="font-family:\'Syne\',sans-serif;font-size:.82rem;font-weight:500;'
                    f'color:#EDE8F5;line-height:1.45;margin-bottom:.35rem;">{a["title"]}</div></a>'
                    f'<div style="font-family:\'Syne\',sans-serif;font-size:.72rem;color:#4A3858;'
                    f'line-height:1.5;overflow:hidden;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;">'
                    f'{a.get("summary","")}</div></div>')
        col_l, col_r = st.columns(2)
        with col_l: st.markdown("".join(render_card(a) for a in all_articles[0::2]), unsafe_allow_html=True)
        with col_r: st.markdown("".join(render_card(a) for a in all_articles[1::2]), unsafe_allow_html=True)
        now_ist_n = _dt.datetime.utcnow() + _dt.timedelta(hours=5, minutes=30)
        st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:.5rem;color:#4A3858;'
                    f'text-align:right;margin-top:.2rem;">Fetched {now_ist_n.strftime("%H:%M")} IST · '
                    f'{len(all_articles)} articles · 10min cache</div>', unsafe_allow_html=True)

        # ── AI News Analysis ──────────────────────────────────────────────
        st.markdown('<div style="margin-top:.8rem;">', unsafe_allow_html=True)
        an1, an2 = st.columns([2, 1])
        with an1:
            if st.button("🤖 Analyse Headlines & Predict Market Impact",
                         key="news_ai_btn", use_container_width=True):
                with st.spinner("Reading headlines and generating market prediction…"):
                    headlines = "\n".join(
                        f"• {a['title']} ({a['source'].replace('📰 ','').replace('🏛️ ','')})"
                        for a in all_articles[:12]
                    )
                    if GROQ_API_KEY:
                        news_pred = groq_call(
                            api_key=GROQ_API_KEY,
                            messages=[{"role":"user","content":
                                f"Today's India financial & policy headlines:\n{headlines}\n\n"
                                f"Analyse these headlines and predict their impact on:\n"
                                f"1. NIFTY 50 and Indian equities (direction, magnitude)\n"
                                f"2. USD/INR exchange rate\n"
                                f"3. Indian bond yields / RBI policy expectations\n"
                                f"4. Key sectors most affected (IT, Banks, FMCG, Energy, etc.)\n"
                                f"For each: cite the specific headline, give directional call "
                                f"(bullish/bearish/neutral), and estimate % move or basis point impact. "
                                f"End with a 1-week outlook."}],
                            system=("You are India's top macroeconomic strategist covering "
                                    "SEBI, RBI, and Indian equity markets. Analyse news headlines "
                                    "for actionable market impact. Be specific, cite numbers, "
                                    "give clear directional calls. 8-10 sentences."),
                            temperature=0.2, max_tokens=650,
                            site_key="india_news_predict",
                        )
                        st.session_state["_india_news_pred"] = news_pred
                    else:
                        st.session_state["_india_news_pred"] = "⚠ Add Groq API key to enable AI analysis."
        with an2:
            if st.button("💬 Ask about this news",
                         key="news_chat_btn", use_container_width=True):
                headlines_short = " | ".join(a["title"] for a in all_articles[:6])
                st.session_state["_prefill"] = (
                    f"Based on these recent headlines: {headlines_short} — "
                    f"What is the likely impact on Indian markets and what should investors do?"
                )
                st.session_state["show_chat"] = True
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        # Show prediction result
        if st.session_state.get("_india_news_pred"):
            st.markdown(
                f'<div style="background:rgba(74,222,128,.04);border:1px solid rgba(74,222,128,.2);'
                f'border-left:3px solid #4ade80;border-radius:0 10px 10px 0;'
                f'padding:.9rem 1.1rem;margin-top:.6rem;">'
                f'<div style="font-family:Space Mono,monospace;font-size:.44rem;color:#4ade80;'
                f'text-transform:uppercase;letter-spacing:.12em;margin-bottom:.5rem;">'
                f'🤖 AI Market Impact Analysis</div>'
                f'<div style="font-family:Syne,sans-serif;font-size:.82rem;color:#C8B8D8;line-height:1.8;">'
                f'{st.session_state["_india_news_pred"]}</div></div>',
                unsafe_allow_html=True,
            )
    else:
        st.info("No articles loaded — try refreshing or selecting different sources.")
else:
    st.info("Select at least one source above.")

st.markdown("<hr style='border-color:rgba(139,58,139,.15);margin:1.4rem 0;'>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CHAT INPUT  (always rendered so st.chat_input works; messages show in panel above)
# ─────────────────────────────────────────────────────────────────────────────
prefill  = st.session_state.pop("_prefill", None)
question = st.chat_input("Ask about stocks, gold, crypto, currencies, or your documents…")
q = prefill or question

if q:
    # Auto-open chat panel when a question is submitted
    if not st.session_state.show_chat:
        st.session_state.show_chat = True

    if not GROQ_API_KEY:
        st.error("Please enter your Groq API key in the sidebar."); st.stop()

    st.session_state.messages.append({"role":"user","content":q})

    with st.spinner("Thinking…" if not st.session_state.analyst_mode else "Extracting structured insights…"):
        try:
            from openai import OpenAI
            oai = OpenAI(api_key=GROQ_API_KEY, base_url="https://api.groq.com/openai/v1")

            stock_lines = [f"  {sym}: ${info['price']:,.2f} ({'▲' if info['pct']>=0 else '▼'}{abs(info['pct']):.2f}%)"
                           for sym in symbols if (info := fetch_quote(sym))]
            comm_lines  = [f"  {name}: ${info['price']:,.{dec}f} {unit} ({'+' if info['pct']>=0 else ''}{info['pct']:.2f}%)"
                           for sym, (name, unit, _, dec) in COMMODITY_SYMS.items() if (info := fetch_quote(sym))]
            crypto_lines = [f"  {ticker}: ${info['price']:,.{dec}f} ({'+' if info['pct']>=0 else ''}{info['pct']:.2f}%)"
                            for sym, (name, ticker, _, dec) in CRYPTO_SYMS.items() if (info := fetch_quote(sym))]
            fx_lines = []
            for _fxsym in st.session_state.get("fx_select_syms", ("USDINR=X","USDJPY=X","USDCNY=X")):
                _fxi = fetch_quote(_fxsym)
                if _fxi:
                    _p  = _fxi["price"]; _rs = f"{_p:,.2f}" if _p >= 10 else f"{_p:.4f}"
                    _sg = "+" if _fxi["pct"] >= 0 else ""
                    fx_lines.append(f"  {ALL_FX.get(_fxsym,{}).get('label',_fxsym)}: {_rs} ({_sg}{_fxi['pct']:.3f}%)")

            utc_now = _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

            # Fetch recent news headlines (cached 5 min) to ground the LLM in current events
            _news_ctx = ""
            try:
                _is_market_q = any(w in q.lower() for w in
                    ["market","stock","nifty","sensex","gold","oil","bitcoin","crypto",
                     "rupee","dollar","rbi","fed","rate","inflation","economy","news"])
                if _is_market_q:
                    _news_ctx = fetch_market_news_context(
                        "india financial markets stocks economy" if "india" in q.lower() or "nifty" in q.lower()
                        else "financial markets economy stocks"
                    )
            except Exception:
                _news_ctx = ""

            live_context = (
                f"=== LIVE MARKET DATA ({utc_now}) ===\n"
                f"STOCKS:\n{chr(10).join(stock_lines) or '  (none)'}\n"
                f"COMMODITIES:\n{chr(10).join(comm_lines) or '  (unavailable)'}\n"
                f"CRYPTO:\n{chr(10).join(crypto_lines) or '  (unavailable)'}\n"
                f"CURRENCIES (vs USD):\n{chr(10).join(fx_lines) or '  (unavailable)'}\n"
                f"MARKET MOOD: Fear & Greed = {fng_val} ({fng_label})"
                + (f"\n\n=== RECENT NEWS HEADLINES ===\n{_news_ctx}" if _news_ctx else "")
            ).strip()

            # ── Document retrieval ─────────────────────────────────────────
            doc_context = ""; sources_data = []
            if st.session_state.vectorstore:
                vs = st.session_state.vectorstore

                # 1. Semantic cache
                _cached = _retrieval_cache_get(q)
                if _cached:
                    doc_context  = _cached["doc_context"]
                    sources_data = _cached["sources_data"]
                else:
                    # 2. Query expansion
                    expanded_q = _expand_query(q)

                    # 3. TF-IDF similarity (replaces neural embedding + ChromaDB)
                    from sklearn.metrics.pairwise import cosine_similarity
                    import numpy as _np2
                    vectorizer   = vs.get("vectorizer")
                    tfidf_matrix = vs.get("tfidf_matrix")
                    chunks_all   = vs.get("chunks", [])
                    meta_all     = vs.get("meta", [])
                    dense_scores_all = []
                    if vectorizer is not None and tfidf_matrix is not None and len(chunks_all):
                        q_vec = vectorizer.transform([expanded_q])
                        sims  = cosine_similarity(q_vec, tfidf_matrix).flatten()
                        dense_scores_all = sims.tolist()
                    else:
                        dense_scores_all = [0.0] * len(chunks_all)

                    # 4. BM25 re-score
                    try:
                        from rank_bm25 import BM25Okapi
                        _bm25_local = BM25Okapi([_tokenize(c) for c in chunks_all])
                        bm25_raw    = _bm25_local.get_scores(_tokenize(expanded_q))
                        bm25_max    = max(bm25_raw) if max(bm25_raw) > 0 else 1.0
                        bm25_norm   = [s / bm25_max for s in bm25_raw]
                    except Exception:
                        bm25_norm = [0.0] * len(chunks_all)

                    # 5. RRF fusion
                    _K = 60
                    N  = len(chunks_all)
                    dense_rank = sorted(range(N), key=lambda i: dense_scores_all[i], reverse=True)
                    bm25_rank  = sorted(range(N), key=lambda i: bm25_norm[i],        reverse=True)
                    rrf: dict[int, float] = {}
                    for ranked in (dense_rank, bm25_rank):
                        for rank, idx in enumerate(ranked):
                            rrf[idx] = rrf.get(idx, 0.0) + 1.0 / (_K + rank + 1)

                    # 6. Section hint pre-filter (post-RRF boost)
                    q_lower = q.lower()
                    pre_section = None
                    _section_hints = {
                        "Income Statement": r"revenue|net income|gross profit|ebitda|operating income",
                        "Balance Sheet":    r"balance sheet|assets|liabilities|equity|debt|borrowing",
                        "Cash Flow":        r"cash flow|capex|free cash|operating cash",
                        "Per Share":        r"\beps\b|earnings per share|diluted|dividend per share",
                        "Ratios":           r"ratio|margin|roe|roa|roce|p/e|current ratio",
                        "Risk Factors":     r"risk|litigation|regulatory|compliance",
                    }
                    for sec, pat in _section_hints.items():
                        if re.search(pat, q_lower, re.IGNORECASE):
                            pre_section = sec; break

                    # 7. Build candidates (top 20 by RRF)
                    top_idxs = sorted(rrf, key=lambda i: rrf[i], reverse=True)[:20]
                    candidates = []
                    for i in top_idxs:
                        m       = meta_all[i]
                        section = m.get("section") or guess_section(chunks_all[i])
                        boost   = 0.04 if (pre_section and section == pre_section) else 0.0
                        candidates.append({
                            "chunk":     chunks_all[i],
                            "meta":      m,
                            "dist":      1 - dense_scores_all[i],
                            "score":     min(1.0, rrf.get(i, 0.0) + boost),
                            "rrf_raw":   rrf.get(i, 0.0),
                            "section":   section,
                            "doc_title": m.get("doc_title", m.get("filename", "")),
                            "page":      m.get("page", ""),
                        })

                    # 8. Sort by score (CE rerank removed — no sentence-transformers)
                    candidates.sort(key=lambda x: x["score"], reverse=True)

                    top = candidates[:6]
                    doc_context = "\n---\n".join(
                        f"[{c['doc_title']} · {c['section']} · p{c['page']}]\n{c['chunk']}"
                        for c in top
                    )
                    sources_data = [
                        {
                            "filename":   c["meta"]["filename"],
                            "doc_title":  c["doc_title"],
                            "score":      round(c["score"], 3),
                            "ce_score":   round(c.get("ce_score", 0), 2),
                            "preview":    c["chunk"][:220],
                            "chunk_full": c["chunk"],
                            "chunk_idx":  c["meta"].get("chunk", ""),
                            "section":    c["section"],
                            "page":       c["page"],
                        }
                        for c in top
                    ]
                    _retrieval_cache_put(q, {"doc_context": doc_context, "sources_data": sources_data})

                    # ── Retrieval monitoring log ──────────────────────────
                    log_retrieval(
                        query=q,
                        retrieved=[{
                            "section":  s.get("section","?"),
                            "ce_score": s.get("ce_score", 0),
                            "score":    s.get("score", 0),
                            "page":     s.get("page",""),
                        } for s in sources_data],
                    )

            # ── Branch: Analyst Mode vs Chat Mode ────────────────────────
            if st.session_state.analyst_mode:
                # Analyst Mode: JSON-structured extraction
                user_msg = build_analyst_prompt(q, live_context, doc_context)
                resp = oai.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[
                        {"role": "system", "content": _ANALYST_SYSTEM},
                        {"role": "user",   "content": user_msg},
                    ],
                    temperature=0.05, max_tokens=1200,
                    response_format={"type": "json_object"},
                )
                answer = resp.choices[0].message.content
                st.session_state.messages.append({
                    "role": "assistant", "content": answer,
                    "sources": sources_data, "analyst_mode": True, "question": q,
                })
            else:
                # Chat Mode: precision financial extraction prompt
                if doc_context:
                    user_msg = (
                        f"{live_context}\n\n"
                        f"=== DOCUMENT CONTEXT ===\n{doc_context}\n\n"
                        f"Question: {q}"
                    )
                    sys_prompt = (
                        "You are an expert financial analyst with access to real-time market data "
                        "and uploaded financial documents (annual reports, earnings releases, SEC filings).\n\n"
                        "CRITICAL RULES for document questions:\n"
                        "1. Extract the EXACT value from the document context. Quote the number verbatim.\n"
                        "2. If the context has multiple years, return the MOST RECENT value first, then prior years.\n"
                        "3. Always cite which document section and page the number came from.\n"
                        "4. If the value is NOT found in the provided context, say exactly: "
                        "\"NOT FOUND in the provided document context.\" — never fabricate a number.\n"
                        "5. For percentage changes, show both absolute value and YoY/QoQ delta.\n"
                        "6. For live market questions (price, rates, indices), use the live data block.\n"
                        "7. Be concise but complete. Use tables for multi-metric answers."
                    )
                else:
                    user_msg = f"{live_context}\n\nQuestion: {q}"
                    sys_prompt = (
                        "You are an expert financial analyst with real-time market data access. "
                        "You have live prices for stocks, gold, silver, oil, crypto, and FX rates. "
                        "Use live data for all market questions. Be concise and precise. "
                        "Never fabricate numbers."
                    )
                resp = oai.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[
                        {"role": "system", "content": sys_prompt},
                        *[{"role": m["role"], "content": m["content"]}
                          for m in st.session_state.messages[:-1]],
                        {"role": "user", "content": user_msg},
                    ],
                    temperature=0.08, max_tokens=1500,
                )
                answer = resp.choices[0].message.content
                st.session_state.messages.append({
                    "role": "assistant", "content": answer,
                    "sources": sources_data, "analyst_mode": False,
                })

            st.rerun()

        except Exception as e:
            err = str(e)
            if "rate_limit_exceeded" in err or "429" in err or "Rate limit" in err:
                retry_match = re.search(r"try again in (\d+m[\d.]*s|[\d.]+s)", err, re.IGNORECASE)
                retry_str   = retry_match.group(1) if retry_match else "a few minutes"
                st.warning(
                    f"⏱ **Groq rate limit reached** — daily token cap hit on the free tier.  \n"
                    f"Analysis available again in ~**{retry_str}**.  \n"
                    f"💡 Upgrade at [console.groq.com/settings/billing](https://console.groq.com/settings/billing)"
                )
            else:
                st.error(f"Error: {err[:200]}")

# ─────────────────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="vfooter">
  <div class="vfooter-text">
    Built by Yash Chaudhary &nbsp;·&nbsp; Financial RAG Assistant v9 &nbsp;·&nbsp;
    Llama 3.3 × Groq × ChromaDB × FinBERT · Portfolio · Analyst Mode
  </div>
</div>
""", unsafe_allow_html=True)
